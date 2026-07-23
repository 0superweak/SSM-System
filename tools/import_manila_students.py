"""Idempotent Manila workbook migration into the SSM Supabase database.

The default mode is a read-only dry run. Pass ``--commit`` only after reviewing
the generated plan. Existing non-empty student fields are preserved, source
duplicates are collapsed to their newest/best workbook, and expense inserts
are deduplicated. Embedded document scans are never treated as profile photos.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import sys
import unicodedata
import zipfile
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from office_app.repositories.audit_repository import AuditRepository
from office_app.services.supabase_client import get_supabase


SUMMARY_FILES = {
    "college ministry master file.xlsx",
    "with honor students.xlsx",
}
STUDENT_COLUMNS = (
    "id,last_name,first_name,gender,grade,address,city,area,birthday,"
    "sponsor,contact,school,parents,course,remarks,status,photo_url"
)
PROFILE_LABELS = {
    "studentsname": "name",
    "studentname": "name",
    "gender": "gender",
    "address": "address",
    "level": "grade",
    "grade": "grade",
    "course": "course",
    "school": "school",
    "dob": "birthday",
    "birthday": "birthday",
    "contactno": "contact",
    "contactnumber": "contact",
    "emailfb": "email",
    "email": "email",
    "fathersname": "father",
    "mothersname": "mother",
    "parentsname": "parents",
    "sponsorsname": "sponsor",
    "sponsorname": "sponsor",
    "coordinator": "coordinator",
    "studentsupdate": "student_update",
    "studentupdate": "student_update",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def normalized_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value).casefold())
    return re.sub(r"[^a-z0-9]+", "", text)


def identity_key(last_name: Any, first_name: Any) -> str:
    return f"{normalized_text(last_name)}|{normalized_text(first_name)}"


def first_value_to_right(sheet, row: int, start_column: int = 2) -> Any:
    for column in range(start_column, min(sheet.max_column, 16) + 1):
        value = sheet.cell(row, column).value
        if clean_text(value):
            return value
    return None


def parse_date(value: Any, *, epoch=None) -> Optional[dt.date]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 90000:
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date() if isinstance(converted, dt.datetime) else converted
        except Exception:
            return None
    text = clean_text(value)
    if not text:
        return None
    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d-%b-%y",
        "%d-%b-%Y",
    ):
        try:
            return dt.datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def school_year_for_date(value: dt.date) -> str:
    start = value.year if value.month >= 6 else value.year - 1
    return f"{start}-{start + 1}"


def numeric_value(value: Any) -> Optional[float]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return None if math.isnan(number) else number
    text = (
        clean_text(value)
        .replace(",", "")
        .replace("PHP", "")
        .replace("₱", "")
    )
    if not text or text.startswith("#"):
        return None
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except ValueError:
        return None


def contact_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if len(digits) == 10 and digits.startswith("9"):
        return f"0{digits}"
    return text


def split_name(profile_name: str, file_name: str) -> tuple[str, str]:
    def remove_version_suffix(value: str) -> str:
        cleaned = clean_text(value)
        cleaned = re.sub(r"[_-]\d{6,8}$", "", cleaned)
        cleaned = re.sub(
            r"[_-](?:graduate|graduated|quit|transfer|g\d+).*$",
            "",
            cleaned,
            flags=re.IGNORECASE,
        )
        return clean_text(cleaned)

    text = remove_version_suffix(profile_name)
    if "," in text:
        last_name, first_name = text.split(",", 1)
        return remove_version_suffix(last_name), remove_version_suffix(first_name)

    stem = re.sub(
        r"(?:[_-](?:graduate|graduated|quit|g\d+|transfer).*)$",
        "",
        Path(file_name).stem,
        flags=re.IGNORECASE,
    )
    if "," in stem:
        last_name, first_name = stem.split(",", 1)
        return remove_version_suffix(last_name), remove_version_suffix(first_name)

    parts = text.split()
    if len(parts) >= 2:
        return parts[-1], " ".join(parts[:-1])
    return text, ""


def status_from_source(file_name: str, grade: str) -> str:
    marker = f"{file_name} {grade}".casefold()
    if "graduat" in marker:
        return "Graduated"
    if any(token in marker for token in ("quit", "transfer", "removed", "inactive")):
        return "Inactive/Removed"
    return "Active"


def media_entries(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as archive:
        return [
            name
            for name in archive.namelist()
            if name.casefold().startswith("xl/media/")
            and not name.endswith("/")
            and Path(name).suffix.casefold()
            in {".png", ".jpg", ".jpeg", ".bmp", ".webp"}
        ]


def workbook_recency_score(path: Path) -> tuple[int, int, int]:
    stem = path.stem
    year_score = 0
    for token in re.findall(r"(?<!\d)(\d{6,8})(?!\d)", stem):
        for pattern in ("%m%d%Y", "%m%d%y"):
            try:
                parsed = dt.datetime.strptime(token, pattern).date()
                year_score = max(year_score, parsed.toordinal())
            except ValueError:
                continue
    grade_version = 1 if re.search(r"(?:^|[_\s-])G\d+", stem, re.I) else 0
    return year_score, grade_version, len(stem)


def expense_description(
    values: list[Any],
    *,
    particulars_column: int,
    balance_column: int,
    student_name: str,
) -> str:
    candidates = []
    student_key = normalized_text(student_name)
    for column in range(particulars_column, max(particulars_column, balance_column)):
        value = values[column - 1] if column - 1 < len(values) else None
        if numeric_value(value) is not None:
            continue
        text = clean_text(value)
        if not text or normalized_text(text) == student_key:
            continue
        if re.fullmatch(r"\d{4,8}(?:\s*[A-Za-z])?", text):
            continue
        if text.casefold() in {"--split--", "debit", "credit", "balance"}:
            continue
        candidates.append(text)
    unique = []
    for candidate in candidates:
        if normalized_text(candidate) not in {
            normalized_text(existing) for existing in unique
        }:
            unique.append(candidate)
    return " — ".join(unique)[:500]


def parse_expenses(sheet, header_row: int, student_name: str) -> list[dict[str, Any]]:
    header_columns = {}
    for column in range(1, min(sheet.max_column, 16) + 1):
        key = normalized_text(sheet.cell(header_row, column).value)
        if key:
            header_columns[key] = column

    date_column = header_columns.get("date", 1)
    particulars_column = header_columns.get("particulars", 2)
    debit_column = header_columns.get("debit")
    credit_column = header_columns.get("credit")
    balance_column = header_columns.get(
        "balance", min(sheet.max_column, (credit_column or 6) + 1)
    )
    remarks_column = header_columns.get("remarks")
    previous_balance = None
    expenses = []

    for row in range(header_row + 1, sheet.max_row + 1):
        values = [
            sheet.cell(row, column).value
            for column in range(1, min(sheet.max_column, 16) + 1)
        ]
        row_date = parse_date(
            values[date_column - 1] if date_column - 1 < len(values) else None,
            epoch=sheet.parent.epoch,
        )
        balance = (
            numeric_value(values[balance_column - 1])
            if balance_column and balance_column - 1 < len(values)
            else None
        )
        amount = None
        if row_date:
            entry_end = min(
                len(values),
                max(debit_column or 0, credit_column or 0),
            )
            money_start = max(particulars_column + 1, 3)
            entry_values = [
                numeric_value(values[index - 1])
                for index in range(money_start, entry_end + 1)
            ]
            negative_values = [
                abs(value)
                for value in entry_values
                if value is not None and value < 0
            ]
            if negative_values:
                amount = max(negative_values)
            elif debit_column and debit_column - 1 < len(values):
                debit = numeric_value(values[debit_column - 1])
                credit = (
                    numeric_value(values[credit_column - 1])
                    if credit_column and credit_column - 1 < len(values)
                    else None
                )
                if debit is not None and debit > 0 and not (credit and credit > 0):
                    amount = debit
            if (
                amount is None
                and balance is not None
                and balance < 0
                and not any(value is not None for value in entry_values)
            ):
                # A small group of legacy sheets stored a standalone expense
                # directly in the Balance column and left Debit/Credit empty.
                amount = abs(balance)
            if (
                amount is None
                and previous_balance is not None
                and balance is not None
                and previous_balance > balance
            ):
                amount = previous_balance - balance

        if row_date and amount is not None and amount > 0:
            description = expense_description(
                values,
                particulars_column=particulars_column,
                balance_column=balance_column,
                student_name=student_name,
            )
            if not description:
                description = "Historical student expense"
            if remarks_column and remarks_column - 1 < len(values):
                remarks = clean_text(values[remarks_column - 1])
                if remarks and normalized_text(remarks) not in normalized_text(description):
                    description = f"{description} — {remarks}"[:500]
            expenses.append(
                {
                    "description": description,
                    "amount": round(float(amount), 2),
                    "date": row_date.isoformat(),
                    "school_year": school_year_for_date(row_date),
                }
            )
        if balance is not None:
            previous_balance = balance

    deduped = {}
    for expense in expenses:
        key = (
            expense["date"],
            round(expense["amount"], 2),
            normalized_text(expense["description"]),
        )
        deduped[key] = expense
    return list(deduped.values())


def parse_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = max(
            workbook.worksheets,
            key=lambda candidate: candidate.max_row * candidate.max_column,
        )
        fields = {}
        header_row = None
        for row in range(1, min(sheet.max_row, 100) + 1):
            label = normalized_text(sheet.cell(row, 1).value)
            field = PROFILE_LABELS.get(label)
            if field and field not in fields:
                fields[field] = first_value_to_right(sheet, row)
            row_labels = {
                normalized_text(sheet.cell(row, column).value)
                for column in range(1, min(sheet.max_column, 16) + 1)
            }
            if "date" in row_labels and (
                "particulars" in row_labels
                or "debit" in row_labels
                or "credit" in row_labels
            ):
                header_row = row

        last_name, first_name = split_name(
            clean_text(fields.get("name")),
            path.name,
        )
        father = clean_text(fields.get("father"))
        mother = clean_text(fields.get("mother"))
        parents = clean_text(fields.get("parents"))
        if not parents:
            parent_parts = []
            if father:
                parent_parts.append(f"Father: {father}")
            if mother:
                parent_parts.append(f"Mother: {mother}")
            parents = " / ".join(parent_parts)

        birthday = parse_date(fields.get("birthday"), epoch=workbook.epoch)
        grade = clean_text(fields.get("grade"))
        notes = []
        student_update = clean_text(fields.get("student_update"))
        coordinator = clean_text(fields.get("coordinator"))
        email = clean_text(fields.get("email"))
        if student_update:
            notes.append(student_update)
        if coordinator:
            notes.append(f"Coordinator: {coordinator}")
        if email:
            notes.append(f"Email/FB: {email}")

        student_name = f"{last_name}, {first_name}".strip(", ")
        expenses = (
            parse_expenses(sheet, header_row, student_name)
            if header_row
            else []
        )
        media = media_entries(path)
        student = {
            "last_name": last_name,
            "first_name": first_name,
            "gender": clean_text(fields.get("gender")),
            "grade": grade,
            "address": clean_text(fields.get("address")),
            "city": "Manila",
            "area": "Manila",
            "birthday": birthday.isoformat() if birthday else None,
            "sponsor": clean_text(fields.get("sponsor")),
            "contact": contact_text(fields.get("contact")),
            "school": clean_text(fields.get("school")),
            "parents": parents,
            "course": clean_text(fields.get("course")),
            "remarks": " | ".join(notes),
            "status": status_from_source(path.name, grade),
        }
        return {
            "source_file": path.name,
            "source_path": str(path),
            "student": student,
            "expenses": expenses,
            "media": media,
            "score": (
                *workbook_recency_score(path),
                sum(bool(clean_text(value)) for value in student.values()),
                len(expenses),
                len(media),
            ),
        }
    finally:
        workbook.close()


def choose_canonical(records: Iterable[dict[str, Any]]) -> tuple[list[dict], dict]:
    grouped = defaultdict(list)
    for record in records:
        student = record["student"]
        grouped[identity_key(student["last_name"], student["first_name"])].append(record)

    exact_selected = []
    duplicates = {}
    for key, candidates in sorted(grouped.items()):
        ordered = sorted(candidates, key=lambda record: record["score"], reverse=True)
        exact_selected.append(ordered[0])
        if len(ordered) > 1:
            duplicates[key] = {
                "selected": ordered[0]["source_file"],
                "ignored": [record["source_file"] for record in ordered[1:]],
            }

    selected = []
    for record in sorted(
        exact_selected,
        key=lambda candidate: candidate["score"],
        reverse=True,
    ):
        student = record["student"]
        source_name = normalized_text(
            f"{student['last_name']}{student['first_name']}"
        )
        duplicate_of = None
        for existing in selected:
            existing_student = existing["student"]
            same_last = normalized_text(student["last_name"]) == normalized_text(
                existing_student["last_name"]
            )
            same_first = normalized_text(student["first_name"]) == normalized_text(
                existing_student["first_name"]
            )
            existing_name = normalized_text(
                f"{existing_student['last_name']}{existing_student['first_name']}"
            )
            similarity = SequenceMatcher(
                None, source_name, existing_name
            ).ratio()
            if similarity >= 0.95 and (same_last or same_first):
                duplicate_of = existing
                break
        if duplicate_of is None:
            selected.append(record)
            continue

        duplicate_student = duplicate_of["student"]
        key = (
            "close:"
            f"{identity_key(duplicate_student['last_name'], duplicate_student['first_name'])}"
        )
        detail = duplicates.setdefault(
            key,
            {
                "selected": duplicate_of["source_file"],
                "ignored": [],
            },
        )
        detail["ignored"].append(record["source_file"])

    return sorted(selected, key=lambda record: record["source_file"]), duplicates


def source_expense_key(student_id: Any, expense: dict[str, Any]) -> tuple:
    return (
        str(student_id),
        expense["date"],
        round(float(expense["amount"]), 2),
        normalized_text(expense["description"]),
    )


def merge_missing_fields(existing: dict, source: dict) -> dict:
    fields = {}
    for key, value in source.items():
        if key == "status":
            continue
        if clean_text(value) and not clean_text(existing.get(key)):
            fields[key] = value
    if not clean_text(existing.get("status")):
        fields["status"] = source.get("status") or "Active"
    return fields


def close_match(source: dict, existing_rows: list[dict]) -> Optional[tuple[dict, float]]:
    source_name = normalized_text(
        f"{source.get('last_name', '')}{source.get('first_name', '')}"
    )
    best = None
    best_score = 0.0
    for row in existing_rows:
        candidate_name = normalized_text(
            f"{row.get('last_name', '')}{row.get('first_name', '')}"
        )
        score = SequenceMatcher(None, source_name, candidate_name).ratio()
        if score > best_score:
            best = row
            best_score = score
    if best is not None and best_score >= 0.93:
        same_last = normalized_text(source.get("last_name")) == normalized_text(
            best.get("last_name")
        )
        same_first = normalized_text(source.get("first_name")) == normalized_text(
            best.get("first_name")
        )
        if same_last or same_first:
            return best, best_score
    return None


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )


def chunks(items: list[Any], size: int = 50) -> Iterable[list[Any]]:
    for index in range(0, len(items), size):
        yield items[index:index + size]


def insert_with_fallback(client, table: str, rows: list[dict]) -> int:
    inserted = 0
    for batch in chunks(rows, 50):
        try:
            response = client.table(table).insert(batch).execute()
            inserted += len(response.data or batch)
        except Exception:
            for row in batch:
                response = client.table(table).insert(row).execute()
                inserted += len(response.data or [row])
    return inserted


def build_plan(source_dir: Path, client, output_dir: Path) -> tuple[dict, list[dict]]:
    workbooks = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if path.name.casefold() not in SUMMARY_FILES
    )
    parsed = [parse_workbook(path) for path in workbooks]
    selected, source_duplicates = choose_canonical(parsed)

    existing_rows = list(
        client.table("students").select(STUDENT_COLUMNS).execute().data or []
    )
    existing_by_identity = defaultdict(list)
    for row in existing_rows:
        existing_by_identity[
            identity_key(row.get("last_name"), row.get("first_name"))
        ].append(row)

    new_records = []
    updates = []
    matched = []
    fuzzy_matches = []
    close_matches = []
    historical_inactive_defaulted = []
    for record in selected:
        student = record["student"]
        source_identity = identity_key(
            student["last_name"], student["first_name"]
        )
        matches = existing_by_identity.get(source_identity, [])
        match_type = "exact"
        similarity = 1.0
        if not matches:
            fuzzy = close_match(student, existing_rows)
            if fuzzy:
                fuzzy_row, similarity = fuzzy
                matches = [fuzzy_row]
                match_type = "fuzzy"
        if matches:
            existing = matches[0]
            fields = merge_missing_fields(existing, student)
            if fields:
                updates.append(
                    {
                        "id": existing["id"],
                        "fields": fields,
                        "source_file": record["source_file"],
                    }
                )
            matched.append(
                {
                    "source_file": record["source_file"],
                    "source_identity": source_identity,
                    "id": existing["id"],
                    "name": f"{student['last_name']}, {student['first_name']}",
                    "match_type": match_type,
                }
            )
            if match_type == "fuzzy":
                fuzzy_matches.append(
                    {
                        "source_file": record["source_file"],
                        "source": f"{student['last_name']}, {student['first_name']}",
                        "candidate": (
                            f"{existing.get('last_name')}, "
                            f"{existing.get('first_name')}"
                        ),
                        "candidate_id": existing.get("id"),
                        "similarity": round(similarity, 3),
                    }
                )
        else:
            insert_record = dict(student)
            source_date_ordinal = record["score"][0]
            recent_activity = any(
                expense["date"] >= "2025-06-01"
                for expense in record["expenses"]
            ) or source_date_ordinal >= dt.date(2025, 6, 1).toordinal()
            if insert_record.get("status") == "Active" and not recent_activity:
                insert_record["status"] = "Inactive/Removed"
                historical_note = (
                    "Historical Manila workbook import; current status needs review."
                )
                current_remarks = clean_text(insert_record.get("remarks"))
                insert_record["remarks"] = (
                    f"{current_remarks} | {historical_note}"
                    if current_remarks
                    else historical_note
                )
                historical_inactive_defaulted.append(
                    {
                        "name": (
                            f"{student['last_name']}, {student['first_name']}"
                        ),
                        "source_file": record["source_file"],
                    }
                )
            new_records.append(
                {
                    "record": insert_record,
                    "source_file": record["source_file"],
                }
            )
            candidate = close_match(student, existing_rows)
            if candidate:
                candidate_row, candidate_score = candidate
                close_matches.append(
                    {
                        "source_file": record["source_file"],
                        "source": f"{student['last_name']}, {student['first_name']}",
                        "candidate": (
                            f"{candidate_row.get('last_name')}, "
                            f"{candidate_row.get('first_name')}"
                        ),
                        "candidate_id": candidate_row.get("id"),
                        "similarity": round(candidate_score, 3),
                    }
                )

    expense_rows = [
        {
            "source_file": record["source_file"],
            "name": (
                f"{record['student']['last_name']}, "
                f"{record['student']['first_name']}"
            ),
            **expense,
        }
        for record in selected
        for expense in record["expenses"]
    ]
    expense_by_year = defaultdict(lambda: {"rows": 0, "total": 0.0})
    for expense in expense_rows:
        summary = expense_by_year[expense["school_year"]]
        summary["rows"] += 1
        summary["total"] = round(
            summary["total"] + float(expense["amount"]), 2
        )
    expense_outliers = sorted(
        (
            expense
            for expense in expense_rows
            if float(expense["amount"]) >= 50000
        ),
        key=lambda expense: float(expense["amount"]),
        reverse=True,
    )

    id_by_source_identity = {
        match["source_identity"]: match["id"] for match in matched
    }
    existing_expense_keys = set()
    matched_ids = sorted(set(id_by_source_identity.values()), key=str)
    for batch in chunks(matched_ids, 100):
        rows = (
            client.table("expenses")
            .select("student_id,description,amount,date,school_year")
            .in_("student_id", batch)
            .execute()
            .data
            or []
        )
        existing_expense_keys.update(
            source_expense_key(row["student_id"], row) for row in rows
        )
    estimated_new_expenses = 0
    for record in selected:
        student = record["student"]
        student_id = id_by_source_identity.get(
            identity_key(student["last_name"], student["first_name"])
        )
        for expense in record["expenses"]:
            if student_id is None or source_expense_key(
                student_id, expense
            ) not in existing_expense_keys:
                estimated_new_expenses += 1

    plan = {
        "source_directory": str(source_dir),
        "source_workbooks": len(workbooks),
        "canonical_students": len(selected),
        "source_duplicates": source_duplicates,
        "database_students_before": len(existing_rows),
        "exact_database_matches": sum(
            match["match_type"] == "exact" for match in matched
        ),
        "fuzzy_database_matches": fuzzy_matches,
        "database_matches": matched,
        "students_to_insert": len(new_records),
        "historical_students_defaulted_inactive": len(
            historical_inactive_defaulted
        ),
        "historical_status_review": historical_inactive_defaulted,
        "students_to_fill_missing_fields": len(updates),
        "close_matches_not_automatically_merged": close_matches,
        "source_expense_candidates": sum(
            len(record["expenses"]) for record in selected
        ),
        "expenses_to_insert_estimate": estimated_new_expenses,
        "expense_summary_by_school_year": dict(sorted(expense_by_year.items())),
        "expense_outliers_at_least_50000": expense_outliers,
        "embedded_document_images": sum(
            len(record["media"]) for record in selected
        ),
        "profile_photos_to_upload": 0,
        "embedded_images_note": (
            "The workbook images are grade reports and school documents, "
            "not student portraits; they are intentionally not uploaded as "
            "profile photos."
        ),
        "new_students": new_records,
        "updates": updates,
    }
    write_json(output_dir / "plan.json", plan)
    return plan, selected


def commit_plan(
    plan: dict,
    selected: list[dict],
    client,
    output_dir: Path,
) -> dict:
    existing_rows = list(
        client.table("students").select(STUDENT_COLUMNS).execute().data or []
    )
    matched_ids = sorted(
        {match["id"] for match in plan["database_matches"]},
        key=str,
    )
    backup_expenses = []
    for batch in chunks(matched_ids, 100):
        if not batch:
            continue
        backup_expenses.extend(
            client.table("expenses")
            .select("*")
            .in_("student_id", batch)
            .execute()
            .data
            or []
        )
    write_json(
        output_dir / "backup_before_commit.json",
        {
            "students": existing_rows,
            "expenses_for_matched_students": backup_expenses,
        },
    )

    updated = 0
    for change in plan["updates"]:
        response = (
            client.table("students")
            .update(change["fields"])
            .eq("id", change["id"])
            .execute()
        )
        updated += len(response.data or [])

    inserted_students = insert_with_fallback(
        client,
        "students",
        [item["record"] for item in plan["new_students"]],
    )

    refreshed = list(
        client.table("students").select(STUDENT_COLUMNS).execute().data or []
    )
    id_by_identity = {
        identity_key(row.get("last_name"), row.get("first_name")): row["id"]
        for row in refreshed
    }
    id_by_source_identity = {
        match["source_identity"]: match["id"]
        for match in plan["database_matches"]
    }
    id_by_source_identity.update(id_by_identity)
    source_expenses = []
    for record in selected:
        student = record["student"]
        student_id = id_by_source_identity.get(
            identity_key(student["last_name"], student["first_name"])
        )
        if not student_id:
            continue
        for expense in record["expenses"]:
            source_expenses.append({"student_id": student_id, **expense})

    relevant_ids = sorted({row["student_id"] for row in source_expenses}, key=str)
    existing_expenses = []
    for batch in chunks(relevant_ids, 100):
        existing_expenses.extend(
            client.table("expenses")
            .select("student_id,description,amount,date,school_year")
            .in_("student_id", batch)
            .execute()
            .data
            or []
        )
    existing_expense_keys = {
        source_expense_key(row["student_id"], row)
        for row in existing_expenses
    }
    new_expenses = [
        row
        for row in source_expenses
        if source_expense_key(row["student_id"], row)
        not in existing_expense_keys
    ]
    inserted_expenses = insert_with_fallback(client, "expenses", new_expenses)

    result = {
        "students_updated": updated,
        "students_inserted": inserted_students,
        "expenses_inserted": inserted_expenses,
        "photos_uploaded": 0,
        "embedded_document_images_skipped": plan["embedded_document_images"],
        "database_students_after": len(refreshed),
    }
    write_json(output_dir / "commit_result.json", result)
    AuditRepository(client).log(
        operator="Codex Manila Import",
        action="import",
        entity_type="workbook_archive",
        details={
            "source": "Manila Students File.zip",
            **{
                key: value
                for key, value in result.items()
                if isinstance(value, (int, float, str))
            },
        },
    )
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_directory", type=Path)
    parser.add_argument("--commit", action="store_true")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("outputs") / "manila_students_import",
    )
    args = parser.parse_args()
    source_dir = args.source_directory.resolve()
    output_dir = args.output.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    client = get_supabase()
    plan, selected = build_plan(source_dir, client, output_dir)
    print(
        json.dumps(
            {
                key: value
                for key, value in plan.items()
                if key
                not in {
                    "new_students",
                    "updates",
                    "source_duplicates",
                    "close_matches_not_automatically_merged",
                    "database_matches",
                    "historical_status_review",
                }
            },
            indent=2,
        )
    )
    print(f"PLAN={output_dir / 'plan.json'}")
    if not args.commit:
        print("DRY_RUN_ONLY")
        return 0

    result = commit_plan(plan, selected, client, output_dir)
    print(json.dumps(result, indent=2))
    print(f"RESULT={output_dir / 'commit_result.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
