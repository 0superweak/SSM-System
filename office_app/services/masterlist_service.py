"""Business logic for current master-list reference data."""

from __future__ import annotations

import os
import re
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook

from office_app.services.workbook_import_service import WorkbookImportService

StudentKey = Tuple[str, str, str]
ReferenceMap = Dict[StudentKey, Dict[str, str]]


class MasterListService:
    """Resolve the latest master-list sheet and apply its student status reference."""

    def __init__(self, workbook_import_service: Optional[WorkbookImportService] = None) -> None:
        self.workbook_import_service = workbook_import_service or WorkbookImportService()
        self._cache_key: Optional[Tuple[Any, ...]] = None
        self._cache: Optional[ReferenceMap] = None

    def filter_current_rows(self, rows: Sequence[Mapping[str, Any]], reference: ReferenceMap) -> List[Dict[str, Any]]:
        if not reference:
            return [dict(row) for row in rows]

        current_rows: List[Dict[str, Any]] = []
        for row in rows:
            master_row = reference.get(self.student_reference_key(row))
            if not master_row:
                continue
            merged = dict(row)
            merged["status"] = master_row["status"]
            merged["_status_source"] = "masterlist"
            current_rows.append(merged)
        return current_rows

    def apply_current_status(self, row: Mapping[str, Any], reference: ReferenceMap) -> Dict[str, Any]:
        rows = self.filter_current_rows([row], reference)
        return rows[0] if rows else dict(row)

    def current_student_keys(self, reference: ReferenceMap) -> set[StudentKey]:
        return set(reference)

    def current_student_reference(
        self,
        *,
        workbook: Optional[Any] = None,
        workbook_path: Optional[str] = None,
        workbook_revision: int = 0,
        saved_workbook_path: str = "",
        cwd: Optional[str] = None,
        fallback_paths: Optional[Iterable[str]] = None,
    ) -> ReferenceMap:
        if workbook is not None:
            latest = self.latest_master_sheet_name_from_names(workbook.sheetnames)
            if latest:
                cache_key = ("open", workbook_path, latest, workbook_revision)
                if self._cache_key == cache_key and self._cache is not None:
                    return self._cache
                reference = self.master_sheet_student_reference(workbook[latest])
                self._cache_key = cache_key
                self._cache = reference
                return reference

        path = self.current_masterlist_path(
            workbook_path=workbook_path,
            saved_workbook_path=saved_workbook_path,
            cwd=cwd,
            fallback_paths=fallback_paths,
        )
        if not path:
            return {}

        cache_key = ("file", path, os.path.getmtime(path))
        if self._cache_key == cache_key and self._cache is not None:
            return self._cache

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            latest = self.latest_master_sheet_name_from_names(wb.sheetnames)
            if not latest:
                return {}
            reference = self.master_sheet_student_reference(wb[latest])
            self._cache_key = cache_key
            self._cache = reference
            return reference
        finally:
            wb.close()

    def invalidate_cache(self) -> None:
        self._cache_key = None
        self._cache = None

    def current_masterlist_path(
        self,
        *,
        workbook_path: Optional[str] = None,
        saved_workbook_path: str = "",
        cwd: Optional[str] = None,
        fallback_paths: Optional[Iterable[str]] = None,
    ) -> str:
        candidates = [
            workbook_path,
            saved_workbook_path,
            os.path.join(cwd or os.getcwd(), "SSM Masterlist.xlsx"),
        ]
        candidates.extend(fallback_paths or [])
        for path in candidates:
            if path and os.path.exists(path):
                return path
        return ""

    def latest_master_sheet_name_from_names(self, sheet_names: Sequence[str]) -> Optional[str]:
        master_sheets = [name for name in sheet_names if "master" in name.lower()]
        if not master_sheets:
            return None
        return max(master_sheets, key=self.sheet_year_score)

    def master_sheet_student_keys(self, worksheet: Any) -> set[StudentKey]:
        return set(self.master_sheet_student_reference(worksheet))

    def master_sheet_student_reference(self, worksheet: Any) -> ReferenceMap:
        rows = list(worksheet.iter_rows(values_only=True))
        header_idx = self.workbook_import_service.find_header_row(rows, ("last name", "first name"))
        if header_idx is None:
            return {}

        headers = self.workbook_import_service.header_map(rows[header_idx])
        last_i = headers.get("last name")
        first_i = headers.get("first name")
        birthday_i = headers.get("birthday")
        grade_i = headers.get("level") or headers.get("grade") or headers.get("grade level")
        if last_i is None or first_i is None:
            return {}

        reference: ReferenceMap = {}
        for row in rows[header_idx + 1:]:
            first_cell = self.workbook_import_service.safe_cell(row, 0)
            last = self.workbook_import_service.safe_cell(row, last_i)
            first = self.workbook_import_service.safe_cell(row, first_i)
            if not last and not first:
                continue
            if not last or not first or last.lower() == "last name":
                continue

            key = (
                self.normalize_student_key_value(last),
                self.normalize_student_key_value(first),
                self.normalize_student_key_value(self.workbook_import_service.safe_cell(row, birthday_i)),
            )
            marker = first_cell.lower()
            grade_text = self.workbook_import_service.safe_cell(row, grade_i).lower()
            if marker in ("g", "graduated") or "graduat" in grade_text:
                status = "Graduated"
            elif not marker or marker == "x":
                status = "Inactive/Removed"
            else:
                status = "Active"
            reference[key] = {"status": status}
        return reference

    def student_reference_key(self, row: Mapping[str, Any]) -> StudentKey:
        return (
            self.normalize_student_key_value(row.get("last_name")),
            self.normalize_student_key_value(row.get("first_name")),
            self.normalize_student_key_value(row.get("birthday")),
        )

    def normalize_student_key_value(self, value: Any) -> str:
        if hasattr(value, "strftime"):
            value = value.strftime("%Y-%m-%d")
        return " ".join(str(value or "").strip().lower().split())

    def sheet_year_score(self, sheet_name: str) -> int:
        pairs = re.findall(r"(20\d{2}|\d{2})\s*-\s*(20\d{2}|\d{2})", sheet_name)
        if pairs:
            _, end = pairs[-1]
            return int(end if len(end) == 4 else "20" + end)
        years = re.findall(r"20\d{2}", sheet_name)
        if years:
            return int(years[-1])
        return 0