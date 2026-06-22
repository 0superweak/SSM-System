"""Excel import/export business logic for student records."""

from __future__ import annotations

import datetime as _dt
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from office_app.repositories.student_repository import StudentRepository


class StudentExcelService:
    """Create student Excel exports and parse/import student Excel workbooks."""

    EXPORT_COLUMNS = [
        "Last Name", "First Name", "Gender", "Grade/Level",
        "Address", "City", "Area/Coordinator", "Birthday",
        "Sponsor", "Contact No.", "School", "Parents & Occupation",
        "Course", "Remarks", "Status",
    ]
    EXPORT_KEYS = [
        "last_name", "first_name", "gender", "grade",
        "address", "city", "area", "birthday",
        "sponsor", "contact", "school", "parents",
        "course", "remarks", "status",
    ]

    def __init__(self, student_repository: Optional[StudentRepository] = None) -> None:
        self.student_repository = student_repository or StudentRepository()

    def export_students(self, rows: Sequence[Mapping[str, Any]], path: str) -> int:
        data = [[row.get(key, "") or "" for key in self.EXPORT_KEYS] for row in rows]
        df = pd.DataFrame(data, columns=self.EXPORT_COLUMNS)
        df.to_excel(path, index=False, sheet_name="Students")
        self._polish_export_workbook(path, rows)
        return len(rows)

    def import_students(self, path: str, *, chunk_size: int = 50) -> Tuple[str, int]:
        sheet_name, records = self.parse_students(path)
        imported = self.student_repository.import_students(records)
        return sheet_name, imported

    def parse_students(self, path: str) -> Tuple[str, List[Dict[str, Any]]]:
        sheet_name = self._select_import_sheet(path)
        df = pd.read_excel(path, sheet_name=sheet_name, header=None)
        header_row = self._find_import_header_row(df)
        col_map = self._import_column_map(df, header_row)

        records: List[Dict[str, Any]] = []
        current_area = ""
        for row_index, row in df.iterrows():
            if row_index <= header_row:
                continue

            last = str(self._series_value(row, col_map.get("last_name", 1)) or "").strip()
            first = str(self._series_value(row, col_map.get("first_name", 2)) or "").strip()
            first_value = self._series_value(row, col_map.get("first_name", 2))
            if (not last or last.lower() in ("nan", "last name")) and pd.isna(first_value):
                cell0 = str(self._series_value(row, 0) or "").strip()
                if cell0 and cell0.lower() not in ("nan", ""):
                    current_area = cell0
                continue
            if not last or last.lower() == "nan":
                continue
            if not first or first.lower() == "nan":
                continue

            birthday = self._format_birthday(self._series_value(row, col_map.get("birthday", 8)))
            row_marker = str(self._series_value(row, 0) or "").strip().lower()
            status = self._import_status(
                status_raw=self._cell_text(row, col_map, "status").lower(),
                row_marker=row_marker,
                grade_text=self._cell_text(row, col_map, "grade").lower(),
            )

            records.append({
                "last_name": last,
                "first_name": first,
                "gender": self._cell_text(row, col_map, "gender"),
                "grade": self._cell_text(row, col_map, "grade"),
                "address": self._cell_text(row, col_map, "address"),
                "city": self._cell_text(row, col_map, "city"),
                "area": self._cell_text(row, col_map, "area") or current_area,
                "birthday": birthday,
                "sponsor": self._cell_text(row, col_map, "sponsor"),
                "contact": self._cell_text(row, col_map, "contact"),
                "school": self._cell_text(row, col_map, "school"),
                "parents": self._cell_text(row, col_map, "parents"),
                "course": self._cell_text(row, col_map, "course"),
                "remarks": self._cell_text(row, col_map, "remarks"),
                "status": status,
            })

        return sheet_name, records

    def _polish_export_workbook(self, path: str, rows: Sequence[Mapping[str, Any]]) -> None:
        wb = load_workbook(path)
        ws = wb.active

        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        active_fill = PatternFill("solid", fgColor="E8F5E9")
        inactive_fill = PatternFill("solid", fgColor="FFEBEE")
        graduated_fill = PatternFill("solid", fgColor="F3E8FF")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[1].height = 28

        status_col = self.EXPORT_COLUMNS.index("Status") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_val = row[status_col - 1].value or ""
            if status_val == "Inactive/Removed":
                row_fill = inactive_fill
            elif status_val == "Graduated":
                row_fill = graduated_fill
            else:
                row_fill = active_fill
            for cell in row:
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"
        total = len(rows)
        active_cnt = sum(1 for row in rows if row.get("status") == "Active")
        inactive_cnt = sum(1 for row in rows if row.get("status") == "Inactive/Removed")
        graduated_cnt = sum(1 for row in rows if row.get("status") == "Graduated")
        ws.append([])
        ws.append([
            f"Total: {total} students  |  Active: {active_cnt}  |  "
            f"Inactive: {inactive_cnt}  |  Graduated: {graduated_cnt}"
        ])
        ws.cell(ws.max_row, 1).font = Font(bold=True, italic=True, color="555555")
        wb.save(path)

    def _select_import_sheet(self, path: str) -> str:
        wb = load_workbook(path, read_only=True)
        try:
            sheets = wb.sheetnames
            master_sheets = [sheet for sheet in sheets if "master" in sheet.lower()]
            return master_sheets[-1] if master_sheets else sheets[0]
        finally:
            wb.close()

    def _find_import_header_row(self, df: pd.DataFrame) -> int:
        for index, row in df.iterrows():
            if any(str(value).strip().lower() == "last name" for value in row if pd.notna(value)):
                return int(index)
        return 1

    def _import_column_map(self, df: pd.DataFrame, header_row: int) -> Dict[str, int]:
        col_map: Dict[str, int] = {}
        for index, value in enumerate(df.iloc[header_row]):
            if pd.isna(value):
                continue
            text = str(value).strip().lower()
            if text == "last name":
                col_map["last_name"] = index
            elif text == "first name":
                col_map["first_name"] = index
            elif text == "gender":
                col_map["gender"] = index
            elif text in ("level", "grade level", "grade"):
                col_map["grade"] = index
            elif text == "location" and "address" not in col_map:
                col_map["address"] = index
            elif text == "birthday":
                col_map["birthday"] = index
            elif text == "sponsor":
                col_map["sponsor"] = index
            elif text == "contact no.":
                col_map["contact"] = index
            elif text == "school":
                col_map["school"] = index
            elif text == "parents":
                col_map["parents"] = index
            elif text == "course":
                col_map["course"] = index
            elif text == "remarks":
                col_map["remarks"] = index
            elif text == "status":
                col_map["status"] = index

        if "address" in col_map:
            address_index = col_map["address"]
            col_map["city"] = address_index + 1
            col_map["area"] = address_index + 2
        return col_map

    def _cell_text(self, row: Any, col_map: Mapping[str, int], key: str, default: str = "") -> str:
        index = col_map.get(key)
        if index is None:
            return default
        value = self._series_value(row, index)
        if pd.isna(value):
            return default
        return str(value).strip().replace("nan", "")

    def _series_value(self, row: Any, index: Optional[int]) -> Any:
        if index is None:
            return None
        try:
            return row[index]
        except Exception:
            return None

    def _format_birthday(self, value: Any) -> str:
        if pd.isna(value):
            return ""
        try:
            if isinstance(value, (_dt.datetime, _dt.date)):
                return value.strftime("%Y-%m-%d")
            return str(value).strip()
        except Exception:
            return str(value).strip()

    def _import_status(self, *, status_raw: str, row_marker: str, grade_text: str) -> str:
        if "graduated" in status_raw or row_marker in ("g", "graduated") or "graduat" in grade_text:
            return "Graduated"
        if "inactive" in status_raw or "removed" in status_raw or row_marker == "x":
            return "Inactive/Removed"
        return "Active"
