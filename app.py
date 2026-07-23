import sys
import os
import logging
import shutil
import tempfile
import threading
import time
from datetime import datetime

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLineEdit, QListWidget, QListWidgetItem, QTextEdit,
    QLabel, QPushButton, QStackedWidget, QFormLayout, QFileDialog,
    QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QScrollArea, QComboBox, QStatusBar, QDialog, QTabWidget,
    QProgressBar, QProgressDialog, QFrame, QGridLayout, QScrollBar, QDateEdit,
    QGraphicsOpacityEffect, QGraphicsDropShadowEffect, QSizePolicy, QStyle,
    QMenu, QSpacerItem
)
from PyQt6.QtWidgets import (
    QComboBox as NativeComboBox,
    QLabel as NativeLabel,
    QProgressBar as NativeProgressBar,
    QPushButton as NativePushButton,
)
from PyQt6.QtCore import (
    Qt, QTimer, pyqtSignal, QObject, QRectF, QSettings, QSize, QDate,
    QPropertyAnimation, QEasingCurve, QThreadPool, pyqtProperty, QPointF
)
from PyQt6.QtGui import (
    QPixmap, QPainter, QColor, QPen, QIcon, QImage, QLinearGradient, QFont,
    QRadialGradient, QPainterPath, QAction, QKeySequence, QFontDatabase,
    QRegion,
)

from supabase import Client
from office_app.app_config import (
    APP_ICON_ASSET,
    KEEPALIVE_INTERVAL_MS,
    LOGO_ASSET,
    USERS,
    get_sheet_sync_token,
)
from office_app.services.supabase_client import get_supabase, reset_supabase_clients
from office_app.utils.paths import resource_path
from office_app.utils.background_tasks import BackgroundTask
from office_app.repositories.coordinator_repository import CoordinatorRepository
from office_app.repositories.audit_repository import AuditRepository
from office_app.repositories.student_repository import StudentRepository
from office_app.repositories.workbook_repository import WorkbookRepository
from office_app.services.expense_service import ExpenseService
from office_app.services.dashboard_service import DashboardService
from office_app.services.masterlist_service import MasterListService
from office_app.services.photo_service import PhotoService
from office_app.services.student_service import StudentService
from office_app.services.student_excel_service import StudentExcelService
from office_app.services.student_list_service import StudentListService
from office_app.services.workbook_import_service import WorkbookImportService
from office_app.services.google_sheet_sync_service import GoogleSheetSyncService
from office_app.ui import (
    ActionButton,
    Card,
    EmptyState,
    Spacing,
    StatusBadge,
    set_content_hugging_button,
)
from office_app.ui.motion import (
    MotionCard,
    PulseController,
    animate_count,
    animate_progress,
    attach_press_feedback,
    fade_in,
)
from office_app.ui.fluent import (
    BodyLabel as QLabel,
    CardWidget,
    ComboBox as QComboBox,
    FluentIcon,
    LineEdit as QLineEdit,
    ListWidget as QListWidget,
    NavigationInterface,
    NavigationItemPosition,
    PrimaryPushButton,
    ProgressBar as QProgressBar,
    PushButton as QPushButton,
    StrongBodyLabel,
    SubtitleLabel,
    TableWidget as QTableWidget,
    TextEdit as QTextEdit,
    TitleLabel,
    set_fluent_theme,
)
from office_app.ui.theme import (
    get_active_theme_tokens,
    set_active_theme,
    set_large_text,
    theme_color,
)
from office_app.ui.configuration_dialog import (
    ConfigurationDialog,
    friendly_connection_error,
)
from office_app.ui.views import StudentListView
from office_app.ui.views.settings_figma_view import SettingsView
from office_app.services.updater_service import UpdaterService

def app_version_label() -> str:
    return f"SSM v{UpdaterService.CURRENT_VERSION} - YWAM Balut"


def render_qss(template: str) -> str:
    """Expand @token references because Qt QSS does not support variables."""
    # The bundled variable file registers with Qt under its internal family
    # name, "Inter Variable". Figma calls the same face "Inter".
    template = template.replace(
        'font-family: "Inter",',
        'font-family: "Inter Variable",',
    )
    dropdown_icon = resource_path(
        os.path.join("assets", "icons", "chevron-down.svg")
    ).replace("\\", "/")
    sidebar_dropdown_icon = resource_path(
        os.path.join("assets", "icons", "chevron-down-sidebar.svg")
    ).replace("\\", "/")
    template = template.replace("@dropdown_icon", dropdown_icon)
    template = template.replace(
        "@sidebar_dropdown_icon", sidebar_dropdown_icon
    )
    tokens = get_active_theme_tokens()
    # Replace longer names first so @primary does not partially consume
    # @primary_hover, @primary_pressed, and similar prefixed tokens.
    for name in sorted(tokens, key=len, reverse=True):
        value = tokens[name]
        template = template.replace(f"@{name}", value)
    return template


_APP_FONTS_REGISTERED = False


def register_app_fonts() -> None:
    """Register the bundled Figma typeface once for source and frozen builds."""
    global _APP_FONTS_REGISTERED
    if _APP_FONTS_REGISTERED:
        return
    for file_name in (
        "InterVariable.ttf",
        "NotoSansSymbols2-Regular.ttf",
    ):
        font_path = resource_path(os.path.join("assets", "fonts", file_name))
        if QFontDatabase.addApplicationFont(font_path) < 0:
            logging.getLogger(__name__).warning(
                "Could not register bundled font at %s", font_path
            )
    _APP_FONTS_REGISTERED = True


def css_color(name: str, alpha: int | None = None) -> str:
    color = theme_color(name, alpha)
    return f"rgba({color.red()}, {color.green()}, {color.blue()}, {color.alpha()})"


def set_on_brand_text(label, *, alpha: int | None = None) -> None:
    label.setStyleSheet(
        f"color: {css_color('on_brand', alpha)}; background: transparent;"
    )





def user_initials(name: str) -> str:
    parts = [part for part in name.split() if part]
    return "".join(part[0].upper() for part in parts[:2]) or name[:1].upper()


def avatar_icon(name: str, *, selected: bool, size: int = 36) -> QIcon:
    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    margin = 2
    bounds = QRectF(margin, margin, size - margin * 2, size - margin * 2)
    border = theme_color("primary_selected") if selected else theme_color("border")
    painter.setPen(QPen(border, 2))
    painter.setBrush(theme_color("primary_soft") if selected else theme_color("surface_subtle"))
    painter.drawEllipse(bounds)
    painter.setPen(theme_color("primary_pressed") if selected else theme_color("text_secondary"))
    font = QFont("Segoe UI", 9)
    font.setBold(True)
    painter.setFont(font)
    painter.drawText(bounds, Qt.AlignmentFlag.AlignCenter, user_initials(name))
    painter.end()
    return QIcon(pixmap)


def sidebar_operator_pixmap(name: str, *, size: int = 30) -> QPixmap:
    """Render the audit operator as a compact, legible sidebar squircle."""
    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    bounds = QRectF(1, 1, size - 2, size - 2)
    painter.setPen(QPen(theme_color("accent", 200), 1))
    painter.setBrush(theme_color("accent"))
    painter.drawRoundedRect(bounds, 8, 8)
    painter.setPen(theme_color("sidebar"))
    font = QFont("Segoe UI", 9)
    font.setBold(True)
    painter.setFont(font)
    painter.drawText(bounds, Qt.AlignmentFlag.AlignCenter, user_initials(name))
    painter.end()
    return pixmap


def startup_operator_pixmap(name: str, *, selected: bool, size: int = 32) -> QPixmap:
    """Render the compact rounded-square operator mark used in Figma startup rows."""
    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    bounds = QRectF(0, 0, size, size)
    painter.setPen(Qt.PenStyle.NoPen)
    painter.setBrush(
        theme_color("accent") if selected else theme_color("primary_soft")
    )
    painter.drawRoundedRect(bounds, 8, 8)
    painter.setPen(
        theme_color("primary") if selected else theme_color("success")
    )
    font = QFont("Segoe UI", 9)
    font.setBold(True)
    painter.setFont(font)
    painter.drawText(bounds, Qt.AlignmentFlag.AlignCenter, user_initials(name))
    painter.end()
    return pixmap


def sidebar_navigation_icon(symbol: str) -> QIcon:
    """Render the exact symbol glyphs used by the Figma navigation rail."""
    icon = QIcon()
    for selected, color in (
        (False, theme_color("sidebar_muted")),
        (True, theme_color("sidebar_text")),
    ):
        pixmap = QPixmap(20, 20)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setPen(color)
        font = QFont("Noto Sans Symbols 2")
        font.setPixelSize(16)
        font.setWeight(QFont.Weight.Medium)
        painter.setFont(font)
        painter.drawText(
            QRectF(0, -2, 20, 20),
            Qt.AlignmentFlag.AlignCenter,
            symbol,
        )
        painter.end()
        icon.addPixmap(
            pixmap,
            QIcon.Mode.Normal,
            QIcon.State.On if selected else QIcon.State.Off,
        )
    return icon


def trim_transparent_pixmap(pixmap, alpha_threshold=8):
    image = pixmap.toImage().convertToFormat(QImage.Format.Format_ARGB32)
    left, top = image.width(), image.height()
    right, bottom = -1, -1
    for y in range(image.height()):
        for x in range(image.width()):
            if image.pixelColor(x, y).alpha() > alpha_threshold:
                left = min(left, x)
                top = min(top, y)
                right = max(right, x)
                bottom = max(bottom, y)
    if right < left or bottom < top:
        return pixmap
    return pixmap.copy(left, top, right - left + 1, bottom - top + 1)


def set_logo_pixmap(label, width, height, fallback_text="\u271d"):
    # Draw the compact Figma mark at its final size so the SSM monogram stays
    # crisp in both the navigation rail and the startup chooser.
    pix = QPixmap(width, height)
    pix.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pix)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    diameter = min(width, height)
    bounds = QRectF((width - diameter) / 2, (height - diameter) / 2, diameter, diameter)
    painter.setPen(Qt.PenStyle.NoPen)
    painter.setBrush(theme_color("accent"))
    painter.drawEllipse(bounds)
    painter.setPen(theme_color("primary"))
    font = QFont("Inter Variable", max(6, int(diameter * 0.21)))
    font.setWeight(QFont.Weight.Bold)
    painter.setFont(font)
    painter.drawText(bounds, Qt.AlignmentFlag.AlignCenter, "SSM")
    painter.end()
    label.setAlignment(Qt.AlignmentFlag.AlignCenter)
    label.setFixedSize(width, height)
    label.setPixmap(pix)
    label.setObjectName("BrandLogo")


# ── SIGNALS (used to safely call UI from background threads) ──────────────────
class WorkerSignals(QObject):
    connected = pyqtSignal()
    failed    = pyqtSignal(str)


class AnimatedSplashPanel(QWidget):
    """Paint a softly moving brand gradient behind the splash content."""

    def __init__(self, parent=None, *, animate=True):
        super().__init__(parent)
        self._phase = 0.0
        self._background_animation = None
        if animate:
            self._background_animation = QPropertyAnimation(
                self, b"phase", self
            )
            self._background_animation.setDuration(12000)
            self._background_animation.setKeyValueAt(0.0, 0.0)
            self._background_animation.setKeyValueAt(0.5, 1.0)
            self._background_animation.setKeyValueAt(1.0, 0.0)
            self._background_animation.setEasingCurve(
                QEasingCurve.Type.InOutSine
            )
            self._background_animation.setLoopCount(-1)
            self._background_animation.start()

    def get_phase(self):
        return self._phase

    def set_phase(self, value):
        self._phase = float(value)
        self.update()

    phase = pyqtProperty(float, get_phase, set_phase)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        bounds = QRectF(self.rect()).adjusted(1, 1, -1, -1)
        clip = QPainterPath()
        clip.addRoundedRect(bounds, 18, 18)
        painter.setClipPath(clip)

        width = max(1, self.width())
        height = max(1, self.height())
        drift = (self._phase - 0.5) * width * 0.16
        gradient = QLinearGradient(
            QPointF(drift, 0), QPointF(width + drift, height)
        )
        gradient.setColorAt(0.0, theme_color("splash_start"))
        gradient.setColorAt(0.52, theme_color("primary"))
        gradient.setColorAt(1.0, theme_color("splash_end"))
        painter.fillPath(clip, gradient)

        bloom_x = width * (0.32 + 0.36 * self._phase)
        bloom = QRadialGradient(QPointF(bloom_x, height * 0.16), width * 0.58)
        bloom.setColorAt(0.0, theme_color("on_brand", 22))
        bloom.setColorAt(0.48, theme_color("on_brand", 8))
        bloom.setColorAt(1.0, theme_color("on_brand", 0))
        painter.fillPath(clip, bloom)

        painter.setClipping(False)
        painter.setPen(QPen(theme_color("on_brand", 54), 1))
        painter.drawRoundedRect(bounds, 18, 18)


# ── STARTUP SPLASH ────────────────────────────────────────────────────────────

class StartupDialog(QDialog):
    """Two-phase splash: (1) user selection, (2) connection progress."""

    _SPLASH_SS = """
        QDialog {
            background: transparent;
        }
        QDialog QWidget#SplashPanel {
            background: @primary_hover;
            border: 1px solid @primary_selected;
            border-radius: 18px;
        }
        QDialog QWidget#SplashPage { background: transparent; }
        QDialog QLabel { background: transparent; }
        QDialog QStackedWidget { background: transparent; }
        QDialog QWidget#StartupBrand { background: transparent; }
        QDialog QFrame#StartupSurface {
            background: @surface;
            border: none;
            border-top-left-radius: 0px;
            border-bottom-left-radius: 0px;
            border-top-right-radius: 18px;
            border-bottom-right-radius: 18px;
        }
        QDialog QFrame#StartupSurface QLabel { color: @text_primary; }
        QDialog *#StartupEyebrow {
            color: @success;
            font-family: "Segoe UI", sans-serif;
            font-size: 9px;
            font-weight: 700;
        }
        QDialog *#StartupTitle {
            color: @text_primary;
            font-family: "Segoe UI", sans-serif;
            font-size: 22px;
            font-weight: 700;
        }
        QDialog *#StartupDescription {
            color: @text_secondary;
            font-size: 11px;
            font-weight: 400;
        }
        QDialog QFrame#LoadingProfile {
            background: @primary_soft;
            border: 1px solid @border;
            border-radius: 9px;
        }
        QDialog *#LoadingProfileName { color: @text_primary; font-size: 13px; font-weight: 600; }
        QDialog *#LoadingProfileMeta { color: @text_secondary; font-size: 10px; }
        QDialog *#StartupStep { color: @disabled; font-size: 11px; font-weight: 400; }
        QDialog *#StartupStep[state="active"] { color: @primary; font-weight: 650; }
        QDialog *#StartupStep[state="complete"] { color: @success; font-weight: 600; }
        QDialog *#StartupStep[state="danger"] { color: @danger; }
        QDialog *#SplashOrg {
            color: #BBD3CA;
            font-size: 10px;
            font-weight: 600;
        }
        QDialog *#SplashTitle {
            color: white;
            font-size: 22px;
            font-weight: 700;
        }
        QDialog *#SplashPrompt {
            color: #C4D9D1;
            font-size: 12px;
            font-weight: 400;
        }
        QDialog *#SplashVersion {
            color: #8FC2B2;
            font-size: 10px;
            font-weight: 500;
        }
        QDialog QFrame#StartupReleaseLog {
            background: transparent;
            border: none;
        }
        QDialog QFrame#StartupReleaseDivider {
            background: rgba(255, 255, 255, 0.18);
            border: none;
        }
        QDialog *#StartupReleaseEyebrow {
            color: #F6C90E;
            font-size: 8px;
            font-weight: 700;
        }
        QDialog *#StartupReleaseItem {
            color: #D9E9E3;
            font-size: 9px;
            font-weight: 400;
        }
        QDialog *#StartupLogoText {
            color: @primary;
            font-size: 12px;
            font-weight: 700;
        }
        QDialog *#SplashStatus {
            color: @success;
            font-size: 10px;
            font-weight: 600;
        }
        QDialog *#SplashProgressValue {
            color: @text_primary;
            font-size: 13px;
            font-weight: 700;
        }
        QDialog *#StartupMotionNote {
            color: @disabled;
            font-size: 8px;
            font-weight: 500;
        }
        QDialog QFrame#StartupSurface QProgressBar {
            border: none;
            border-radius: 4px;
            background: @border_subtle;
            max-height: 8px;
        }
        QDialog QFrame#StartupSurface QProgressBar::chunk { background: @success; border-radius: 4px; }

        QDialog QPushButton#ContinueBtn {
            background: @primary;
            color: @on_brand;
            border: none;
            border-radius: 8px;
            padding: 0px;
            font-size: 12px;
            font-weight: 600;
        }
        QDialog QPushButton#ContinueBtn:hover { background: @primary_hover; color: @on_brand; }
        QDialog QPushButton#ContinueBtn:pressed { background: @primary_pressed; }

        QDialog QPushButton#RetryBtn {
            background: @primary; color: @on_brand; border: none;
            border-radius: 7px; padding: 7px 14px;
            font-weight: 700; font-size: 12px;
        }
        QDialog QPushButton#RetryBtn:hover { background: @primary_hover; }
        QDialog QPushButton#OfflineBtn {
            background: @surface; color: @text_primary;
            border: 1px solid @border;
            border-radius: 7px; padding: 7px 12px;
            font-weight: 650; font-size: 12px;
        }
        QDialog QPushButton#OfflineBtn:hover { background: @surface_subtle; border-color: @primary; }
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._reduce_motion = QSettings(
            "YWAMBalut", "SSMStudentProfilingSystem"
        ).value("reduce_motion", False, type=bool)
        self.setWindowTitle("YWAM Balut SSM")
        icon_path = resource_path(APP_ICON_ASSET)
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.setFixedSize(640, 420)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.CustomizeWindowHint |
            Qt.WindowType.FramelessWindowHint
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        # Apply stylesheet AFTER setting flags; prepend * reset to override app-level QPushButton
        ss = "* { font-family: 'Segoe UI', Arial, sans-serif; }\n" + render_qss(self._SPLASH_SS)
        self.setStyleSheet(ss)

        self.success = False
        self.error_msg = ""
        self.selected_user = USERS[0]
        self._welcome_anim = None
        self._progress_anim = None
        self._dot_timer = None
        self._dot_count = 0
        self._loading_status_base = "Connecting to database"
        self._pending_sb = None
        self.recovery_action = ""
        self._failure_actions = None
        self._startup_release_logs = []

        self._signals = WorkerSignals()
        self._signals.connected.connect(self._on_connected)
        self._signals.failed.connect(self._on_failed)

        # ── Root stacked layout (page 0 = user select, page 1 = connecting) ──
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        panel = QWidget()
        panel.setObjectName("SplashPanel")

        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(0, 0, 0, 0)
        self._stack = QStackedWidget()
        panel_layout.addWidget(self._stack)
        root.addWidget(panel)

        self._stack.addWidget(self._build_user_page())
        self._stack.addWidget(self._build_connect_page())
        self._stack.setCurrentIndex(0)
        self._apply_rounded_window_mask()
        self.setWindowOpacity(1.0 if self._reduce_motion else 0.0)
        if not self._reduce_motion:
            QTimer.singleShot(80, self._start_splash_entrance)

    def _apply_rounded_window_mask(self):
        path = QPainterPath()
        path.addRoundedRect(QRectF(self.rect()), 18, 18)
        self.setMask(QRegion(path.toFillPolygon().toPolygon()))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._apply_rounded_window_mask()

    # ── Page 1: User Selection ─────────────────────────────────────────────────
    def _build_startup_brand(self, *, loading=False):
        brand = QWidget()
        brand.setObjectName("StartupBrand")
        brand.setFixedWidth(218)
        layout = QVBoxLayout(brand)
        layout.setContentsMargins(30, 32, 30, 18)
        layout.setSpacing(0)

        logo_mark = QWidget()
        logo_mark.setFixedSize(52, 52)
        mark_image = NativeLabel(logo_mark)
        mark_image.setGeometry(0, 0, 52, 52)
        mark_pixmap = QPixmap(resource_path(os.path.join("assets", "ssm_startup_mark.png")))
        if not mark_pixmap.isNull():
            mark_image.setPixmap(
                mark_pixmap.scaled(
                    52,
                    52,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
            )
        logo_text = NativeLabel("SSM", logo_mark)
        logo_text.setObjectName("StartupLogoText")
        logo_text.setGeometry(0, 0, 52, 52)
        logo_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(logo_mark, alignment=Qt.AlignmentFlag.AlignLeft)
        layout.addSpacing(27)

        org = NativeLabel("YWAM BALUT")
        org.setObjectName("SplashOrg")
        layout.addWidget(org)
        layout.addSpacing(14)

        title = NativeLabel("Student Support\nManagement")
        title.setObjectName("SplashTitle")
        title.setWordWrap(True)
        layout.addWidget(title)
        layout.addSpacing(12)

        caption_text = (
            "Opening the office workspace and checking synchronized services."
            if loading
            else "A focused workspace for student records, attendance, and office follow-up."
        )
        caption = NativeLabel(caption_text)
        caption.setObjectName("SplashPrompt")
        caption.setWordWrap(True)
        layout.addWidget(caption)
        layout.addSpacing(20)

        release_log = QFrame()
        release_log.setObjectName("StartupReleaseLog")
        release_log.setAccessibleName(
            f"What changed in SSM version {UpdaterService.CURRENT_VERSION}"
        )
        release_log.setAccessibleDescription(
            ". ".join(UpdaterService.CURRENT_RELEASE_NOTES)
        )
        release_layout = QVBoxLayout(release_log)
        release_layout.setContentsMargins(0, 0, 0, 0)
        release_layout.setSpacing(4)

        release_divider = QFrame()
        release_divider.setObjectName("StartupReleaseDivider")
        release_divider.setFixedHeight(1)
        release_layout.addWidget(release_divider)
        release_layout.addSpacing(7)

        release_eyebrow = NativeLabel(
            f"WHAT CHANGED  ·  v{UpdaterService.CURRENT_VERSION}"
        )
        release_eyebrow.setObjectName("StartupReleaseEyebrow")
        release_layout.addWidget(release_eyebrow)

        for note in UpdaterService.CURRENT_RELEASE_NOTES:
            note_label = NativeLabel(f"•  {note}")
            note_label.setObjectName("StartupReleaseItem")
            note_label.setWordWrap(True)
            release_layout.addWidget(note_label)

        layout.addWidget(release_log)
        self._startup_release_logs.append(release_log)
        layout.addStretch(1)

        version = NativeLabel(f"SSM v{UpdaterService.CURRENT_VERSION}")
        version.setObjectName("SplashVersion")
        layout.addWidget(version)
        return brand

    def _build_user_page(self):
        page = QWidget()
        page.setObjectName("SplashPage")
        page_layout = QHBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)
        page_layout.addWidget(self._build_startup_brand())

        surface = QFrame()
        surface.setObjectName("StartupSurface")
        surface.setFixedWidth(422)
        surface_layout = QVBoxLayout(surface)
        surface_layout.setContentsMargins(34, 32, 34, 18)
        surface_layout.setSpacing(0)

        eyebrow = QLabel("CHOOSE OPERATOR")
        eyebrow.setObjectName("StartupEyebrow")
        surface_layout.addWidget(eyebrow)
        surface_layout.addSpacing(9)

        title = QLabel("Who is making changes?")
        title.setObjectName("StartupTitle")
        surface_layout.addWidget(title)
        surface_layout.addSpacing(4)

        prompt = QLabel("This identifies the person recorded in the audit history.")
        prompt.setObjectName("StartupDescription")
        prompt.setWordWrap(True)
        surface_layout.addWidget(prompt)
        surface_layout.addSpacing(30)

        self._user_cards = []
        for name in USERS:
            card = QPushButton()
            card.setObjectName("UserCard")
            card.setProperty("user_name", name)
            card.setCheckable(True)
            card.setCursor(Qt.CursorShape.PointingHandCursor)
            card.setAccessibleName(f"Use office profile {name}")
            card.setAccessibleDescription(
                "Selects the name recorded in the activity log."
            )
            card.setFixedHeight(52)
            card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            card.clicked.connect(
                lambda _checked=False, n=name: self._select_user(n)
            )

            card_layout = QHBoxLayout(card)
            card_layout.setContentsMargins(12, 9, 12, 9)
            card_layout.setSpacing(14)

            avatar_lbl = QLabel()
            avatar_lbl.setObjectName("UserAvatar")
            avatar_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            avatar_lbl.setFixedSize(32, 32)

            name_copy = QVBoxLayout()
            name_copy.setContentsMargins(0, 0, 0, 0)
            name_copy.setSpacing(0)

            name_lbl = QLabel(name)
            name_lbl.setObjectName("UserName")
            name_lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

            meta_lbl = QLabel(f"Changes recorded as {name}")
            meta_lbl.setObjectName("UserMeta")
            meta_lbl.setStyleSheet(
                f"color: {css_color('text_secondary')}; background: transparent; "
                "font-size: 10px; font-weight: 400;"
            )
            name_copy.addWidget(name_lbl)
            name_copy.addWidget(meta_lbl)

            state_lbl = QLabel()
            state_lbl.setObjectName("UserState")
            state_lbl.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight)
            state_lbl.setMinimumWidth(58)

            for clickable in (avatar_lbl, name_lbl, meta_lbl, state_lbl):
                clickable.setCursor(Qt.CursorShape.PointingHandCursor)
                clickable.mousePressEvent = lambda event, n=name: self._select_user(n)

            card._avatar_label = avatar_lbl
            card._name_label = name_lbl
            card._meta_label = meta_lbl
            card._state_label = state_lbl
            card_layout.addWidget(avatar_lbl)
            card_layout.addLayout(name_copy, 1)
            card_layout.addWidget(state_lbl)

            surface_layout.addWidget(card)
            surface_layout.addSpacing(12)
            self._user_cards.append(card)

        surface_layout.addSpacing(12)
        continue_btn = QPushButton()
        self._continue_btn = continue_btn
        continue_btn.setObjectName("ContinueBtn")
        continue_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        continue_btn.setAccessibleDescription("Continues to the secure workspace startup check.")
        continue_btn.setFixedHeight(40)
        continue_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        continue_btn.setDefault(True)
        continue_btn.setStyleSheet(self._continue_button_style())
        continue_btn.clicked.connect(self._on_continue)
        surface_layout.addWidget(continue_btn)
        surface_layout.addStretch(1)

        motion_note = QLabel("Motion: 180 ms fade-in  •  selected row uses 140 ms color transition")
        motion_note.setObjectName("StartupMotionNote")
        surface_layout.addWidget(motion_note)

        page_layout.addWidget(surface, 1)
        if self._user_cards:
            for current, following in zip(self._user_cards, self._user_cards[1:]):
                QWidget.setTabOrder(current, following)
            QWidget.setTabOrder(self._user_cards[-1], continue_btn)
        self._select_user(self.selected_user)
        return page

    def _set_loading_phase(self, phase):
        states = {
            "version": ("active", "pending", "pending", "pending"),
            "database": ("complete", "active", "pending", "pending"),
            "records": ("complete", "complete", "active", "pending"),
            "workspace": ("complete", "complete", "complete", "active"),
            "complete": ("complete", "complete", "complete", "complete"),
            "failed": ("complete", "danger", "pending", "pending"),
        }.get(phase, ("pending", "pending", "pending", "pending"))
        copy = (
            {
                "pending": "Checking application version",
                "active": "Checking application version",
                "complete": "Application version checked",
                "danger": "Version check unavailable",
            },
            {
                "pending": "Connecting to office database",
                "active": "Connecting to office database",
                "complete": "Office database connected",
                "danger": "Database connection unavailable",
            },
            {
                "pending": "Loading student records",
                "active": "Loading student records",
                "complete": "Student records ready",
            },
            {
                "pending": "Preparing dashboard",
                "active": "Preparing dashboard",
                "complete": "Dashboard ready",
            },
        )
        glyphs = {
            "complete": "✓",
            "active": "●",
            "pending": " ",
            "danger": "×",
        }
        for label, state, labels in zip(self._loading_step_labels, states, copy):
            previous_state = label.property("state")
            label.setProperty("state", state)
            label.setText(f"{glyphs[state]}     {labels.get(state, labels['pending'])}")
            label.setAccessibleDescription(labels.get(state, labels["pending"]))
            label.style().unpolish(label)
            label.style().polish(label)
            if previous_state != state and state in {"active", "complete"}:
                fade_in(
                    label,
                    motion_enabled=not self._reduce_motion,
                    duration_ms=180,
                )

    def _start_splash_entrance(self):
        """Start compositor-safe splash motion without child paint effects."""
        self._window_entrance = QPropertyAnimation(self, b"windowOpacity", self)
        self._window_entrance.setDuration(180)
        self._window_entrance.setStartValue(0.0)
        self._window_entrance.setEndValue(1.0)
        self._window_entrance.setEasingCurve(QEasingCurve.Type.OutCubic)
        self._window_entrance.start()

    def _select_user(self, name):
        self.selected_user = name
        for card in self._user_cards:
            card_name = card.property("user_name")
            selected = card_name == name
            card.setProperty("selected", selected)
            card.setChecked(selected)
            card.setStyleSheet(self._user_card_style(selected))
            card._avatar_label.setPixmap(
                startup_operator_pixmap(card_name, selected=selected, size=32)
            )
            card._name_label.setStyleSheet(self._user_name_style(selected))
            card._state_label.setText("Selected" if selected else "Choose")
            card._state_label.setStyleSheet(self._user_state_style(selected))
        if hasattr(self, "_continue_btn"):
            self._continue_btn.setText(f"Continue as {name}")
        if hasattr(self, "loading_user_name"):
            self.loading_user_name.setText(name)
            self.loading_user_avatar.setPixmap(startup_operator_pixmap(name, selected=True))
        if hasattr(self, "loading_title"):
            self.loading_title.setText(f"Welcome, {name}")

    def _user_card_style(self, selected: bool) -> str:
        if selected:
            return f"""
                QPushButton#UserCard {{
                    background: {css_color('primary_soft')};
                    border: 1px solid {css_color('success')};
                    border-radius: 9px;
                }}
            """
        return f"""
            QPushButton#UserCard {{
                background: {css_color('surface')};
                border: 1px solid {css_color('border')};
                border-radius: 9px;
            }}
            QPushButton#UserCard:hover {{
                background: {css_color('surface_subtle')};
                border-color: {css_color('primary_selected')};
            }}
        """

    def _user_name_style(self, selected: bool) -> str:
        color = css_color("text_primary")
        return f"""
            color: {color};
            background: transparent;
            font-size: 13px;
            font-weight: 700;
        """

    def _user_state_style(self, selected: bool) -> str:
        color = css_color("primary") if selected else css_color("text_secondary")
        return f"""
            color: {color};
            background: transparent;
            font-size: 10px;
            font-weight: 700;
        """

    def _continue_button_style(self) -> str:
        return f"""
            QPushButton#ContinueBtn {{
                background: {css_color('primary')};
                color: {css_color('on_brand')};
                border: none;
                border-radius: 8px;
                font-weight: 700;
                padding: 0 14px;
            }}
            QPushButton#ContinueBtn:hover {{
                background: {css_color('primary_hover')};
                color: {css_color('on_brand')};
            }}
            QPushButton#ContinueBtn:pressed {{
                background: {css_color('primary_pressed')};
            }}
        """

    def _on_continue(self):
        self._select_user(self.selected_user)
        self._stack.setCurrentIndex(1)
        fade_in(
            self.loading_title,
            motion_enabled=not self._reduce_motion,
            duration_ms=180,
        )
        if self._pending_sb is not None:
            self.start_ping(self._pending_sb)

    def queue_ping(self, sb: Client):
        """Call this before exec() so the ping fires after the user clicks Continue."""
        self._pending_sb = sb

    # ── Page 2: Connecting ─────────────────────────────────────────────────────
    def _build_connect_page(self):
        page = QWidget()
        page.setObjectName("SplashPage")
        page_layout = QHBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)
        page_layout.addWidget(self._build_startup_brand(loading=True))

        surface = QFrame()
        surface.setObjectName("StartupSurface")
        surface.setFixedWidth(422)
        lay = QVBoxLayout(surface)
        lay.setContentsMargins(34, 32, 34, 18)
        lay.setSpacing(0)

        eyebrow = QLabel("OPENING WORKSPACE")
        eyebrow.setObjectName("StartupEyebrow")
        lay.addWidget(eyebrow)
        lay.addSpacing(9)

        self.loading_title = QLabel(f"Welcome, {self.selected_user}")
        self.loading_title.setObjectName("StartupTitle")
        lay.addWidget(self.loading_title)
        lay.addSpacing(4)

        description = QLabel("Preparing your student-support workspace.")
        description.setObjectName("StartupDescription")
        lay.addWidget(description)
        lay.addSpacing(26)

        profile = QFrame()
        profile.setObjectName("LoadingProfile")
        profile.setFixedHeight(56)
        profile_layout = QHBoxLayout(profile)
        profile_layout.setContentsMargins(14, 11, 14, 11)
        profile_layout.setSpacing(14)
        self.loading_user_avatar = QLabel()
        self.loading_user_avatar.setFixedSize(32, 32)
        self.loading_user_avatar.setPixmap(startup_operator_pixmap(self.selected_user, selected=True))
        profile_copy = QVBoxLayout()
        profile_copy.setContentsMargins(0, 0, 0, 0)
        profile_copy.setSpacing(0)
        self.loading_user_name = QLabel(self.selected_user)
        self.loading_user_name.setObjectName("LoadingProfileName")
        profile_meta = QLabel("Changes will be recorded under this operator.")
        profile_meta.setObjectName("LoadingProfileMeta")
        profile_copy.addWidget(self.loading_user_name)
        profile_copy.addWidget(profile_meta)
        profile_layout.addWidget(self.loading_user_avatar)
        profile_layout.addLayout(profile_copy, 1)
        lay.addWidget(profile)
        lay.addSpacing(22)

        progress_header = QHBoxLayout()
        progress_header.setContentsMargins(0, 0, 0, 0)
        progress_header.setSpacing(8)
        self.progress_value_label = QLabel("0%")
        self.progress_value_label.setObjectName("SplashProgressValue")
        progress_header.addWidget(self.progress_value_label)
        progress_header.addStretch(1)

        self.status_label = QLabel("Checking database connection")
        self.status_label.setObjectName("SplashStatus")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        progress_header.addWidget(self.status_label)
        lay.addLayout(progress_header)
        lay.addSpacing(7)

        self.progress = NativeProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setTextVisible(False)
        self.progress.setFixedHeight(8)
        lay.addWidget(self.progress)
        lay.addSpacing(17)

        self._loading_step_labels = []
        for step in (
            "Checking database connection",
            "Loading application settings",
            "Loading student records",
            "Preparing dashboard",
        ):
            label = QLabel(step)
            label.setObjectName("StartupStep")
            label.setProperty("state", "pending")
            label.setFixedHeight(27)
            lay.addWidget(label)
            self._loading_step_labels.append(label)

        lay.addStretch(1)

        self._connect_actions = QWidget()
        self._connect_actions.setVisible(False)
        self._connect_actions_layout = QHBoxLayout(self._connect_actions)
        self._connect_actions_layout.setContentsMargins(0, 0, 0, 0)
        self._connect_actions_layout.setSpacing(8)
        lay.addWidget(self._connect_actions)

        motion_note = QLabel(
            "Four guided checks  •  short ease-out transitions  •  reduced-motion ready"
        )
        motion_note.setObjectName("StartupMotionNote")
        lay.addWidget(motion_note)

        page_layout.addWidget(surface, 1)
        self._set_loading_phase("version")

        return page
    # ── Connection logic ───────────────────────────────────────────────────────
    def start_ping(self, sb: Client):
        self.loading_title.setText(f"Welcome, {self.selected_user}")
        self.status_label.setStyleSheet("")
        self._set_loading_phase("version")
        self._set_loading_status("Checking application version")
        self._set_loading_progress(15)
        self._pending_sb = sb  # keep ref so _on_update_checked can use it

        # --- AUTO UPDATE CHECK (non-blocking) ---
        # check_for_update() is a DB call, so it must NOT run on the main thread.
        # We use BackgroundTask here and continue to normal boot in the callback
        # only if no update is found.
        self.updater = UpdaterService(sb)
        update_task = BackgroundTask(lambda: self.updater.check_for_update())
        update_task.signals.succeeded.connect(self._on_update_checked)
        update_task.signals.failed.connect(
            # If the update check itself errors, just proceed with normal boot
            lambda _err: self._resume_boot_after_version_check_error()
        )
        QThreadPool.globalInstance().start(update_task)

    def _on_update_checked(self, update_info):
        """Called on the main thread once the background update check completes."""
        if not update_info:
            self._show_version_complete_then_connect("Version is up to date")
            return

        latest_version, url = update_info
        if self._dot_timer:
            self._dot_timer.stop()
        self.status_label.setText(f"Downloading update {latest_version}")
        self._set_loading_progress(50)

        # Start the download (safely pushing UI updates back to main thread)
        self.updater.download_and_install(
            url=url,
            progress_callback=lambda p: QTimer.singleShot(0, lambda: self._set_loading_progress(50 + int(p * 0.45), animate=False)),
            success_callback=lambda: QTimer.singleShot(0, lambda: self.status_label.setText("Restarting to apply update")),
            error_callback=lambda err: QTimer.singleShot(0, lambda: self._resume_boot_after_update_error(err))
        )
        # Do NOT start normal boot — we are updating and will os._exit()

    def _resume_boot_after_update_error(self, error):
        """Continue startup if an update cannot be installed."""
        self._set_loading_phase("database")
        self._set_loading_status("Checking database connection")
        self._set_loading_progress(42)
        self._start_normal_boot(self._pending_sb)

    def _resume_boot_after_version_check_error(self):
        """Continue startup when the version check cannot complete."""
        self._show_version_complete_then_connect("Version check skipped")

    def _show_version_complete_then_connect(self, message):
        if self._dot_timer:
            self._dot_timer.stop()
        self._set_loading_progress(50)
        self.status_label.setText(message)
        QTimer.singleShot(
            120 if self._reduce_motion else 520,
            self._continue_after_version_check,
        )

    def _continue_after_version_check(self):
        self._set_loading_phase("database")
        self._set_loading_status("Connecting to office database")
        self._set_loading_progress(42)
        self._start_normal_boot(self._pending_sb)

    def _start_normal_boot(self, sb):
        """Kick off the normal DB ping after the update check finds nothing."""
        task = BackgroundTask(lambda: StudentRepository(client=sb).ping())
        task.signals.succeeded.connect(lambda _rows: self._signals.connected.emit())
        task.signals.failed.connect(
            lambda error: self._signals.failed.emit(error.strip().splitlines()[-1])
        )
        QThreadPool.globalInstance().start(task)

    def _start_dot_anim(self):
        if self._dot_timer:
            self._dot_timer.stop()
        if self._reduce_motion:
            return
        self._dot_count = 0
        self._dot_timer = QTimer(self)
        self._dot_timer.setInterval(420)
        self._dot_timer.timeout.connect(self._tick_dots)
        self._dot_timer.start()

    def _set_loading_status(self, message):
        self._loading_status_base = message
        self.status_label.setText(message)

    def _set_loading_progress(self, value, animate=True):
        value = max(0, min(100, int(value)))
        self.progress_value_label.setText(f"{value}%")
        self.progress.setRange(0, 100)
        if not animate or self._reduce_motion:
            self.progress.setValue(value)
            return
        animation = QPropertyAnimation(self.progress, b"value", self)
        animation.setDuration(260)
        animation.setStartValue(self.progress.value())
        animation.setEndValue(value)
        animation.setEasingCurve(QEasingCurve.Type.OutCubic)
        self._phase_progress_anim = animation
        animation.start()

    def _tick_dots(self):
        self._dot_count = (self._dot_count + 1) % 4

    def _on_connected(self):
        if self._dot_timer:
            self._dot_timer.stop()
        self.success = True
        self.loading_title.setText(f"Welcome, {self.selected_user}")

        if self._reduce_motion:
            self._set_loading_phase("complete")
            self.status_label.setText("Workspace ready")
            self.progress.setValue(100)
            self.progress_value_label.setText("100%")
            QTimer.singleShot(250, self.accept)
            return

        self._set_loading_phase("records")
        self._set_loading_status("Loading student records")
        self._set_loading_progress(68)
        QTimer.singleShot(420, self._show_dashboard_startup_step)

    def _show_dashboard_startup_step(self):
        self._set_loading_phase("workspace")
        self._set_loading_status("Preparing dashboard")
        self._set_loading_progress(88)
        QTimer.singleShot(420, self._finish_startup_sequence)

    def _finish_startup_sequence(self):
        self._set_loading_phase("complete")
        self.status_label.setText("Workspace ready")
        self._progress_anim = QPropertyAnimation(self.progress, b"value", self)
        self._progress_anim.setDuration(240)
        self._progress_anim.setStartValue(self.progress.value())
        self._progress_anim.setEndValue(100)
        self._progress_anim.setEasingCurve(QEasingCurve.Type.OutCubic)

        self.progress_value_label.setText("100%")
        self._progress_anim.start()
        QTimer.singleShot(620, self.accept)

    def _on_failed(self, msg):
        if self._dot_timer:
            self._dot_timer.stop()
        self.error_msg = msg
        self._set_loading_phase("failed")
        self.progress.setRange(0, 1)
        self.progress.setValue(0)
        self.progress_value_label.setText("--")
        self.status_label.setText("Connection unavailable")
        self.status_label.setToolTip(friendly_connection_error(msg))
        self.status_label.setStyleSheet(f"color: {css_color('danger')};")
        if self._failure_actions is not None:
            return

        self._failure_actions = self._connect_actions
        retry_btn = QPushButton("Try again")
        settings_btn = QPushButton("Connection settings")
        exit_btn = QPushButton("Exit")
        settings_btn.setObjectName("RetryBtn")
        retry_btn.setObjectName("OfflineBtn")
        exit_btn.setObjectName("OfflineBtn")
        set_content_hugging_button(settings_btn)
        set_content_hugging_button(retry_btn)
        set_content_hugging_button(exit_btn)
        settings_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        retry_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        exit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._connect_actions_layout.addWidget(settings_btn)
        self._connect_actions_layout.addWidget(retry_btn)
        self._connect_actions_layout.addWidget(exit_btn)
        self._connect_actions.setVisible(True)
        settings_btn.clicked.connect(self._request_connection_settings)
        retry_btn.clicked.connect(self._retry_connection)
        exit_btn.clicked.connect(self._exit_after_connection_failure)

    def _retry_connection(self):
        self.error_msg = ""
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        if self._failure_actions is not None:
            while self._connect_actions_layout.count():
                item = self._connect_actions_layout.takeAt(0)
                if item.widget() is not None:
                    item.widget().deleteLater()
            self._connect_actions.setVisible(False)
            self._failure_actions = None
        self.start_ping(self._pending_sb)

    def _request_connection_settings(self):
        self.recovery_action = "settings"
        self.reject()

    def _exit_after_connection_failure(self):
        self.recovery_action = "exit"
        self.reject()

# ── MAIN APP ──────────────────────────────────────────────────────────────────
class CircularProgress(QWidget):
    def __init__(self, parent=None, size=96):
        super().__init__(parent)
        self.value = 0
        self.setFixedSize(size, size)
        self.setAccessibleName("Profile completion")
        self.setAccessibleDescription("Profile completion is 0 percent")

    def set_value(self, val):
        self.value = max(0, min(100, int(val)))
        self.setAccessibleDescription(
            f"Profile completion is {self.value} percent"
        )
        self.update()

    def get_value(self):
        return self.value

    progressValue = pyqtProperty(int, get_value, set_value)

    def animate_to(self, val, *, motion_enabled=True, duration_ms=260):
        """Animate progress without delaying the synchronized final value."""
        target = max(0, min(100, int(val)))
        previous = getattr(self, "_value_animation", None)
        if previous is not None:
            previous.stop()
        if not motion_enabled or self.value == target:
            self.set_value(target)
            return
        animation = QPropertyAnimation(self, b"progressValue", self)
        animation.setDuration(min(320, max(120, int(duration_ms))))
        animation.setStartValue(self.value)
        animation.setEndValue(target)
        animation.setEasingCurve(QEasingCurve.Type.OutCubic)
        self._value_animation = animation
        animation.start()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        pen_width = max(4, int(self.width() * 0.07))
        margin = pen_width + 2
        rect = QRectF(margin, margin, self.width() - (margin * 2), self.height() - (margin * 2))
        
        # Draw background track
        pen_bg = QPen(theme_color("border_subtle"), pen_width)
        painter.setPen(pen_bg)
        painter.drawArc(rect, 0, 360 * 16)
        
        # Determine color based on completion
        if self.value == 100: color = theme_color("success")
        elif self.value >= 75: color = theme_color("primary")
        elif self.value >= 50: color = theme_color("warning")
        else: color = theme_color("danger")
        
        # Draw progress arc
        pen_fg = QPen(color, pen_width)
        pen_fg.setCapStyle(Qt.PenCapStyle.RoundCap)
        painter.setPen(pen_fg)
        span_angle = int((self.value / 100) * 360 * 16)
        painter.drawArc(rect, 90 * 16, -span_angle) # Start at top (90 degrees)
        
        # Draw percentage text
        painter.setPen(theme_color("text_primary"))
        font = painter.font()
        font.setPixelSize(max(8, int(self.width() * 0.19)))
        font.setBold(True)
        painter.setFont(font)
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, f"{self.value}%")

class StudentApp(QMainWindow):
    def __init__(self, sb: Client, initial_user: str = "Joshua"):
        super().__init__()
        register_app_fonts()
        self.sb = sb
        self.student_service = StudentService()
        self.student_repository = self.student_service.repository
        self.student_excel_service = StudentExcelService(self.student_repository)
        self.student_list_service = StudentListService(self.student_service)
        self.dashboard_service = DashboardService()
        self.photo_service = PhotoService(client=self.sb)
        self.expense_service = ExpenseService()
        self.coordinator_repository = CoordinatorRepository()
        self.audit_repository = AuditRepository()
        self.sheet_sync_service = GoogleSheetSyncService(self.audit_repository)
        self.workbook_repository = WorkbookRepository()
        self.workbook_import_service = WorkbookImportService(
            self.student_repository,
            self.coordinator_repository,
            self.workbook_repository,
        )
        self.masterlist_service = MasterListService(self.workbook_import_service)
        self.thread_pool = QThreadPool(self)
        self._initial_user = initial_user
        self._current_operator = initial_user
        self.current_student_id = None
        self._pending_photo = None
        self._editing_id = None
        self._editing_snapshot = None
        self._workbook = None
        self._workbook_path = None
        self._workbook_mtime_ns = None
        self._workbook_loading_request = 0
        self._workbook_busy = False
        self._workbook_lock = threading.RLock()
        self._loaded_workbook_sheets = set()
        self._workbook_dirty = False
        self._loading_workbook_sheet = False
        self._master_ref_cache_key = None
        self._master_ref_cache = None
        self._workbook_revision = 0
        self._dashboard_request = 0
        self._profile_request = 0
        self._student_form_request = 0
        self._expense_request = 0
        self._coordinator_request = 0
        self._dashboard_motion_cards = []
        self._dashboard_insight_headers = []
        self._dashboard_revealed = False

        self.setWindowTitle("SSM Student Profiling System")
        icon_path = resource_path(APP_ICON_ASSET)
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.resize(1100, 750)
        self.setMinimumSize(980, 680)

        # Set initial accessibility and theme preferences before styling.
        settings = self._settings()
        self._reduce_motion = settings.value(
            "reduce_motion", False, type=bool
        )
        set_large_text(settings.value("large_text", False, type=bool))
        initial_theme = settings.value("theme", "Light", type=str)
        self.change_theme(initial_theme)
        self.status_bar = QStatusBar()
        self.status_bar.setSizeGripEnabled(False)
        self.setStatusBar(self.status_bar)
        # The Figma desktop shell has no persistent footer. Existing services
        # can still post transient messages without consuming workspace space.
        self.status_bar.hide()
        self._last_database_update_at = None
        self.database_updated_label = QLabel("DB: waiting")
        self.database_updated_label.setObjectName("StatusMeta")
        self.status_bar.addPermanentWidget(self.database_updated_label)
        self.database_updated_timer = QTimer(self)
        self.database_updated_timer.setInterval(30000)
        self.database_updated_timer.timeout.connect(self._refresh_database_updated_label)
        self.database_updated_timer.start()

        # Main Layout: Sidebar + Main Content
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 1. Left Sidebar
        self.create_sidebar(main_layout)

        # 2. Main Content Area
        content_widget = QWidget()
        content_widget.setObjectName("MainContent")
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(28, 20, 28, 18)
        # This places every page body on the 96px Figma baseline.
        content_layout.setSpacing(26)

        # Workspace header: page context on the left, live office state on the right.
        header = QFrame()
        header.setObjectName("WorkspaceHeader")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(14)

        title_layout = QVBoxLayout()
        title_layout.setSpacing(3)
        today = datetime.now().strftime("%A, %B %d").replace(" 0", " ")
        self.page_eyebrow_label = QLabel(
            f"STUDENT SUPPORT OFFICE  /  {today.upper()}"
        )
        self.page_eyebrow_label.setObjectName("HeaderEyebrow")
        self.page_eyebrow_label.hide()
        self.page_title_label = TitleLabel("Dashboard")
        self.page_title_label.setObjectName("HeaderTitle")
        self.page_subtitle_label = SubtitleLabel(
            "Student support activity and records at a glance"
        )
        self.page_subtitle_label.setObjectName("HeaderSubtitle")
        title_layout.addWidget(self.page_title_label)
        title_layout.addWidget(self.page_subtitle_label)
        header_layout.addLayout(title_layout, 1)

        header_actions_widget = QWidget()
        header_actions_widget.setObjectName("HeaderActions")
        header_actions_widget.setSizePolicy(
            QSizePolicy.Policy.Maximum,
            QSizePolicy.Policy.Fixed,
        )
        header_actions = QHBoxLayout(header_actions_widget)
        header_actions.setContentsMargins(0, 0, 0, 0)
        header_actions.setSpacing(8)
        self.header_actions_layout = header_actions
        self.connection_badge = StatusBadge(
            "Connecting", state="neutral", role="ConnectionBadge"
        )
        self.connection_badge.setFixedHeight(30)
        self.connection_badge.setToolTip("Live Supabase connection status")
        self.connection_badge.hide()
        self.header_secondary_button = ActionButton("Refresh data", variant="secondary")
        self.header_secondary_button.setObjectName("HeaderSecondaryAction")
        self.header_secondary_button.clicked.connect(self._header_secondary_action_clicked)
        self.header_profile_photo_button = ActionButton(
            "Add photo", variant="secondary"
        )
        self.header_profile_photo_button.setObjectName("HeaderProfilePhotoAction")
        self.header_profile_photo_button.setAccessibleName(
            "Add or change student photo"
        )
        self.header_profile_photo_button.clicked.connect(self.change_photo)
        self.header_profile_photo_button.hide()
        self.header_profile_expenses_button = ActionButton(
            "Expenses", variant="secondary"
        )
        self.header_profile_expenses_button.setObjectName(
            "HeaderProfileExpensesAction"
        )
        self.header_profile_expenses_button.setAccessibleName(
            "Open this student's expenses"
        )
        self.header_profile_expenses_button.clicked.connect(self.nav_expenses)
        self.header_profile_expenses_button.hide()
        self.header_add_button = ActionButton("Refresh data", variant="secondary")
        self.header_add_button.setObjectName("HeaderPrimaryAction")
        self.header_add_button.clicked.connect(self._header_action_clicked)
        header_actions.addWidget(
            self.header_secondary_button,
            alignment=Qt.AlignmentFlag.AlignVCenter,
        )
        header_actions.addWidget(
            self.header_profile_photo_button,
            alignment=Qt.AlignmentFlag.AlignVCenter,
        )
        header_actions.addWidget(
            self.header_profile_expenses_button,
            alignment=Qt.AlignmentFlag.AlignVCenter,
        )
        header_actions.addWidget(
            self.header_add_button,
            alignment=Qt.AlignmentFlag.AlignVCenter,
        )
        header_layout.addWidget(
            header_actions_widget,
            alignment=Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight,
        )

        content_layout.addWidget(header)
        
        # Stacked Widget for Screens
        self.stacked_widget = QStackedWidget()
        content_layout.addWidget(self.stacked_widget)
        
        main_layout.addWidget(content_widget)

        # Build Screens
        self.create_dashboard_screen()  # Index 0
        self.student_list_view = StudentListView(
            self.student_repository,
            self.student_list_service,
            self.student_excel_service,
            self._run_background,
            expense_service=self.expense_service,
            filter_current_rows_fn=self._filter_current_master_rows,
            status_message_fn=self._student_list_status_message,
        )
        self.student_list_view.student_selected.connect(self._open_student_profile)
        self.student_list_view.new_student_requested.connect(self.nav_add)
        self.student_list_view.students_changed.connect(self._on_students_changed)
        self.student_list_view.students_imported.connect(
            lambda count: self._audit(
                "import", "students", details={"record_count": count}
            )
        )
        self.stacked_widget.addWidget(self.student_list_view)  # Index 1
        self.create_profile_screen()    # Index 2
        self.create_add_screen()        # Index 3
        self.create_expenses_screen()   # Index 4
        self.create_workbook_screen()   # Index 5
        self.create_coordinators_screen()  # Index 6
        self.create_settings_screen()   # Index 7
        self.settings_view.theme_changed.connect(self.change_theme)
        self.settings_view.connection_settings_requested.connect(
            self._open_connection_settings
        )
        self.settings_view.test_connection_requested.connect(
            self._test_settings_connection
        )
        self.settings_view.sync_now_requested.connect(
            self.sync_google_sheet_now
        )
        self.settings_view.preferences_changed.connect(
            self._apply_accessibility_preferences
        )
        self.settings_view.sync_token_changed.connect(
            lambda _configured: self.refresh_sync_status()
        )
        self._update_compact_header()
        self._apply_button_cursors()
        self._apply_accessibility_defaults()
        self._install_shortcuts()

        # Initialize Data
        self.nav_dashboard()
        # Keep the Figma rail visually quiet after launch and mouse navigation,
        # while retaining a visible focus ring for keyboard users who Tab into it.
        self.stacked_widget.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.stacked_widget.setFocus(Qt.FocusReason.OtherFocusReason)
        self._start_keepalive()
        # --- Start Auto-Update Poller (Checks every 5 minutes) ---
        self._start_update_poller()

    def _run_background(self, function, on_success=None, on_error=None):
        """Run blocking work in Qt's managed thread pool."""
        task = BackgroundTask(function)
        if on_success is not None:
            task.signals.succeeded.connect(on_success)
        if on_error is not None:
            task.signals.failed.connect(on_error)
        self.thread_pool.start(task)
        return task

    def _mark_database_updated(self):
        self._last_database_update_at = datetime.now()
        self._set_connection_state("Connected", "success")
        self._refresh_database_updated_label()

    def _set_connection_state(self, text: str, state: str) -> None:
        badge = getattr(self, "connection_badge", None)
        if badge is not None:
            badge.set_state(state, text)
        settings_view = getattr(self, "settings_view", None)
        if settings_view is not None:
            settings_view.set_connection_state(text, state)

    def _refresh_database_updated_label(self):
        if self._last_database_update_at is None:
            self.database_updated_label.setText("DB: waiting")
            return
        elapsed = max(0, int((datetime.now() - self._last_database_update_at).total_seconds()))
        if elapsed < 10:
            text = "now"
        elif elapsed < 60:
            text = f"{elapsed}s ago"
        elif elapsed < 3600:
            text = f"{elapsed // 60}m ago"
        else:
            text = f"{elapsed // 3600}h ago"
        self.database_updated_label.setText(f"DB updated: {text}")

    def _student_list_status_message(self, message, timeout=0):
        self.status_bar.showMessage(message, timeout)
        if message.startswith("Loaded ") or message.startswith("Imported "):
            self._mark_database_updated()

    def _switch_page(self, index: int):
        if not hasattr(self, "stacked_widget") or index < 0:
            return
        if index >= self.stacked_widget.count():
            return
        active_nav = self._nav_button_for_page(index)
        if active_nav is not None:
            self._set_active_nav(active_nav)
        self._update_page_header(index)
        if self.stacked_widget.currentIndex() == index:
            return

        self.stacked_widget.setCurrentIndex(index)
        if self._reduce_motion:
            return
        # Fade leaf-level header copy, not the page container. Applying a
        # graphics effect to an input-heavy page can make Qt composite child
        # controls at stale positions after a responsive relayout.
        fade_in(
            self.page_title_label,
            motion_enabled=True,
            duration_ms=170,
        )
        if self.page_subtitle_label.isVisible():
            fade_in(
                self.page_subtitle_label,
                motion_enabled=True,
                duration_ms=190,
                delay_ms=24,
            )

    def _update_page_header(self, index: int) -> None:
        """Keep the workspace header useful and specific to the active screen."""
        form_header = (
            ("Edit student", "Update identity, support, and office record details")
            if self._editing_id
            else ("New student", "Create a complete student record for the shared office workspace")
        )
        pages = {
            0: ("Dashboard", "A clear view of student support, budgets, and recent activity."),
            1: ("Students", "Search, filter, and manage student records."),
            2: ("Student profile", "Review support details, completion, and current-year assistance."),
            3: form_header,
            4: ("Expenses", "Budget and expense history for the selected student"),
            5: ("Workbook", "Review and safely update the local master workbook"),
            6: ("Coordinators", "Contact directory for ministry coordinators"),
            7: (
                "Settings",
                "Configure appearance, accessibility, and office connections.",
            ),
        }
        title, subtitle = pages.get(index, ("SSM Workspace", "Student support records"))
        self.page_title_label.setText(title)
        self.page_subtitle_label.setText(subtitle)
        action_map = {
            0: (("Refresh data", "secondary"), None),
            1: (("Refresh data", "secondary"), None),
            2: (("Back to students", "secondary"), ("Edit record", "primary")),
            3: (("Cancel", "secondary"), ("Save changes" if self._editing_id else "Save student", "primary")),
            4: (("Back to profile", "secondary"), None),
            5: (None, ("Open in Excel", "primary")),
            6: (None, ("New coordinator", "primary")),
            7: (None, None),
        }
        secondary, primary = action_map.get(index, (None, None))
        self._configure_header_button(self.header_secondary_button, secondary)
        self._configure_header_button(self.header_add_button, primary)
        profile_header = index == 2
        self.header_profile_photo_button.setVisible(profile_header)
        self.header_profile_expenses_button.setVisible(profile_header)
        self.header_actions_layout.setContentsMargins(
            0, 0, 10 if profile_header else 0, 0
        )
        self.header_actions_layout.setSpacing(8)
        if profile_header:
            self.header_secondary_button.setFixedWidth(126)
            self.header_profile_photo_button.setFixedSize(106, 38)
            self.header_profile_expenses_button.setFixedSize(92, 38)
            self.header_add_button.setFixedWidth(92)

    @staticmethod
    def _configure_header_button(button, spec) -> None:
        button.setVisible(bool(spec))
        if not spec:
            return
        text, variant = spec
        button.setText(text)
        button.setEnabled(True)
        button.setProperty("variant", variant)
        button.style().unpolish(button)
        button.style().polish(button)
        set_content_hugging_button(button, height=38)

    def _nav_button_for_page(self, index: int):
        """Return the navigation owner for a workspace page.

        Detail pages deliberately remain owned by their parent destination so
        programmatic navigation cannot leave a stale item highlighted.
        """
        if index == 3:
            return self.btn_stud if self._editing_id else self.btn_add
        return {
            0: self.btn_dash,
            1: self.btn_stud,
            2: self.btn_stud,
            4: self.btn_exp,
            5: self.btn_workbook,
            6: self.btn_coordinators,
            7: self.btn_settings,
        }.get(index)

    def _header_action_clicked(self):
        index = self.stacked_widget.currentIndex()
        if index == 2:
            self.open_edit_screen()
        elif index == 3:
            self.save_student_form()
        elif index == 5:
            self.open_workbook_in_excel()
        elif index == 6:
            self._add_coordinator_dialog()
        else:
            self._sidebar_refresh()

    def _header_secondary_action_clicked(self):
        index = self.stacked_widget.currentIndex()
        if index in (2, 3):
            self.nav_students()
        elif index == 4:
            self._switch_page(2) if self.current_student_id else self.nav_students()
        else:
            self._sidebar_refresh()

    def resizeEvent(self, event) -> None:
        """Keep the workspace header usable at the supported minimum width."""
        super().resizeEvent(event)
        self._update_compact_header()
        self._update_profile_layout()
        self._relayout_dashboard_metrics()
        self._relayout_dashboard_insights()
        self._relayout_expense_entry()

    def _relayout_expense_entry(self) -> None:
        grid = getattr(self, "expense_add_grid", None)
        if grid is None:
            return
        compact = self.width() < 1120
        card = self.expense_add_card
        card.setFixedHeight(184 if compact else 118)
        widgets = (
            self.expense_description_label,
            self.expense_amount_label,
            self.expense_date_label,
            self.expense_year_label,
            self.exp_desc,
            self.exp_amount,
            self.exp_date,
            self.exp_sy_entry,
            self.add_expense_btn,
        )
        for widget in widgets:
            grid.removeWidget(widget)
        for column in range(5):
            grid.setColumnStretch(column, 0)

        if compact:
            grid.addWidget(self.expense_description_label, 0, 0, 1, 2)
            grid.addWidget(self.expense_amount_label, 0, 2)
            grid.addWidget(self.exp_desc, 1, 0, 1, 2)
            grid.addWidget(self.exp_amount, 1, 2)
            grid.addWidget(self.expense_date_label, 2, 0)
            grid.addWidget(self.expense_year_label, 2, 1)
            grid.addWidget(self.exp_date, 3, 0)
            grid.addWidget(self.exp_sy_entry, 3, 1)
            grid.addWidget(self.add_expense_btn, 3, 2)
            grid.setColumnStretch(0, 2)
            grid.setColumnStretch(1, 1)
            grid.setColumnStretch(2, 1)
            self.exp_date.setMinimumWidth(0)
        else:
            grid.addWidget(self.expense_description_label, 0, 0)
            grid.addWidget(self.expense_amount_label, 0, 1)
            grid.addWidget(self.expense_date_label, 0, 2)
            grid.addWidget(self.expense_year_label, 0, 3)
            grid.addWidget(self.exp_desc, 1, 0)
            grid.addWidget(self.exp_amount, 1, 1)
            grid.addWidget(self.exp_date, 1, 2)
            grid.addWidget(self.exp_sy_entry, 1, 3)
            grid.addWidget(self.add_expense_btn, 1, 4)
            grid.setColumnStretch(0, 3)
            grid.setColumnStretch(1, 1)
            grid.setColumnStretch(2, 1)
            grid.setColumnStretch(3, 1)
            self.exp_date.setMinimumWidth(190)

    def _update_profile_layout(self) -> None:
        """Preserve the Figma profile geometry while keeping 980px usable."""
        summary_layout = getattr(self, "profile_summary_layout", None)
        if summary_layout is None:
            return
        compact = self.width() < 1120
        summary_layout.setContentsMargins(
            0,
            15,
            18 if compact else 44,
            15,
        )
        self.profile_accent_gap.changeSize(12 if compact else 20, 0)
        self.profile_metric_gap.changeSize(12 if compact else 44, 0)
        self.profile_metric_widget.setFixedWidth(112 if compact else 180)
        self.profile_budget_widget.setFixedWidth(140 if compact else 244)
        self.profile_cards_layout.setVerticalSpacing(10 if compact else 20)
        for grid in self.profile_field_grids:
            for row in range(grid.rowCount()):
                grid.setRowMinimumHeight(row, 34 if compact else 40)
        for card in self.profile_top_cards:
            card.setFixedHeight(200 if compact else 222)
        for card in self.profile_bottom_cards:
            card.setFixedHeight(200 if compact else 220)
        summary_layout.invalidate()

    def _update_compact_header(self) -> None:
        if not hasattr(self, "header_add_button"):
            return
        compact = self.width() < 1120
        self.page_subtitle_label.setVisible(not compact)
        self._update_page_header(self.stacked_widget.currentIndex())
        if hasattr(self, "sync_panel"):
            self.sync_panel.setFixedHeight(126 if compact else 96)
        if hasattr(self, "settings_view"):
            self.settings_view.set_compact(compact)
        self.page_eyebrow_label.hide()

    def _dashboard_greeting(self) -> str:
        hour = time.localtime().tm_hour
        greeting = (
            "Good morning"
            if hour < 12
            else "Good afternoon"
            if hour < 18
            else "Good evening"
        )
        operator = getattr(self, "_current_operator", self._initial_user)
        return f"{greeting}, {operator}"
    def _apply_button_cursors(self):
        for button in self.findChildren(NativePushButton):
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            if not button.accessibleName() and button.text().strip():
                button.setAccessibleName(button.text().replace("&", ""))
            attach_press_feedback(
                button,
                lambda self=self: not self._reduce_motion,
            )
        for combo in self.findChildren(QComboBox):
            combo.setCursor(Qt.CursorShape.PointingHandCursor)

    def _install_shortcuts(self) -> None:
        """Install predictable desktop shortcuts without adding visual chrome."""
        shortcuts = (
            ("New student", QKeySequence.StandardKey.New, self.nav_add),
            ("Find", QKeySequence.StandardKey.Find, self._focus_page_search),
            ("Refresh", QKeySequence.StandardKey.Refresh, self._sidebar_refresh),
            ("Back", QKeySequence.StandardKey.Back, self._navigate_back),
        )
        self._shortcut_actions = []
        for text, sequence, callback in shortcuts:
            action = QAction(text, self)
            action.setShortcut(QKeySequence(sequence))
            action.setShortcutContext(Qt.ShortcutContext.WindowShortcut)
            action.triggered.connect(callback)
            self.addAction(action)
            self._shortcut_actions.append(action)

        self.header_add_button.setToolTip("Create a new student record (Ctrl+N)")
        self.sidebar_refresh_btn.setToolTip("Refresh the current data view (F5)")

    def _focus_page_search(self) -> None:
        """Focus the search field that belongs to the current workspace."""
        index = self.stacked_widget.currentIndex()
        if index == 1:
            field = self.student_list_view.search_name
        elif index == 6:
            field = self.coord_search
        else:
            self.nav_students()
            field = self.student_list_view.search_name
        field.setFocus(Qt.FocusReason.ShortcutFocusReason)
        field.selectAll()

    def _navigate_back(self) -> None:
        index = self.stacked_widget.currentIndex()
        if index in (2, 3, 4):
            self.nav_students()
        elif index != 0:
            self.nav_dashboard()

    def _apply_accessibility_defaults(self) -> None:
        """Add useful names and keyboard focus without changing workflows."""
        self.setAccessibleName("SSM Student Profiling System")
        self.stacked_widget.setAccessibleName("Application workspace")
        self.connection_badge.setAccessibleName("Database connection status")
        self.database_updated_label.setAccessibleName(
            "Database refresh timestamp"
        )
        for line_edit in self.findChildren(QLineEdit):
            if not line_edit.accessibleName() and line_edit.placeholderText():
                line_edit.setAccessibleName(line_edit.placeholderText())
        for table in self.findChildren(QTableWidget):
            table.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        for list_widget in self.findChildren(QListWidget):
            list_widget.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

    def _apply_accessibility_preferences(self) -> None:
        settings = self._settings()
        self._reduce_motion = settings.value(
            "reduce_motion", False, type=bool
        )
        if self._reduce_motion and hasattr(self, "_sync_pulse"):
            self._sync_pulse.stop()
        set_large_text(settings.value("large_text", False, type=bool))
        self.change_theme(settings.value("theme", "Light", type=str))

    def apply_modern_stylesheet(self):
        qss_path = resource_path(os.path.join("assets", "styles", "app.qss"))
        try:
            with open(qss_path, "r", encoding="utf-8") as file:
                self.setStyleSheet(render_qss(file.read()))
        except Exception as e:
            print(f"Stylesheet load error: {e}")

    def change_theme(self, theme_name: str):
        set_active_theme(theme_name)
        set_fluent_theme(theme_name)
        self.apply_modern_stylesheet()
        operator_avatar = getattr(self, "sidebar_operator_avatar", None)
        if operator_avatar is not None:
            operator_avatar.setPixmap(
                sidebar_operator_pixmap(self._current_operator)
            )

    # ── SIDEBAR NAVIGATION ────────────────────────────────────────────────────
    def create_sidebar(self, layout):
        self._using_fluent_sidebar = False
        # The custom shell is intentionally used even when QFluentWidgets is
        # installed. It gives the office app a distinct navigation identity and
        # avoids inheriting the stock Fluent navigation appearance.
        self._create_legacy_sidebar(layout)

    def _create_legacy_sidebar(self, layout):
        self.sidebar = QWidget()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setFixedWidth(224)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(12, 0, 12, 10)
        sidebar_layout.setSpacing(0)

        brand_container = QWidget()
        brand_container.setObjectName("BrandPanel")
        brand_container.setFixedHeight(72)
        brand_layout = QHBoxLayout(brand_container)
        brand_layout.setContentsMargins(12, 24, 0, 11)
        brand_layout.setSpacing(12)

        logo_lbl = NativeLabel()
        logo_lbl.setObjectName("BrandLogo")
        set_logo_pixmap(logo_lbl, 36, 36)

        brand_copy_widget = QWidget()
        brand_copy_widget.setObjectName("BrandCopy")
        brand_copy_widget.setFixedHeight(37)
        brand_copy = QVBoxLayout(brand_copy_widget)
        brand_copy.setContentsMargins(0, 0, 0, 0)
        brand_copy.setSpacing(2)
        self.brand_lbl = NativeLabel("SSM Office")
        self.brand_lbl.setObjectName("BrandTitle")
        self.brand_lbl.setFixedHeight(20)
        self.brand_lbl.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop
        )
        brand_sub = NativeLabel("Student records")
        brand_sub.setObjectName("BrandSubtitle")
        brand_sub.setFixedHeight(15)
        brand_sub.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop
        )
        brand_copy.addWidget(self.brand_lbl)
        brand_copy.addWidget(brand_sub)
        brand_layout.addWidget(logo_lbl, alignment=Qt.AlignmentFlag.AlignTop)
        brand_layout.addWidget(
            brand_copy_widget,
            1,
            alignment=Qt.AlignmentFlag.AlignTop,
        )
        sidebar_layout.addWidget(brand_container)

        sidebar_layout.addSpacing(28)

        self.btn_dash = NativePushButton("Dashboard")
        self.btn_stud = NativePushButton("Students")
        self.btn_exp  = NativePushButton("Expenses")
        self.btn_coordinators = NativePushButton("Coordinators")
        self.btn_add  = NativePushButton("New student")
        self.btn_workbook = NativePushButton("Workbook")
        self.btn_settings = NativePushButton("Settings")

        button_icons = {
            self.btn_dash: "⌂",
            self.btn_stud: "◎",
            self.btn_add: "+",
            self.btn_exp: "▤",
            self.btn_coordinators: "⌕",
            self.btn_workbook: "▧",
            self.btn_settings: "⚙",
        }
        for btn, icon_symbol in button_icons.items():
            btn.setProperty("class", "SidebarBtn")
            btn.setCheckable(True)
            btn.setFocusPolicy(Qt.FocusPolicy.TabFocus)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedHeight(42)
            btn.setSizePolicy(
                QSizePolicy.Policy.Expanding,
                QSizePolicy.Policy.Fixed,
            )
            btn.setIcon(sidebar_navigation_icon(icon_symbol))
            btn.setIconSize(QSize(20, 20))
            btn.setAccessibleName(f"Open {btn.text()}")

        self.btn_dash.clicked.connect(self.nav_dashboard)
        self.btn_stud.clicked.connect(self.nav_students)
        self.btn_add.clicked.connect(self.nav_add)
        self.btn_exp.clicked.connect(self.nav_expenses)
        self.btn_workbook.clicked.connect(self.nav_workbook)
        self.btn_coordinators.clicked.connect(self.nav_coordinators)
        self.btn_settings.clicked.connect(self.nav_settings)

        records_label = NativeLabel("RECORDS")
        records_label.setObjectName("SidebarGroupLabel")
        records_label.setFixedHeight(13)
        records_label.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop
        )
        sidebar_layout.addWidget(records_label)
        sidebar_layout.addSpacing(14)
        for button in (
            self.btn_dash,
            self.btn_stud,
            self.btn_add,
            self.btn_exp,
            self.btn_coordinators,
        ):
            sidebar_layout.addWidget(button)
            if button is not self.btn_coordinators:
                sidebar_layout.addSpacing(6)

        sidebar_layout.addSpacing(23)
        tools_label = NativeLabel("TOOLS")
        tools_label.setObjectName("SidebarGroupLabel")
        tools_label.setFixedHeight(13)
        tools_label.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop
        )
        sidebar_layout.addWidget(tools_label)
        sidebar_layout.addSpacing(14)
        sidebar_layout.addWidget(self.btn_workbook)
        sidebar_layout.addSpacing(6)
        sidebar_layout.addWidget(self.btn_settings)
        sidebar_layout.addStretch()

        user_panel = QWidget()
        user_panel.setObjectName("OperatorPanel")
        user_panel.setFixedHeight(62)
        user_label = NativeLabel("ACTIVE OPERATOR")
        user_label.setObjectName("SidebarCaption")
        user_label.setParent(user_panel)
        user_label.setGeometry(12, 7, 152, 14)
        self.sidebar_operator_avatar = NativeLabel()
        self.sidebar_operator_avatar.setObjectName("OperatorAvatar")
        self.sidebar_operator_avatar.setParent(user_panel)
        self.sidebar_operator_avatar.setFixedSize(0, 0)
        self.sidebar_operator_avatar.hide()
        self.sidebar_operator_avatar.setPixmap(
            sidebar_operator_pixmap(self._initial_user)
        )
        self.user_combo = NativeComboBox()
        self.user_combo.addItems(USERS)
        self.user_combo.setCurrentText(self._initial_user)
        self.user_combo.setObjectName("SidebarUserCombo")
        self.user_combo.setParent(user_panel)
        self.user_combo.setGeometry(12, 24, 176, 31)
        self.user_combo.setAccessibleName("Audit identity")
        self.user_combo.setAccessibleDescription(
            "Choose the office user whose name will be written to the activity log."
        )
        self.user_combo.setToolTip(
            "Choose the name written to the activity log"
        )
        self.user_combo.currentTextChanged.connect(self._on_user_changed)
        sidebar_layout.addWidget(user_panel)
        sidebar_layout.addSpacing(6)

        self.sidebar_refresh_btn = NativePushButton("Refresh data")
        self.sidebar_refresh_btn.setObjectName("SidebarRefreshBtn")
        self.sidebar_refresh_btn.setIcon(
            QIcon(resource_path(os.path.join("assets", "icons", "nav-refresh.svg")))
        )
        self.sidebar_refresh_btn.setIconSize(QSize(17, 17))
        self.sidebar_refresh_btn.setToolTip("Refresh data from Supabase")
        self.sidebar_refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.sidebar_refresh_btn.setFixedHeight(40)
        self.sidebar_refresh_btn.clicked.connect(self._sidebar_refresh)
        self.sidebar_refresh_btn.hide()

        version_label = NativeLabel(
            f"SSM   /   v{UpdaterService.CURRENT_VERSION}"
        )
        version_label.setObjectName("SidebarVersion")
        version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        version_label.setFixedHeight(10)
        sidebar_layout.addWidget(version_label)

        layout.addWidget(self.sidebar)

    def _create_fluent_sidebar(self, layout):
        self.sidebar = QWidget()
        self.sidebar.setObjectName("FluentSidebar")
        self.sidebar.setFixedWidth(218)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 24)
        sidebar_layout.setSpacing(8)

        brand_container = QWidget()
        brand_container.setObjectName("BrandPanel")
        brand_container.setFixedHeight(116)
        brand_layout = QVBoxLayout(brand_container)
        brand_layout.setContentsMargins(16, 14, 16, 12)
        brand_layout.setSpacing(4)
        brand_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        logo_lbl = QLabel()
        set_logo_pixmap(logo_lbl, 72, 52)

        self.brand_lbl = QLabel("SSM Workspace")
        self.brand_lbl.setObjectName("BrandTitle")
        self.brand_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)

        brand_sub = QLabel("Student support records")
        brand_sub.setObjectName("BrandSubtitle")
        brand_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)

        brand_layout.addWidget(logo_lbl, alignment=Qt.AlignmentFlag.AlignHCenter)
        brand_layout.addWidget(self.brand_lbl, alignment=Qt.AlignmentFlag.AlignHCenter)
        brand_layout.addWidget(brand_sub, alignment=Qt.AlignmentFlag.AlignHCenter)
        sidebar_layout.addWidget(brand_container)

        self.fluent_nav = NavigationInterface(
            self.sidebar,
            showMenuButton=True,
            showReturnButton=False,
        )
        if hasattr(self.fluent_nav, "setExpandWidth"):
            self.fluent_nav.setExpandWidth(218)
        if hasattr(self.fluent_nav, "setMinimumExpandWidth"):
            self.fluent_nav.setMinimumExpandWidth(0)
        self.fluent_nav.setMinimumHeight(310)

        self.btn_dash = "dashboard"
        self.btn_stud = "students"
        self.btn_exp = "expenses"
        self.btn_coordinators = "coordinators"
        self.btn_add = "add"
        self.btn_workbook = "workbook"
        self.btn_settings = "settings"

        self._add_fluent_nav_item(
            self.btn_dash,
            self._fluent_icon("HOME", QStyle.StandardPixmap.SP_DesktopIcon),
            "Dashboard",
            self.nav_dashboard,
        )
        self._add_fluent_nav_item(
            self.btn_stud,
            self._fluent_icon("PEOPLE", QStyle.StandardPixmap.SP_FileDialogDetailedView),
            "Students",
            self.nav_students,
        )
        self._add_fluent_nav_item(
            self.btn_exp,
            self._fluent_icon("PIE_SINGLE", QStyle.StandardPixmap.SP_DriveHDIcon),
            "Expenses",
            self.nav_expenses,
        )
        self._add_fluent_nav_item(
            self.btn_coordinators,
            self._fluent_icon("PHONE", QStyle.StandardPixmap.SP_FileDialogListView),
            "Coordinators",
            self.nav_coordinators,
        )
        self._add_fluent_nav_item(
            self.btn_add,
            self._fluent_icon("ADD", QStyle.StandardPixmap.SP_FileDialogNewFolder),
            "Add student",
            self.nav_add,
        )
        self._add_fluent_nav_item(
            self.btn_workbook,
            self._fluent_icon("DOCUMENT", QStyle.StandardPixmap.SP_FileIcon),
            "Workbook tabs",
            self.nav_workbook,
        )
        self._add_fluent_nav_item(
            self.btn_settings,
            self._fluent_icon("SETTING", QStyle.StandardPixmap.SP_FileDialogDetailedView),
            "Settings",
            self.nav_settings,
        )
        if hasattr(self.fluent_nav, "expand"):
            self.fluent_nav.expand(useAni=False)
        sidebar_layout.addWidget(self.fluent_nav, 1)

        user_label = QLabel("Changes recorded as")
        user_label.setObjectName("SidebarCaption")
        self.user_combo = QComboBox()
        self.user_combo.addItems(USERS)
        self.user_combo.setCurrentText(self._initial_user)
        self.user_combo.setObjectName("SidebarUserCombo")
        self.user_combo.setAccessibleName("Audit identity")
        self.user_combo.setAccessibleDescription(
            "Choose the office user whose name will be written to the activity log."
        )
        self.user_combo.setToolTip(
            "Choose the name written to the activity log"
        )
        self.user_combo.currentTextChanged.connect(self._on_user_changed)
        sidebar_layout.addWidget(user_label)
        sidebar_layout.addWidget(self.user_combo)
        sidebar_layout.addSpacing(8)

        self.sidebar_refresh_btn = QPushButton("Refresh")
        self.sidebar_refresh_btn.setObjectName("SidebarRefreshBtn")
        self.sidebar_refresh_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_BrowserReload))
        self.sidebar_refresh_btn.setIconSize(QSize(15, 15))
        self.sidebar_refresh_btn.setToolTip("Refresh data from Supabase")
        self.sidebar_refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.sidebar_refresh_btn.setMinimumHeight(34)
        self.sidebar_refresh_btn.setMaximumHeight(36)
        self.sidebar_refresh_btn.clicked.connect(self._sidebar_refresh)
        sidebar_layout.addWidget(self.sidebar_refresh_btn)

        layout.addWidget(self.sidebar)

    def _fluent_icon(self, icon_name, fallback_pixmap):
        if FluentIcon is not None and hasattr(FluentIcon, icon_name):
            return getattr(FluentIcon, icon_name)
        return self.style().standardIcon(fallback_pixmap)

    def _add_fluent_nav_item(self, route_key, icon, text, callback):
        try:
            self.fluent_nav.addItem(
                routeKey=route_key,
                icon=icon,
                text=text,
                onClick=callback,
                selectable=True,
                position=NavigationItemPosition.TOP,
            )
        except TypeError:
            self.fluent_nav.addItem(route_key, icon, text, callback)

    def nav_dashboard(self):
        self._set_active_nav(self.btn_dash)
        self.refresh_dashboard()
        self.refresh_sync_status()
        self._switch_page(0)
        if not self._dashboard_revealed:
            QTimer.singleShot(20, self._reveal_dashboard_once)

    def _reveal_dashboard_once(self) -> None:
        """Orchestrate the first dashboard reveal without delaying its data."""
        if self._dashboard_revealed:
            return
        self._dashboard_revealed = True
        fade_in(
            self._dashboard_intro,
            motion_enabled=not self._reduce_motion,
            duration_ms=180,
        )
        for index, card in enumerate(self._dashboard_motion_cards):
            card.reveal(45 + (index * 45))
        for index, label in enumerate(self._dashboard_insight_headers):
            fade_in(
                label,
                motion_enabled=not self._reduce_motion,
                duration_ms=180,
                delay_ms=210 + (index * 35),
            )

    def nav_students(self):
        self._set_active_nav(self.btn_stud)
        self.student_list_view.refresh_grade_filter()
        if self.student_list_view.student_list_mode:
            self.student_list_view.load_student_list()
        else:
            self.student_list_view.show_student_view_prompt()
        self._switch_page(1)

    def _on_dashboard_student_click(self, item):
        student_id = item.data(Qt.ItemDataRole.UserRole)
        if student_id:
            self._open_student_profile(str(student_id))

    def _open_student_profile(self, student_id: str):
        self.current_student_id = student_id
        self._set_active_nav(self.btn_stud)
        self._load_profile(student_id)
        self._switch_page(2)

    def _on_students_changed(self):
        self._mark_database_updated()
        self.refresh_dashboard()
        self.student_list_view.refresh_filter_options()
        if self.stacked_widget.currentIndex() == 1:
            self.student_list_view.load_student_list()

    def nav_add(self):
        self._set_active_nav(self.btn_add)
        self._open_add_screen()

    def nav_expenses(self):
        if self.current_student_id:
            self._set_active_nav(self.btn_exp)
            self.open_expenses_screen()
        else:
            QMessageBox.information(self, "Select student", "Select a student from the Students list first to view expenses.")
            self.nav_students()

    def nav_workbook(self):
        self._set_active_nav(self.btn_workbook)
        self._switch_page(5)

    def nav_coordinators(self):
        self._set_active_nav(self.btn_coordinators)
        self._switch_page(6)
        self.load_coordinators()

    def nav_settings(self):
        self._set_active_nav(self.btn_settings)
        self._switch_page(7)
        self.settings_view.load_settings()
        self.refresh_sync_status()
        self._test_settings_connection(silent=True)

    def _on_user_changed(self, name):
        if not name:
            return
        self._current_operator = name
        self._update_operator_labels(name)
        if self.stacked_widget.currentIndex() == 0:
            self._update_page_header(0)

    def _update_operator_labels(self, name: str) -> None:
        operator_avatar = getattr(self, "sidebar_operator_avatar", None)
        if operator_avatar is not None:
            operator_avatar.setPixmap(sidebar_operator_pixmap(name))
        selector = getattr(self, "user_combo", None)
        if selector is not None:
            selector.setToolTip(
                f"Changes are currently recorded as {name}"
            )
            selector.setAccessibleDescription(
                f"Changes are currently recorded as {name}. "
                "Choose another office user to change the activity log name."
            )

    def _open_connection_settings(self) -> None:
        dialog = ConfigurationDialog(
            "Update the Supabase connection saved on this computer."
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        QMessageBox.information(
            self,
            "Connection saved",
            (
                "The new connection has been saved for this computer.\n\n"
                "Close and reopen the application to use it."
            ),
        )

    def _test_settings_connection(self, *_args, silent=False) -> None:
        """Test the live repository without blocking the Settings screen."""
        self.settings_view.set_connection_state("Checking", "loading")

        def connected(_rows):
            self._mark_database_updated()
            if not silent:
                self.status_bar.showMessage("Database connection is working.", 4000)

        def failed(error):
            self._set_connection_state("Connection issue", "danger")
            self.status_bar.showMessage(
                "Could not reach the office database. Check the saved connection.",
                7000,
            )
            logging.getLogger(__name__).warning(
                "Settings connection test failed: %s",
                str(error).strip().splitlines()[-1],
            )

        self._run_background(self.student_repository.ping, connected, failed)

    def _audit(self, action, entity_type, entity_id=None, details=None):
        operator = getattr(self, "_current_operator", self._initial_user)
        self._run_background(
            lambda: self.audit_repository.log(
                operator=operator,
                action=action,
                entity_type=entity_type,
                entity_id=entity_id,
                details=details,
            ),
            on_error=lambda error: logging.getLogger(__name__).error(
                "Audit log failed: %s", error.strip().splitlines()[-1]
            ),
        )

    def _sidebar_refresh(self):
        button = getattr(self, "sidebar_refresh_btn", None)
        if button is not None:
            button.setEnabled(False)
            button.setText("Refreshing…")
        self.status_bar.showMessage("Refreshing…", 1200)

        idx = self.stacked_widget.currentIndex()
        task = None
        if idx == 0:
            task = self.refresh_dashboard()
        elif idx == 1:
            task = self.student_list_view.load_student_list()
        elif idx == 2 and self.current_student_id:
            task = self._load_profile(self.current_student_id)
        elif idx == 4 and self.current_student_id:
            task = self._refresh_expenses_view()
        elif idx == 6:
            task = self.load_coordinators()

        if button is not None and task is not None:
            def restore_refresh_button():
                button.setEnabled(True)
                button.setText("Refresh data")

            task.signals.finished.connect(restore_refresh_button)
            if getattr(task, "done", False):
                QTimer.singleShot(0, restore_refresh_button)
        elif button is not None:
            QTimer.singleShot(
                450,
                lambda: (
                    button.setEnabled(True),
                    button.setText("Refresh data"),
                ),
            )
    def _set_active_nav(self, active_btn):
        if isinstance(active_btn, str):
            nav = getattr(self, "fluent_nav", None)
            if nav is not None and hasattr(nav, "setCurrentItem"):
                try:
                    nav.setCurrentItem(active_btn)
                except Exception:
                    logging.getLogger(__name__).debug(
                        "Could not set Fluent nav route %s", active_btn,
                        exc_info=True,
                    )
            return
        for btn in [self.btn_dash, self.btn_stud, self.btn_exp, self.btn_coordinators, self.btn_add, self.btn_workbook, self.btn_settings]:
            btn.setChecked(btn is active_btn)

    # ── KEEPALIVE ─────────────────────────────────────────────────────────────
    def _start_keepalive(self):
        self.keepalive_timer = QTimer(self)
        self.keepalive_timer.setInterval(KEEPALIVE_INTERVAL_MS)
        self.keepalive_timer.timeout.connect(self._do_keepalive)
        self.keepalive_timer.start()
        self._mark_database_updated()
        self.status_bar.showMessage("Connected to Supabase", 5000)

    def _do_keepalive(self):
        self._run_background(
            self.student_repository.ping,
            lambda _rows: (self._mark_database_updated(), self.status_bar.showMessage("Supabase keepalive OK", 4000)),
            lambda error: (
                self._set_connection_state("Connection issue", "danger"),
                self.status_bar.showMessage(
                    "Could not reach the office database. Use Refresh data to try again.",
                    8000,
                ),
                logging.getLogger(__name__).warning(
                    "Supabase keepalive failed: %s",
                    error.strip().splitlines()[-1],
                ),
            ),
        )
    
    # ── LIVE AUTO-UPDATER ─────────────────────────────────────────────────────
    def _start_update_poller(self):
        self.update_timer = QTimer(self)
        self.update_timer.setInterval(300000) # 5 minutes in milliseconds
        self.update_timer.timeout.connect(self._check_for_updates_bg)
        self.update_timer.start()

    def _check_for_updates_bg(self):
        """Silently check for updates in the background."""
        self._run_background(
            lambda: UpdaterService(self.sb).check_for_update(),
            self._on_update_check_result,
            lambda e: None # If internet drops briefly, fail silently
        )

    def _on_update_check_result(self, update_info):
        if not update_info:
            return

        latest_version, url = update_info
        
        # Pause the timer so we don't spam them with popups every 5 minutes
        self.update_timer.stop()

        reply = QMessageBox.question(
            self,
            "Update available",
            f"A new version of the SSM System (v{latest_version}) has just been published!\n\nWould you like to download and install it now?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self._trigger_live_update(latest_version, url)
        else:
            # If they say no, wait 30 minutes before reminding them again
            self.update_timer.start(1800000)

    def _trigger_live_update(self, latest_version, url):
        # 1. Safety Check: Don't restart if they have unsaved work!
        if getattr(self, "_workbook_dirty", False):
            QMessageBox.warning(
                self, 
                "Unsaved work",
                "You have unsaved changes in your workbook!\n\nPlease save your workbook first, then the update prompt will reappear shortly."
            )
            # Check again in 1 minute
            self.update_timer.start(60000)
            return

        # 2. Show a progress dialog overlay
        self.update_progress = QProgressDialog("Downloading update...", "Cancel", 0, 100, self)
        self.update_progress.setWindowTitle(f"Updating to v{latest_version}")
        self.update_progress.setWindowModality(Qt.WindowModality.WindowModal)
        self.update_progress.setAutoClose(False)
        self.update_progress.setAutoReset(False)
        self.update_progress.show()

        # 3. Trigger the download (using the safe UI thread routing we set up earlier)
        updater = UpdaterService(self.sb)

        # Wire Cancel button to abort the download thread before it can call os._exit()
        self.update_progress.canceled.connect(lambda: (
            updater.cancel(),
            self.update_progress.close(),
            self.update_timer.start(1800000)  # Re-prompt in 30 min
        ))

        updater.download_and_install(
            url=url,
            progress_callback=lambda p: QTimer.singleShot(0, lambda: self.update_progress.setValue(p)),
            success_callback=lambda: QTimer.singleShot(0, lambda: self.update_progress.setLabelText("Restarting app to apply update...")),
            error_callback=lambda err: QTimer.singleShot(0, lambda: self._on_live_update_failed(err))
        )

    def _on_live_update_failed(self, error):
        progress = getattr(self, "update_progress", None)
        if progress is not None:
            progress.close()
        QMessageBox.critical(self, "Update failed", str(error))
        self.update_timer.start(1800000)

    def _update_field(self, table: str, field: str, value, record_id):
        try:
            if table != "students":
                raise ValueError(f"Unsupported table for field update: {table}")
            self.student_repository.update_student(record_id, {field: value})
            self.status_bar.showMessage(f"{field.replace('_', ' ').capitalize()} saved", 3000)
        except Exception as e:
            self.status_bar.showMessage(f"Error saving {field} ({type(e).__name__}): {e}", 8000)

    # ── SCREEN 0: DASHBOARD ───────────────────────────────────────────────────
    def create_dashboard_screen(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)
        page_layout.setSpacing(0)

        scroll = QScrollArea()
        scroll.setObjectName("DashboardScroll")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )
        scroll.setAccessibleName("Dashboard overview")
        widget = QWidget()
        widget.setObjectName("DashboardContent")
        layout = QVBoxLayout(widget)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        layout.setSpacing(16)
        layout.setContentsMargins(0, 0, 6, 18)
        scroll.setWidget(widget)
        page_layout.addWidget(scroll)

        overview_row = QHBoxLayout()
        overview_lbl = StrongBodyLabel("Current synchronized records")
        overview_lbl.setObjectName("SectionTitle")
        self._dashboard_intro = overview_lbl
        overview_lbl.hide()
        self.dashboard_state_label = QLabel("Preparing data")
        self.dashboard_state_label.setObjectName("DashboardState")
        self.dashboard_state_label.setFixedHeight(26)
        self.dashboard_state_label.setAccessibleName("Dashboard refresh status")
        self.dashboard_state_label.hide()
        overview_row.addWidget(overview_lbl)
        overview_row.addStretch()
        overview_row.addWidget(self.dashboard_state_label)
        layout.addLayout(overview_row)

        self.dashboard_cards_grid = QGridLayout()
        self.dashboard_cards_grid.setHorizontalSpacing(12)
        self.dashboard_cards_grid.setVerticalSpacing(12)

        self.lbl_total_val    = QLabel("--"); self.lbl_total_val.setObjectName("CardValue")
        self.lbl_active_val   = QLabel("--"); self.lbl_active_val.setObjectName("CardValue")
        self.lbl_inactive_val = QLabel("--"); self.lbl_inactive_val.setObjectName("CardValue")
        self.lbl_graduated_val = QLabel("--"); self.lbl_graduated_val.setObjectName("CardValue")
        self.dashboard_budget_value = QLabel("--")
        self.dashboard_budget_value.setObjectName("CardValue")
        self.dashboard_pending_value = QLabel("--")
        self.dashboard_pending_value.setObjectName("CardValue")

        self.dashboard_metric_cards = [
            self._build_card(
                "ACTIVE STUDENTS", self.lbl_active_val, "success",
                "Currently supported",
            ),
            self._build_card(
                "GRADUATING", self.lbl_graduated_val, "graduated",
                "Completing the program",
            ),
            self._build_card(
                "BUDGET USED", self.dashboard_budget_value, "warning",
                "Current school year",
            ),
            self._build_card(
                "PENDING RECORDS", self.dashboard_pending_value, "danger",
                "Profiles that need review",
            ),
        ]
        self._relayout_dashboard_metrics()
        layout.addLayout(self.dashboard_cards_grid)

        overview = QHBoxLayout()
        overview.setSpacing(16)

        budget_card = Card()
        budget_card.setObjectName("DashboardPanel")
        budget_card.setFixedHeight(242)
        budget_layout = QVBoxLayout(budget_card)
        budget_layout.setContentsMargins(20, 16, 20, 14)
        budget_layout.setSpacing(7)
        budget_title = StrongBodyLabel("Budget overview")
        budget_title.setObjectName("CardHeading")
        budget_caption = QLabel("Current-school-year allocation and utilization")
        budget_caption.setObjectName("Caption")
        budget_layout.addWidget(budget_title)
        budget_layout.addWidget(budget_caption)
        budget_amount_row = QHBoxLayout()
        self.dashboard_budget_spent = QLabel("PHP --")
        self.dashboard_budget_spent.setObjectName("DashboardHeroValue")
        self.dashboard_budget_of = QLabel("used of PHP --")
        self.dashboard_budget_of.setObjectName("Caption")
        budget_amount_row.addWidget(self.dashboard_budget_spent)
        budget_amount_row.addWidget(self.dashboard_budget_of)
        budget_amount_row.addStretch()
        budget_layout.addLayout(budget_amount_row)
        budget_meta = QHBoxLayout()
        self.dashboard_budget_percent = QLabel("0% used")
        self.dashboard_budget_percent.setObjectName("BudgetStatus")
        self.dashboard_budget_remaining = QLabel("PHP -- remaining")
        self.dashboard_budget_remaining.setObjectName("Caption")
        budget_meta.addWidget(self.dashboard_budget_percent)
        budget_meta.addStretch()
        budget_meta.addWidget(self.dashboard_budget_remaining)
        budget_layout.addLayout(budget_meta)
        self.dashboard_budget_bar = QProgressBar()
        self.dashboard_budget_bar.setObjectName("BudgetProgress")
        self.dashboard_budget_bar.setRange(0, 100)
        self.dashboard_budget_bar.setTextVisible(False)
        self.dashboard_budget_bar.setFixedHeight(7)
        budget_layout.addWidget(self.dashboard_budget_bar)
        budget_scale = QHBoxLayout()
        zero = QLabel("0")
        zero.setObjectName("Caption")
        self.dashboard_budget_scale_max = QLabel("PHP 0")
        self.dashboard_budget_scale_max.setObjectName("Caption")
        budget_scale.addWidget(zero)
        budget_scale.addStretch()
        budget_scale.addWidget(self.dashboard_budget_scale_max)
        budget_layout.addLayout(budget_scale)
        budget_rule = QFrame()
        budget_rule.setObjectName("Divider")
        budget_rule.setFixedHeight(1)
        budget_layout.addWidget(budget_rule)
        budget_footer = QHBoxLayout()
        pace_label = QLabel("Monthly pace")
        pace_label.setObjectName("Caption")
        self.dashboard_monthly_pace = StrongBodyLabel("PHP --")
        self.dashboard_budget_health = StatusBadge("Unallocated", state="neutral")
        budget_footer.addWidget(pace_label)
        budget_footer.addWidget(self.dashboard_monthly_pace)
        budget_footer.addStretch()
        budget_footer.addWidget(self.dashboard_budget_health)
        budget_layout.addLayout(budget_footer)
        overview.addWidget(budget_card, 5)

        progress_card = Card()
        progress_card.setObjectName("DashboardPanel")
        progress_card.setFixedHeight(242)
        progress_layout = QVBoxLayout(progress_card)
        progress_layout.setContentsMargins(20, 16, 20, 14)
        progress_layout.setSpacing(5)
        progress_title = StrongBodyLabel("Student progress")
        progress_title.setObjectName("CardHeading")
        progress_caption = QLabel("Profile completion across active records")
        progress_caption.setObjectName("Caption")
        progress_layout.addWidget(progress_title)
        progress_layout.addWidget(progress_caption)
        self.dashboard_progress_rows = {}
        for key, label_text, tone in (
            ("complete", "Complete", "success"),
            ("progress", "In progress", "warning"),
            ("review", "Needs review", "danger"),
        ):
            row_meta = QHBoxLayout()
            label = StrongBodyLabel(label_text)
            value = StrongBodyLabel("--")
            row_meta.addWidget(label)
            row_meta.addStretch()
            row_meta.addWidget(value)
            bar = QProgressBar()
            bar.setObjectName("DashboardProfileProgress")
            bar.setProperty("tone", tone)
            bar.setRange(0, 100)
            bar.setValue(0)
            bar.setTextVisible(False)
            bar.setFixedHeight(7)
            progress_layout.addLayout(row_meta)
            progress_layout.addWidget(bar)
            self.dashboard_progress_rows[key] = (value, bar)
        progress_layout.addStretch()
        overview.addWidget(progress_card, 3)
        layout.addLayout(overview)

        recent_card = Card()
        recent_card.setObjectName("DashboardPanel")
        recent_card.setFixedHeight(246)
        recent_layout = QVBoxLayout(recent_card)
        recent_layout.setContentsMargins(14, 12, 14, 10)
        recent_layout.setSpacing(7)
        recent_header = QHBoxLayout()
        recent_copy = QVBoxLayout()
        recent_copy.setSpacing(2)
        recent_title = StrongBodyLabel("Recent student updates")
        recent_title.setObjectName("CardHeading")
        recent_caption = QLabel("Latest synchronized student records")
        recent_caption.setObjectName("Caption")
        recent_copy.addWidget(recent_title)
        recent_copy.addWidget(recent_caption)
        recent_header.addLayout(recent_copy)
        recent_header.addStretch()
        view_all = ActionButton("View all students", variant="secondary")
        view_all.clicked.connect(self.nav_students)
        set_content_hugging_button(view_all, height=38)
        recent_header.addWidget(view_all)
        recent_layout.addLayout(recent_header)
        self.dashboard_recent_table = QTableWidget()
        self.dashboard_recent_table.setObjectName("DashboardRecentTable")
        self.dashboard_recent_table.setColumnCount(5)
        self.dashboard_recent_table.setHorizontalHeaderLabels(
            ["STUDENT", "STATUS", "PROFILE", "BUDGET", "UPDATED"]
        )
        self.dashboard_recent_table.verticalHeader().hide()
        self.dashboard_recent_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.dashboard_recent_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.dashboard_recent_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.dashboard_recent_table.setShowGrid(False)
        self.dashboard_recent_table.setAlternatingRowColors(False)
        self.dashboard_recent_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for column in range(1, 5):
            self.dashboard_recent_table.horizontalHeader().setSectionResizeMode(column, QHeaderView.ResizeMode.ResizeToContents)
        self.dashboard_recent_table.cellDoubleClicked.connect(self._open_dashboard_recent)
        recent_layout.addWidget(self.dashboard_recent_table)
        layout.addWidget(recent_card)

        sync_panel = self._build_sync_panel()
        sync_panel.setFixedHeight(84)
        sync_panel.hide()
        layout.addWidget(sync_panel)
        self.stacked_widget.addWidget(page)

    def _build_sync_panel(self):
        panel = QFrame()
        panel.setObjectName("SyncPanel")
        self.sync_panel = panel
        panel.setFixedHeight(96)
        panel.setAccessibleName("Google Sheets synchronization")
        panel.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Maximum,
        )
        layout = QGridLayout(panel)
        layout.setContentsMargins(0, 3, 0, 12)
        layout.setHorizontalSpacing(16)
        layout.setVerticalSpacing(7)
        layout.setColumnStretch(1, 1)

        accent_rail = QFrame()
        accent_rail.setObjectName("SyncAccentRail")
        accent_rail.setFixedWidth(3)
        layout.addWidget(accent_rail, 0, 0, 2, 1)

        title_row = QHBoxLayout()
        title_row.setSpacing(8)
        title = StrongBodyLabel("Workbook sync")
        title.setObjectName("SyncPanelTitle")
        self.sync_state_badge = StatusBadge(
            "Checking",
            state="neutral",
            role="SyncStateBadge",
        )
        self.sync_state_badge.setFixedHeight(28)
        self.sync_state_badge.setAccessibleName(
            "Google Sheets synchronization status"
        )
        self._sync_pulse = PulseController(
            self.sync_state_badge,
            lambda: not self._reduce_motion,
        )
        title_row.addWidget(title)
        title_row.addWidget(self.sync_state_badge)
        title_row.addStretch()
        layout.addLayout(title_row, 0, 1)

        meta = QGridLayout()
        meta.setHorizontalSpacing(20)
        meta.setVerticalSpacing(2)
        meta.setColumnStretch(0, 1)
        meta.setColumnStretch(1, 2)
        last_caption = QLabel("Last sync")
        last_caption.setObjectName("SyncMetaLabel")
        self.sync_last_label = QLabel("Checking…")
        self.sync_last_label.setObjectName("SyncMetaValue")
        self.sync_last_label.setAccessibleName("Last successful sync time")
        records_caption = QLabel("Latest result")
        records_caption.setObjectName("SyncMetaLabel")
        self.sync_records_label = QLabel("—")
        self.sync_records_label.setObjectName("SyncMetaValue")
        self.sync_records_label.setAccessibleName("Latest sync record counts")
        self.sync_records_label.setWordWrap(True)
        self.sync_records_label.setMinimumWidth(0)
        self.sync_records_label.setSizePolicy(
            QSizePolicy.Policy.Ignored,
            QSizePolicy.Policy.Preferred,
        )
        meta.addWidget(last_caption, 0, 0)
        meta.addWidget(self.sync_last_label, 1, 0)
        meta.addWidget(records_caption, 0, 1)
        meta.addWidget(self.sync_records_label, 1, 1)
        layout.addLayout(meta, 1, 1)

        actions = QHBoxLayout()
        actions.setContentsMargins(0, 0, 0, 0)
        actions.setSpacing(6)
        self.sync_now_button = ActionButton("Sync now")
        self.sync_now_button.setAccessibleDescription(
            "Synchronizes the current Google workbook into Supabase."
        )
        self.sync_now_button.clicked.connect(self.sync_google_sheet_now)
        self.refresh_data_button = ActionButton(
            "Refresh", variant="tertiary"
        )
        self.refresh_data_button.setAccessibleDescription(
            "Reloads dashboard data already stored in Supabase."
        )
        self.refresh_data_button.clicked.connect(self._sidebar_refresh)
        actions.addWidget(self.sync_now_button)
        actions.addWidget(self.refresh_data_button)
        actions_widget = QWidget()
        actions_widget.setObjectName("SyncActions")
        actions_widget.setSizePolicy(
            QSizePolicy.Policy.Maximum,
            QSizePolicy.Policy.Fixed,
        )
        actions_widget.setFixedHeight(40)
        actions_widget.setLayout(actions)
        layout.addWidget(
            actions_widget,
            0,
            2,
            2,
            1,
            Qt.AlignmentFlag.AlignVCenter,
        )
        return panel

    def refresh_sync_status(self):
        """Load the last server-side sync audit entry without blocking Qt."""
        if not hasattr(self, "sync_state_badge"):
            return None
        self.sync_state_badge.set_state("loading", "Checking")
        self._sync_pulse.start()
        self.sync_last_label.setText("Checking…")

        def apply_status(entry):
            self._sync_pulse.stop()
            has_token = bool(get_sheet_sync_token())
            if not entry:
                self.sync_state_badge.set_state(
                    "warning", "Setup needed" if not has_token else "Not synced"
                )
                self.sync_last_label.setText("No successful sync recorded")
                self.sync_records_label.setText("—")
                self.settings_view.set_sync_status(
                    configured=has_token,
                    state="Ready to sync" if has_token else "Token needed",
                    last_sync="No successful sync recorded",
                )
                return
            details = entry.get("details") or {}
            self.sync_state_badge.set_state(
                "success" if has_token else "warning",
                "Ready" if has_token else "Token needed",
            )
            self.sync_last_label.setText(
                self.sheet_sync_service.format_timestamp(
                    entry.get("created_at")
                )
            )
            self.sync_records_label.setText(
                self._format_sync_counts(details)
            )
            self.settings_view.set_sync_status(
                configured=has_token,
                state="Ready to sync" if has_token else "Token needed",
                last_sync=self.sheet_sync_service.format_timestamp(
                    entry.get("created_at")
                ),
                active_rows=int(details.get("students") or 0),
                source=str(
                    details.get("source")
                    or details.get("source_sheet_name")
                    or "SSM Masterlist / Current workbook"
                ),
            )

        def failed(error):
            self._sync_pulse.stop()
            logging.getLogger(__name__).warning(
                "Could not load sync status: %s", error
            )
            self.sync_state_badge.set_state("warning", "Status unavailable")
            self.sync_last_label.setText("Could not read the audit log")
            self.sync_records_label.setText("—")
            self.settings_view.set_sync_status(
                configured=bool(get_sheet_sync_token()),
                state="Status unavailable",
                last_sync="Could not read the audit log",
            )

        return self._run_background(
            self.sheet_sync_service.latest_success,
            apply_status,
            failed,
        )

    def sync_google_sheet_now(self):
        """Run the protected server sync and refresh the visible workspace."""
        token = get_sheet_sync_token()
        if not token:
            self.sync_state_badge.set_state("warning", "Token needed")
            self.status_bar.showMessage(
                "Add the private Google Sheet sync token in Settings.", 6000
            )
            self.nav_settings()
            QTimer.singleShot(0, self.settings_view.prompt_replace_token)
            return None

        decision = QMessageBox.question(
            self,
            "Synchronize Google Sheets",
            "Update students, donor assignments, movements, and coordinators "
            "from the current Google workbook?\n\n"
            "The server will roll back every change if any part fails.",
            QMessageBox.StandardButton.Yes |
            QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Cancel,
        )
        if decision != QMessageBox.StandardButton.Yes:
            return None

        self._set_sync_busy(True)
        self.sync_state_badge.set_state("loading", "Syncing")
        self._sync_pulse.start()
        self.sync_last_label.setText("Reading the current workbook…")
        self.status_bar.showMessage("Synchronizing Google Sheets…")

        def synchronize():
            return self.sheet_sync_service.synchronize(token)

        def synchronized(result):
            self._set_sync_busy(False)
            self._sync_pulse.stop()
            database_result = result.get("database_result") or {}
            excluded_rows = result.get("excluded_donor_rows") or []
            self.sync_state_badge.set_state("success", "Synced")
            self.sync_last_label.setText("Completed just now")
            self.sync_records_label.setText(
                self._format_sync_counts(database_result)
            )
            fade_in(
                self.sync_state_badge,
                motion_enabled=not self._reduce_motion,
                duration_ms=180,
            )
            if excluded_rows:
                self.sync_state_badge.setToolTip(
                    "Excluded donor source rows: " +
                    ", ".join(str(row) for row in excluded_rows)
                )
            self._mark_database_updated()
            self.status_bar.showMessage(
                "Google workbook synchronized successfully.", 7000
            )
            self.refresh_dashboard()
            self.student_list_view.refresh_filter_options()
            if self.stacked_widget.currentIndex() == 1:
                self.student_list_view.load_student_list()
            if self.stacked_widget.currentIndex() == 6:
                self.load_coordinators()
            self.refresh_sync_status()

        def failed(error):
            self._set_sync_busy(False)
            self._sync_pulse.stop()
            message = self._friendly_background_error(error)
            self.sync_state_badge.set_state("danger", "Sync failed")
            self.sync_last_label.setText(message)
            self.sync_records_label.setText("No database changes applied")
            self.status_bar.showMessage(message, 9000)

        return self._run_background(synchronize, synchronized, failed)

    def _set_sync_busy(self, busy: bool) -> None:
        self.sync_now_button.setEnabled(not busy)
        self.sync_now_button.setText("Syncing…" if busy else "Sync now")
        self.refresh_data_button.setEnabled(not busy)
        self.settings_view.set_sync_busy(busy)

    @staticmethod
    def _format_sync_counts(details) -> str:
        if not isinstance(details, dict) or not details:
            return "—"
        students = int(details.get("students") or 0)
        donors = int(details.get("donor_students") or 0)
        movements = int(details.get("movements") or 0)
        coordinators = int(details.get("coordinators") or 0)
        return (
            f"{students} students · {donors} donors · "
            f"{movements} movements · {coordinators} coordinators"
        )

    @staticmethod
    def _friendly_background_error(error: str) -> str:
        last_line = str(error or "").strip().splitlines()[-1]
        if ": " in last_line:
            last_line = last_line.split(": ", 1)[1]
        return last_line or "The operation failed. Try again."

    def _build_card(
        self,
        title_text,
        value_label,
        tone="primary",
        caption_text="",
    ):
        card = MotionCard(
            motion_enabled=lambda: not self._reduce_motion,
            hover_depth=False,
        )
        card.setObjectName("DashboardMetricCard")
        card.setProperty("tone", tone)
        card.setMinimumWidth(0)
        card.setFixedHeight(104)
        shell = QHBoxLayout(card)
        shell.setContentsMargins(0, 0, 0, 0)
        shell.setSpacing(0)
        marker = QFrame()
        marker.setObjectName("MetricMarker")
        marker.setProperty("tone", tone)
        marker.setFixedWidth(4)
        marker.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)
        shell.addWidget(marker)
        content = QWidget()
        l = QVBoxLayout(content)
        l.setContentsMargins(14, 12, 14, 11)
        l.setSpacing(2)
        t = StrongBodyLabel(title_text)
        t.setObjectName("CardTitle")
        t.setWordWrap(True)
        l.addWidget(t)
        l.addWidget(value_label)
        caption = QLabel(caption_text)
        caption.setObjectName("MetricCaption")
        l.addWidget(caption)
        shell.addWidget(content, 1)
        value_label.setAccessibleName(f"{title_text} count")
        value_label.setAccessibleDescription(caption_text)
        card.setAccessibleName(f"{title_text} dashboard metric")
        self._dashboard_motion_cards.append(card)
        return card

    def _relayout_dashboard_metrics(self) -> None:
        grid = getattr(self, "dashboard_cards_grid", None)
        cards = getattr(self, "dashboard_metric_cards", None)
        if grid is None or not cards:
            return
        for card in cards:
            grid.removeWidget(card)
        columns = 4 if self.width() >= 1180 else 2
        for column in range(4):
            grid.setColumnStretch(column, 1 if column < columns else 0)
        for index, card in enumerate(cards):
            grid.addWidget(card, index // columns, index % columns)

    def _relayout_dashboard_insights(self) -> None:
        grid = getattr(self, "dashboard_insights_grid", None)
        cards = getattr(self, "dashboard_insight_cards", None)
        if grid is None or not cards:
            return

        for card in cards:
            grid.removeWidget(card)
        wide = self.width() >= 1240
        columns = 3 if wide else 2
        for column in range(3):
            grid.setColumnStretch(column, 1 if column < columns else 0)
        if wide:
            self.dashboard_attention_card.setFixedHeight(382)
            self.dashboard_attention_list.setMinimumHeight(286)
            self.dashboard_attention_list.setMaximumHeight(286)
            for card in (self.dashboard_area_card, self.dashboard_sponsor_card):
                card.setFixedHeight(183)
                card._dashboard_list_widget.setMinimumHeight(108)
                card._dashboard_list_widget.setMaximumHeight(108)
            grid.addWidget(self.dashboard_attention_card, 0, 0, 2, 2)
            grid.addWidget(self.dashboard_area_card, 0, 2)
            grid.addWidget(self.dashboard_sponsor_card, 1, 2)
            return

        for card in cards:
            card.setFixedHeight(196 if card is not self.dashboard_attention_card else 210)
            card._dashboard_list_widget.setMinimumHeight(116)
            card._dashboard_list_widget.setMaximumHeight(130)
        grid.addWidget(self.dashboard_area_card, 0, 0)
        grid.addWidget(self.dashboard_sponsor_card, 0, 1)
        grid.addWidget(self.dashboard_attention_card, 1, 0, 1, 2)

    def _build_dashboard_list_card(
        self,
        title_text,
        caption_text,
        list_widget,
        *,
        role="summary",
        eyebrow_text="",
    ):
        card = Card()
        card.setObjectName("DashboardListCard")
        card.setProperty("role", role)
        card.setMinimumWidth(0)
        layout = QVBoxLayout(card)
        margins = (14, 14, 14, 10) if role == "attention" else (4, 10, 4, 8)
        layout.setContentsMargins(*margins)
        layout.setSpacing(6)

        eyebrow = QLabel(eyebrow_text)
        eyebrow.setObjectName("InsightEyebrow")
        eyebrow.setVisible(bool(eyebrow_text))

        title = StrongBodyLabel(title_text)
        title.setObjectName("CardHeading")
        caption = QLabel(caption_text)
        caption.setObjectName("Caption")
        caption.setWordWrap(True)
        list_widget.setFrameShape(QFrame.Shape.NoFrame)
        list_widget.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        list_widget.setProperty("dashboard", True)

        layout.addWidget(eyebrow)
        layout.addWidget(title)
        layout.addWidget(caption)
        layout.addWidget(list_widget, 1)
        card._dashboard_list_widget = list_widget
        card._dashboard_caption_label = caption
        card.setAccessibleName(title_text)
        self._dashboard_insight_headers.append(title)
        return card

    def refresh_dashboard(self):
        """Fetch dashboard data without blocking Qt's event loop."""
        self._dashboard_request += 1
        request_id = self._dashboard_request
        for label in (self.lbl_total_val, self.lbl_active_val, self.lbl_inactive_val, self.lbl_graduated_val):
            label.setText("—")
        self.dashboard_state_label.setText("Updating overview…")
        self.dashboard_state_label.setProperty("state", "loading")
        self.dashboard_state_label.style().unpolish(self.dashboard_state_label)
        self.dashboard_state_label.style().polish(self.dashboard_state_label)

        def fetch_rows():
            raw_rows = self.student_repository.list_students(columns=(
                "id,last_name,first_name,status,sponsor,area,grade,contact,"
                "school,birthday,photo_url,gender,address,city,parents,course,"
                "remarks,sheet_synced_at,source_sheet_name"
            ))
            roster_rows = self.dashboard_service.latest_sync_cohort(raw_rows)
            rows = self.dashboard_service.dedupe_students(roster_rows)
            student_ids = [row.get("id") for row in rows if row.get("id")]
            school_year = self._current_school_year()
            summaries = self.expense_service.get_financial_summaries(
                student_ids, school_year
            )
            return raw_rows, rows, summaries, school_year

        def apply_rows(result):
            if request_id != self._dashboard_request:
                return
            raw_rows, rows, summaries, school_year = result
            # Sync preserves older/manual rows for history. The overview uses
            # only the newest transaction cohort so its totals match the
            # latest workbook result on every office PC.
            has_sync_metadata = any(
                str(row.get("sheet_synced_at") or "").strip()
                for row in raw_rows
            )
            counts = self.dashboard_service.summary_counts(rows)
            animate_count(
                self.lbl_total_val,
                counts["total"],
                motion_enabled=not self._reduce_motion,
            )
            animate_count(
                self.lbl_active_val,
                counts["active"],
                motion_enabled=not self._reduce_motion,
            )
            animate_count(
                self.lbl_inactive_val,
                counts["inactive"],
                motion_enabled=not self._reduce_motion,
            )
            animate_count(
                self.lbl_graduated_val,
                counts["graduated"],
                motion_enabled=not self._reduce_motion,
            )

            active_students = [
                row
                for row in rows
                if self.dashboard_service.status_bucket(row.get("status")) == "active"
            ]
            dashboard = self.dashboard_service.build_lists(active_students)
            pending = int(dashboard.get("attention_count") or 0)
            animate_count(
                self.dashboard_pending_value,
                pending,
                motion_enabled=not self._reduce_motion,
            )

            total_budget = sum(
                float(summary.get("total_budget") or 0)
                for summary in summaries.values()
            )
            total_spent = sum(
                float(summary.get("total_expenses") or 0)
                for summary in summaries.values()
            )
            usage = self.expense_service.budget_usage({
                "total_budget": total_budget,
                "total_expenses": total_spent,
                "remaining_balance": total_budget - total_spent,
            })
            self.dashboard_budget_value.setText(f"PHP {total_spent:,.0f}")
            self.dashboard_budget_spent.setText(f"PHP {total_spent:,.0f}")
            self.dashboard_budget_of.setText(f"used of PHP {total_budget:,.0f}")
            self.dashboard_budget_percent.setText(f"{usage['percent']}% used")
            self.dashboard_budget_remaining.setText(
                f"PHP {usage['remaining']:,.0f} remaining"
                if total_budget > 0 else "Budget unallocated this year"
            )
            self.dashboard_budget_scale_max.setText(f"PHP {total_budget:,.0f}")
            elapsed_months = max(1, ((datetime.now().month - 6) % 12) + 1)
            self.dashboard_monthly_pace.setText(
                f"PHP {total_spent / elapsed_months:,.0f}"
            )
            state = usage["state"]
            self.dashboard_budget_bar.setProperty("state", state)
            self.dashboard_budget_bar.style().unpolish(self.dashboard_budget_bar)
            self.dashboard_budget_bar.style().polish(self.dashboard_budget_bar)
            animate_progress(
                self.dashboard_budget_bar,
                usage["percent"],
                motion_enabled=not self._reduce_motion,
            )
            health_text = {
                "success": "Healthy",
                "warning": "Watch closely",
                "danger": "Over budget",
                "neutral": "Unallocated",
            }.get(state, "Unallocated")
            self.dashboard_budget_health.set_state(state, health_text)

            progress_counts = {"complete": 0, "progress": 0, "review": 0}
            for student in active_students:
                completion = self._profile_completion_percent(student)
                if completion >= 90:
                    progress_counts["complete"] += 1
                elif completion >= 60:
                    progress_counts["progress"] += 1
                else:
                    progress_counts["review"] += 1
            active_total = max(1, len(active_students))
            for key, count in progress_counts.items():
                value, bar = self.dashboard_progress_rows[key]
                value.setText(str(count))
                animate_progress(
                    bar,
                    round(count * 100 / active_total),
                    motion_enabled=not self._reduce_motion,
                )

            self._populate_dashboard_recent(rows, summaries)
            record_kind = "synced" if has_sync_metadata else "database"
            self.dashboard_state_label.setText(
                f"Updated just now  /  {counts['total']} {record_kind} records"
            )
            self.dashboard_state_label.setProperty("state", "success")
            self.dashboard_state_label.style().unpolish(self.dashboard_state_label)
            self.dashboard_state_label.style().polish(self.dashboard_state_label)
            fade_in(
                self.dashboard_state_label,
                motion_enabled=not self._reduce_motion,
                duration_ms=160,
            )
            self._mark_database_updated()

        def show_error(error):
            if request_id != self._dashboard_request:
                return
            for label in (
                self.lbl_total_val,
                self.lbl_active_val,
                self.lbl_inactive_val,
                self.lbl_graduated_val,
            ):
                label.setText("—")
            self.dashboard_state_label.setText("Overview unavailable  /  refresh to retry")
            self.dashboard_state_label.setProperty("state", "danger")
            self.dashboard_state_label.style().unpolish(self.dashboard_state_label)
            self.dashboard_state_label.style().polish(self.dashboard_state_label)
            self._set_connection_state("Connection issue", "danger")
            self.status_bar.showMessage(
                "Could not update the dashboard. Check the office connection and refresh.",
                8000,
            )
            logging.getLogger(__name__).error("Dashboard refresh failed:\n%s", error)

        return self._run_background(fetch_rows, apply_rows, show_error)

    @staticmethod
    def _current_school_year() -> str:
        now = datetime.now()
        start = now.year if now.month >= 6 else now.year - 1
        return f"{start}-{start + 1}"

    def _populate_dashboard_recent(self, rows, summaries) -> None:
        table = self.dashboard_recent_table
        ordered = sorted(
            rows,
            key=lambda row: (
                str(row.get("sheet_synced_at") or ""),
                str(row.get("last_name") or ""),
            ),
            reverse=True,
        )[:3]
        table.setRowCount(len(ordered))
        for row_index, student in enumerate(ordered):
            full_status, status_text, _ = self._status_style(student.get("status"))
            gender = str(student.get("gender") or "").strip().upper()
            name = f"{student.get('last_name', '')}, {student.get('first_name', '')}".strip(", ")
            if gender:
                name += f" ({gender})"
            completion = self._profile_completion_percent(student)
            budget = self.expense_service.budget_card_status(
                summaries.get(student.get("id"))
            )
            budget_text = budget.get("detail") or "Unallocated this year"
            updated = str(student.get("sheet_synced_at") or "").strip()
            if updated:
                try:
                    updated = datetime.fromisoformat(
                        updated.replace("Z", "+00:00")
                    ).strftime("%b %d").replace(" 0", " ")
                except ValueError:
                    updated = updated[:10]
            else:
                updated = "Current"
            values = [name or "Unnamed student", status_text, f"{completion}%", budget_text, updated]
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, student.get("id"))
                table.setItem(row_index, column, item)
            table.setRowHeight(row_index, 38)
        table.setAccessibleName("Recent synchronized student updates")

    def _open_dashboard_recent(self, row, _column):
        item = self.dashboard_recent_table.item(row, 0)
        if item is None:
            return
        student_id = item.data(Qt.ItemDataRole.UserRole)
        if student_id:
            self._open_student_profile(str(student_id))
    def _filter_current_master_rows(self, rows):
        with self._workbook_lock:
            return self.masterlist_service.filter_current_rows(
                rows,
                self._current_master_student_reference(),
            )

    def _apply_current_master_status(self, row):
        return self.masterlist_service.apply_current_status(
            row,
            self._current_master_student_reference(),
        )

    def _current_master_student_keys(self):
        return self.masterlist_service.current_student_keys(
            self._current_master_student_reference(),
        )

    def _current_master_student_reference(self):
        try:
            return self.masterlist_service.current_student_reference(
                workbook=self._workbook,
                workbook_path=self._workbook_path,
                workbook_revision=self._workbook_revision,
                saved_workbook_path=self._settings().value("workbook_path", "", type=str),
                cwd=os.getcwd(),
                fallback_paths=[],
            )
        except Exception as e:
            print(f"Masterlist reference error: {e}")
            return {}

    def _invalidate_master_reference_cache(self):
        self._master_ref_cache_key = None
        self._master_ref_cache = None
        self.masterlist_service.invalidate_cache()

    def _current_masterlist_path(self):
        return self.masterlist_service.current_masterlist_path(
            workbook_path=self._workbook_path,
            saved_workbook_path=self._settings().value("workbook_path", "", type=str),
            cwd=os.getcwd(),
            fallback_paths=[],
        )

    def _latest_master_sheet_name_from_names(self, sheet_names):
        return self.masterlist_service.latest_master_sheet_name_from_names(sheet_names)

    def _master_sheet_student_keys(self, worksheet):
        return self.masterlist_service.master_sheet_student_keys(worksheet)

    def _master_sheet_student_reference(self, worksheet):
        return self.masterlist_service.master_sheet_student_reference(worksheet)

    def _student_reference_key(self, row):
        return self.masterlist_service.student_reference_key(row)

    def _normalize_student_key_value(self, value):
        return self.masterlist_service.normalize_student_key_value(value)
    def _status_style(self, status):
        return self.student_service.status_style(status)

    def _profile_status_color(self, status_key):
        return {
            "active": theme_color("success"),
            "inactive": theme_color("danger"),
            "graduated": theme_color("graduated"),
        }.get(status_key, theme_color("danger"))

    def _profile_completion_percent(self, student):
        return self.student_service.profile_completion_percent(student)

    def _dedupe_dashboard_students(self, rows):
        return self.dashboard_service.dedupe_students(rows)

    def _dashboard_student_key(self, row):
        return self.dashboard_service.student_key(row)

    def _dashboard_row_score(self, row):
        return self.dashboard_service.row_score(row)

    def _refresh_dashboard_lists(self, rows):
        dashboard = self.dashboard_service.build_lists(rows)
        area_total = sum(count for _, count in dashboard["area_counts"])
        sponsor_total = sum(count for _, count in dashboard["sponsor_counts"])
        self.dashboard_area_caption.setText(
            f"{area_total} active students across "
            f"{len(dashboard['area_counts'])} areas"
        )
        self.dashboard_sponsor_caption.setText(
            f"{sponsor_total} active students across "
            f"{len(dashboard['sponsor_counts'])} sponsor groups"
        )
        self.dashboard_attention_caption.setText(
            f"{dashboard['attention_count']} active profiles are missing "
            "an important detail"
        )

        self.dashboard_area_list.clear()
        for area, count in dashboard["area_counts"]:
            item = QListWidgetItem(f"{area}  /  {count}")
            item.setData(Qt.ItemDataRole.UserRole, area)
            item.setToolTip(f"{count} students in {area}")
            self.dashboard_area_list.addItem(item)
        if not dashboard["area_counts"]:
            self._add_dashboard_empty_item(
                self.dashboard_area_list, "No active student areas yet"
            )

        self.dashboard_sponsor_list.clear()
        for sponsor, count in dashboard["sponsor_counts"]:
            item = QListWidgetItem(f"{sponsor}  /  {count}")
            item.setData(Qt.ItemDataRole.UserRole, sponsor)
            item.setToolTip(f"{count} students for {sponsor}")
            self.dashboard_sponsor_list.addItem(item)
        if not dashboard["sponsor_counts"]:
            self._add_dashboard_empty_item(
                self.dashboard_sponsor_list, "No active sponsors yet"
            )

        self.dashboard_attention_list.clear()
        attention = dashboard["attention"]
        for entry in attention:
            item = QListWidgetItem(f"{entry['name']}    Missing: {entry['missing_text']}")
            item.setData(Qt.ItemDataRole.UserRole, entry["student_id"])
            item.setToolTip("Open student profile")
            self.dashboard_attention_list.addItem(item)
        if not attention:
            self._add_dashboard_empty_item(
                self.dashboard_attention_list,
                "All active profiles look complete",
            )

    def _add_dashboard_empty_item(self, widget, text):
        item = QListWidgetItem(text)
        item.setData(Qt.ItemDataRole.UserRole, None)
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
        item.setForeground(theme_color("text_secondary"))
        widget.addItem(item)

    def _open_dashboard_area(self, item):
        area = item.data(Qt.ItemDataRole.UserRole)
        if not area:
            return
        self._set_active_nav(self.btn_stud)
        self.student_list_view._select_area(str(area))
        self._switch_page(1)

    def _open_dashboard_sponsor(self, item):
        sponsor = item.data(Qt.ItemDataRole.UserRole)
        if not sponsor:
            return
        self.student_list_view.search_sponsor.setText(str(sponsor))
        self._set_active_nav(self.btn_stud)
        self._switch_page(1)
    # ── SCREEN 2: PROFILE ─────────────────────────────────────────────────────
    def _create_profile_screen_legacy(self):
        widget = QWidget()
        outer = QVBoxLayout(widget)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(16)

        top_bar = QHBoxLayout()
        back_btn = QPushButton("Back"); back_btn.setObjectName("SecondaryBtn")
        set_content_hugging_button(back_btn)
        back_btn.clicked.connect(self.nav_students)
        self.edit_btn = PrimaryPushButton("Edit profile"); self.edit_btn.clicked.connect(self.open_edit_screen)
        set_content_hugging_button(self.edit_btn)
        self.deactivate_btn = QPushButton("Mark inactive"); self.deactivate_btn.setObjectName("WarningBtn")
        set_content_hugging_button(self.deactivate_btn)
        self.deactivate_btn.clicked.connect(self.toggle_active_status)
        self.remove_student_btn = QPushButton("Remove")
        self.remove_student_btn.setObjectName("DangerBtn")
        set_content_hugging_button(self.remove_student_btn)
        self.remove_student_btn.setToolTip("Permanently remove this student and related records")
        self.remove_student_btn.clicked.connect(self.remove_current_student)
        
        top_bar.addWidget(back_btn); top_bar.addStretch()
        top_bar.addWidget(self.edit_btn)
        top_bar.addWidget(self.deactivate_btn)
        top_bar.addWidget(self.remove_student_btn)
        outer.addLayout(top_bar)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        content = Card()
        content.setProperty("component", "profile")
        layout = QVBoxLayout(content)
        layout.setContentsMargins(18, 16, 18, 18)
        layout.setSpacing(14)

        top_row = QHBoxLayout()
        top_row.setSpacing(16)
        top_row.setAlignment(Qt.AlignmentFlag.AlignTop)

        profile_accent = QFrame()
        profile_accent.setObjectName("ProfileAccent")
        profile_accent.setFixedWidth(4)
        profile_accent.setMinimumHeight(250)
        top_row.addWidget(profile_accent)
        
        # Left side: Photo
        pcol = QVBoxLayout()
        pcol.setSpacing(8)
        pcol.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.photo_label = QLabel("No photo")
        self.photo_label.setFixedSize(112, 136)
        self.photo_label.setObjectName("PhotoFrame")
        self.photo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.change_photo_btn = QPushButton("Change photo"); self.change_photo_btn.setObjectName("SecondaryBtn")
        self.change_photo_btn.clicked.connect(self.change_photo)
        set_content_hugging_button(self.change_photo_btn)
        self.remove_photo_btn = QPushButton("Remove photo"); self.remove_photo_btn.setObjectName("DangerBtn")
        self.remove_photo_btn.clicked.connect(self.remove_photo)
        set_content_hugging_button(self.remove_photo_btn)
        self.remove_photo_btn.setVisible(False)
        pcol.addWidget(self.photo_label)
        pcol.addWidget(
            self.change_photo_btn,
            alignment=Qt.AlignmentFlag.AlignHCenter,
        )
        pcol.addWidget(
            self.remove_photo_btn,
            alignment=Qt.AlignmentFlag.AlignHCenter,
        )
        top_row.addLayout(pcol)
        
        # Middle: Info (native QGridLayout)
        info_layout = QVBoxLayout()
        info_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        info_layout.setSpacing(14)
        
        # Header Row: Name & Status
        self.lbl_profile_name = TitleLabel("Student name")
        self.lbl_profile_name.setObjectName("ProfileTitle")
        self.lbl_profile_name.setWordWrap(True)
        self.lbl_profile_name.setMinimumWidth(0)
        self.lbl_profile_name.setSizePolicy(
            QSizePolicy.Policy.Ignored,
            QSizePolicy.Policy.Preferred,
        )
        self.lbl_profile_status = QLabel("Active")
        self.lbl_profile_status.setObjectName("StatusLabel")
        self.lbl_profile_status.setMinimumHeight(28)
        self.lbl_profile_status.setTextFormat(Qt.TextFormat.RichText)
        
        header_row = QHBoxLayout()
        header_row.setSpacing(18)
        header_text = QVBoxLayout()
        header_text.setSpacing(4)
        header_text.addWidget(self.lbl_profile_name)
        header_text.addWidget(self.lbl_profile_status)
        header_row.addLayout(header_text, 1)

        progress_layout = QVBoxLayout()
        progress_layout.setSpacing(6)
        progress_layout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        self.profile_progress = CircularProgress(size=70)
        self.profile_progress.setFixedSize(70, 70)
        progress_lbl = StrongBodyLabel("Profile\ncompletion")
        progress_lbl.setObjectName("CaptionStrong")
        progress_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        progress_layout.addWidget(self.profile_progress, alignment=Qt.AlignmentFlag.AlignHCenter)
        progress_layout.addWidget(progress_lbl, alignment=Qt.AlignmentFlag.AlignHCenter)
        header_row.addLayout(progress_layout)
        info_layout.addLayout(header_row)

        details_lbl = StrongBodyLabel("Student details")
        details_lbl.setObjectName("CardHeading")
        info_layout.addWidget(details_lbl)

        # Data Grid Layout — one field per row, value column takes all
        # remaining width. A single column avoids two text blocks ever
        # competing for width on the same row, so this never forces a
        # horizontal scrollbar, no matter how the window is resized.
        self.profile_grid = QGridLayout()
        self.profile_grid.setHorizontalSpacing(16)
        self.profile_grid.setVerticalSpacing(10)
        self.profile_grid.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.profile_grid.setColumnMinimumWidth(0, 78)
        self.profile_grid.setColumnMinimumWidth(2, 78)
        self.profile_grid.setColumnStretch(1, 1)
        self.profile_grid.setColumnStretch(3, 1)

        # Dictionary to store our dynamic value labels for easy updating
        self.profile_data_labels = {}

        # Helper function to create a uniform row in the grid
        def add_grid_item(row, column, label_text, key):
            grid_col = column * 2
            title_lbl = QLabel(label_text)
            title_lbl.setObjectName("FieldLabel")
            val_lbl = QLabel("--")
            val_lbl.setObjectName("FieldValue")
            val_lbl.setWordWrap(True)
            val_lbl.setMinimumWidth(0)
            val_lbl.setSizePolicy(
                QSizePolicy.Policy.Ignored,
                QSizePolicy.Policy.Preferred,
            )
            self.profile_grid.addWidget(title_lbl, row, grid_col, Qt.AlignmentFlag.AlignTop)
            self.profile_grid.addWidget(val_lbl, row, grid_col + 1)
            self.profile_data_labels[key] = val_lbl

        # Grid Population
        add_grid_item(0, 0, "Gender", "gender")
        add_grid_item(0, 1, "Grade / level", "grade")
        add_grid_item(1, 0, "Area", "area")
        add_grid_item(1, 1, "City", "city")
        add_grid_item(2, 0, "Address", "address")
        add_grid_item(2, 1, "Birthday", "birthday")
        add_grid_item(3, 0, "Contact", "contact")
        add_grid_item(3, 1, "Sponsor", "sponsor")
        self.profile_data_labels["sponsor"].setProperty("emphasis", True)
        
        add_grid_item(4, 0, "School", "school")
        add_grid_item(4, 1, "Parents", "parents")
        add_grid_item(5, 0, "Course", "course")

        info_layout.addLayout(self.profile_grid)
        top_row.addLayout(info_layout, 1)

        top_row.setContentsMargins(14, 14, 14, 14)
        profile_summary = QWidget()
        profile_summary.setObjectName("ProfileSummary")
        profile_summary.setLayout(top_row)
        profile_summary.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Maximum,
        )
        layout.addWidget(profile_summary)

        divider = QFrame()
        divider.setObjectName("Divider")
        divider.setFrameShape(QFrame.Shape.HLine)
        layout.addWidget(divider)

        # Remarks
        remarks_header = QHBoxLayout()
        remarks_header.setSpacing(10)
        remarks_lbl = StrongBodyLabel("Remarks")
        remarks_lbl.setObjectName("CardHeading")
        remarks_copy = QLabel("Context shared with the student-support office")
        remarks_copy.setObjectName("Caption")
        remarks_title = QVBoxLayout()
        remarks_title.setSpacing(2)
        remarks_title.addWidget(remarks_lbl)
        remarks_title.addWidget(remarks_copy)
        remarks_header.addLayout(remarks_title)
        remarks_header.addStretch()
        self.save_remarks_btn = PrimaryPushButton("Save remarks")
        self.save_remarks_btn.clicked.connect(self.save_remarks)
        self.save_remarks_btn.setEnabled(False)
        set_content_hugging_button(self.save_remarks_btn)
        exp_btn  = QPushButton("Expenses")
        exp_btn.setObjectName("WarningBtn")
        exp_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        set_content_hugging_button(exp_btn)
        exp_btn.clicked.connect(self.nav_expenses)
        remarks_header.addWidget(self.save_remarks_btn)
        remarks_header.addWidget(exp_btn)
        layout.addLayout(remarks_header)

        self.remarks_edit = QTextEdit()
        self.remarks_edit.setFixedHeight(128)
        self.remarks_edit.setObjectName("RemarksEditor")
        self.remarks_edit.setAccessibleName("Student office remarks")
        self.remarks_edit.textChanged.connect(self._update_remarks_dirty_state)
        self._loaded_remarks = ""
        layout.addWidget(self.remarks_edit)

        scroll.setWidget(content); outer.addWidget(scroll)
        self.stacked_widget.addWidget(widget)

    # ── SCREEN 3: ADD / EDIT ──────────────────────────────────────────────────
    def create_profile_screen(self):
        """Build the compact student profile approved in Figma."""
        widget = QWidget()
        outer = QVBoxLayout(widget)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(20)

        # Compatibility actions stay available to the profile loader and
        # shortcuts; visible navigation lives in the shared Figma header.
        self.edit_btn = ActionButton("Edit record", widget)
        self.edit_btn.clicked.connect(self.open_edit_screen)
        self.edit_btn.hide()
        self.deactivate_btn = ActionButton("Mark inactive", widget, variant="tertiary")
        self.deactivate_btn.clicked.connect(self.toggle_active_status)
        self.deactivate_btn.hide()
        self.remove_student_btn = ActionButton("Remove", widget, variant="danger")
        self.remove_student_btn.clicked.connect(self.remove_current_student)
        self.remove_student_btn.hide()
        profile_menu = QMenu(widget)
        self.profile_change_photo_action = profile_menu.addAction("Change photo")
        self.profile_change_photo_action.triggered.connect(self.change_photo)
        self.profile_remove_photo_action = profile_menu.addAction("Remove photo")
        self.profile_remove_photo_action.triggered.connect(self.remove_photo)
        self.profile_remove_photo_action.setVisible(False)
        profile_menu.addSeparator()
        self.profile_status_action = profile_menu.addAction("Mark inactive")
        self.profile_status_action.triggered.connect(self.toggle_active_status)
        self.profile_remove_action = profile_menu.addAction("Remove student")
        self.profile_remove_action.triggered.connect(self.remove_current_student)

        summary = QFrame()
        summary.setObjectName("ProfileSummary")
        summary.setFixedHeight(126)
        summary_layout = QHBoxLayout(summary)
        summary_layout.setContentsMargins(0, 15, 44, 15)
        summary_layout.setSpacing(0)
        self.profile_summary_layout = summary_layout
        accent = QFrame()
        accent.setObjectName("ProfileAccent")
        accent.setFixedWidth(4)
        self.profile_summary_accent = accent
        summary_layout.addWidget(accent)

        self.profile_accent_gap = QSpacerItem(
            20, 0, QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Minimum
        )
        summary_layout.addSpacerItem(self.profile_accent_gap)

        self.photo_label = QLabel("Photo")
        self.photo_label.setObjectName("ProfileAvatar")
        self.photo_label.setFixedSize(96, 96)
        self.photo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        summary_layout.addWidget(self.photo_label)
        summary_layout.addSpacing(18)

        identity_widget = QWidget()
        identity_widget.setObjectName("ProfileIdentity")
        identity = QVBoxLayout(identity_widget)
        identity.setContentsMargins(0, 0, 0, 0)
        identity.setSpacing(2)
        self.lbl_profile_name = TitleLabel("Student name")
        self.lbl_profile_name.setObjectName("ProfileTitle")
        self.lbl_profile_name.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred
        )
        self.profile_meta_label = QLabel("Area not set  /  Sponsor not set")
        self.profile_meta_label.setObjectName("Caption")
        self.profile_meta_label.setSizePolicy(
            QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred
        )
        self.lbl_profile_status = QLabel("Active")
        self.lbl_profile_status.setObjectName("ProfileStatusBadge")
        self.lbl_profile_status.setFixedHeight(24)
        self.lbl_profile_status.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        identity.addWidget(self.lbl_profile_name)
        identity.addWidget(self.profile_meta_label)
        identity.addWidget(self.lbl_profile_status)
        identity.addStretch()
        summary_layout.addWidget(identity_widget, 1)

        self.profile_metric_widget = QWidget()
        self.profile_metric_widget.setObjectName("ProfileMetric")
        self.profile_metric_widget.setFixedWidth(180)
        profile_metric = QVBoxLayout(self.profile_metric_widget)
        profile_metric.setContentsMargins(0, 3, 0, 0)
        profile_metric.setSpacing(5)
        profile_label = QLabel("PROFILE")
        profile_label.setObjectName("ProfileMetricLabel")
        self.profile_completion_label = StrongBodyLabel("0%")
        self.profile_completion_label.setObjectName("ProfileMetricValue")
        self.profile_progress = NativeProgressBar()
        self.profile_progress.setObjectName("ProfileLinearProgress")
        self.profile_progress.setRange(0, 100)
        self.profile_progress.setValue(0)
        self.profile_progress.setTextVisible(False)
        self.profile_progress.setFixedHeight(7)
        profile_metric.addWidget(profile_label)
        profile_metric.addWidget(self.profile_completion_label)
        profile_metric.addWidget(self.profile_progress)
        profile_metric.addStretch()
        summary_layout.addWidget(self.profile_metric_widget)

        self.profile_metric_gap = QSpacerItem(
            44, 0, QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Minimum
        )
        summary_layout.addSpacerItem(self.profile_metric_gap)

        self.profile_budget_widget = QWidget()
        self.profile_budget_widget.setObjectName("ProfileMetric")
        self.profile_budget_widget.setFixedWidth(244)
        budget_metric = QVBoxLayout(self.profile_budget_widget)
        budget_metric.setContentsMargins(0, 3, 0, 0)
        budget_metric.setSpacing(5)
        budget_label = QLabel("BUDGET")
        budget_label.setObjectName("ProfileMetricLabel")
        self.profile_budget_label = QLabel("Unallocated this year")
        self.profile_budget_label.setObjectName("ProfileBudgetValue")
        self.profile_budget_bar = NativeProgressBar()
        self.profile_budget_bar.setObjectName("BudgetProgress")
        self.profile_budget_bar.setRange(0, 100)
        self.profile_budget_bar.setValue(0)
        self.profile_budget_bar.setTextVisible(False)
        self.profile_budget_bar.setFixedHeight(7)
        budget_metric.addWidget(budget_label)
        budget_metric.addWidget(self.profile_budget_label)
        budget_metric.addWidget(self.profile_budget_bar)
        budget_metric.addStretch()
        summary_layout.addWidget(self.profile_budget_widget)
        outer.addWidget(summary)

        self.profile_data_labels = {}
        self.profile_field_grids = []

        def info_card(title_text, fields, *, label_width=140):
            card = Card()
            card.setObjectName("ProfileInfoCard")
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(20, 16, 20, 16)
            card_layout.setSpacing(6)
            title = StrongBodyLabel(title_text)
            title.setObjectName("CardHeading")
            card_layout.addWidget(title)
            grid = QGridLayout()
            grid.setHorizontalSpacing(18)
            grid.setVerticalSpacing(0)
            grid.setColumnMinimumWidth(0, label_width)
            grid.setColumnStretch(1, 1)
            for row, (label_text, key) in enumerate(fields):
                label = QLabel(label_text)
                label.setObjectName("ProfileFieldLabel")
                value = QLabel("--")
                value.setObjectName("ProfileFieldValue")
                value.setWordWrap(True)
                value.setMinimumWidth(0)
                grid.addWidget(label, row, 0, Qt.AlignmentFlag.AlignTop)
                grid.addWidget(value, row, 1, Qt.AlignmentFlag.AlignTop)
                grid.setRowMinimumHeight(row, 40)
                self.profile_data_labels[key] = value
            self.profile_field_grids.append(grid)
            card_layout.addLayout(grid)
            card_layout.addStretch()
            return card

        cards = QGridLayout()
        self.profile_cards_layout = cards
        cards.setHorizontalSpacing(18)
        cards.setVerticalSpacing(20)
        student_card = info_card("Student details", (
            ("Grade / year", "grade"), ("Birthday", "birthday"),
            ("School", "school"),
            ("Course / track", "course"),
        ))
        support_card = info_card("Support details", (
            ("Status", "support_status"), ("Area", "area"),
            ("Sponsor", "sponsor"), ("School year", "school_year"),
        ), label_width=142)
        contact_card = info_card("Contact and family", (
            ("Contact", "contact"), ("Parent / guardian", "parents"),
            ("City", "city"), ("Address", "address"),
        ))

        notes_card = Card()
        notes_card.setObjectName("ProfileInfoCard")
        notes_layout = QVBoxLayout(notes_card)
        notes_layout.setContentsMargins(20, 16, 20, 26)
        notes_layout.setSpacing(0)
        notes_title = StrongBodyLabel("Office notes")
        notes_title.setObjectName("CardHeading")
        notes_layout.addWidget(notes_title)
        notes_layout.addSpacing(8)
        remarks_label = QLabel("REMARKS")
        remarks_label.setObjectName("ProfileRemarksLabel")
        notes_layout.addWidget(remarks_label)
        notes_layout.addSpacing(6)
        self.remarks_edit = QTextEdit()
        self.remarks_edit.setObjectName("RemarksEditor")
        self.remarks_edit.setPlaceholderText("Add a concise office note")
        self.remarks_edit.textChanged.connect(self._update_remarks_dirty_state)
        self._loaded_remarks = ""
        self.remarks_edit.hide()
        self.remarks_display = QLabel("No office notes recorded.")
        self.remarks_display.setObjectName("RemarksText")
        self.remarks_display.setWordWrap(True)
        self.remarks_display.setAlignment(Qt.AlignmentFlag.AlignTop)
        notes_layout.addWidget(self.remarks_display, 1)
        next_action = QFrame()
        next_action.setObjectName("NextActionCallout")
        next_layout = QVBoxLayout(next_action)
        next_layout.setContentsMargins(12, 8, 12, 8)
        next_layout.setSpacing(2)
        next_title = QLabel("NEXT ACTION")
        next_title.setObjectName("ProfileNextActionTitle")
        self.profile_next_action_copy = QLabel(
            "Review during the next office follow-up."
        )
        self.profile_next_action_copy.setObjectName("ProfileNextActionCopy")
        self.profile_next_action_copy.setWordWrap(True)
        next_layout.addWidget(next_title)
        next_layout.addWidget(self.profile_next_action_copy)
        next_action.setFixedHeight(58)
        notes_layout.addWidget(next_action)
        self.save_remarks_btn = ActionButton("Save notes")
        self.save_remarks_btn.clicked.connect(self.save_remarks)
        self.save_remarks_btn.setEnabled(False)
        self.save_remarks_btn.hide()
        expense_btn = ActionButton("Expenses", variant="secondary")
        expense_btn.clicked.connect(self.nav_expenses)
        expense_btn.hide()
        self.change_photo_btn = ActionButton("Change photo", widget, variant="tertiary")
        self.change_photo_btn.clicked.connect(self.change_photo)
        self.change_photo_btn.hide()
        self.remove_photo_btn = ActionButton("Remove photo", widget, variant="tertiary")
        self.remove_photo_btn.clicked.connect(self.remove_photo)
        self.remove_photo_btn.hide()
        for button in (
            self.save_remarks_btn, expense_btn, self.change_photo_btn,
            self.remove_photo_btn, self.deactivate_btn, self.remove_student_btn,
        ):
            set_content_hugging_button(button, height=34)
        student_card.setFixedHeight(222)
        support_card.setFixedHeight(222)
        contact_card.setFixedHeight(220)
        notes_card.setFixedHeight(220)
        self.profile_top_cards = (student_card, support_card)
        self.profile_bottom_cards = (contact_card, notes_card)
        cards.addWidget(student_card, 0, 0)
        cards.addWidget(support_card, 0, 1)
        cards.addWidget(contact_card, 1, 0)
        cards.addWidget(notes_card, 1, 1)
        cards.setColumnStretch(0, 520)
        cards.setColumnStretch(1, 548)
        outer.addLayout(cards)
        outer.addStretch()
        self.stacked_widget.addWidget(widget)
        self._update_profile_layout()

    def create_add_screen(self):
        widget = QWidget()
        outer = QVBoxLayout(widget)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self.form_title_label = QLabel("New student record", widget)
        self.form_title_label.hide()
        self.add_photo_label = QLabel("No photo", widget)
        self.add_photo_label.hide()

        content = Card()
        content.setProperty("component", "studentForm")
        content.setFixedHeight(608)
        shell = QHBoxLayout(content)
        shell.setContentsMargins(0, 0, 0, 0)
        shell.setSpacing(0)
        accent = QFrame()
        accent.setObjectName("FormAccent")
        accent.setFixedWidth(4)
        shell.addWidget(accent)
        form = QWidget()
        layout = QVBoxLayout(form)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)
        shell.addWidget(form, 1)

        self.inp_last    = QLineEdit()
        self.inp_first   = QLineEdit()
        self.inp_gender  = QComboBox(); self.inp_gender.addItems(["F", "M", ""])
        self.inp_grade   = QLineEdit()
        self.inp_address = QLineEdit()
        self.inp_city    = QLineEdit()
        self.inp_area    = QLineEdit()
        self.inp_bday    = QDateEdit(); self.inp_bday.setCalendarPopup(True); self.inp_bday.setDisplayFormat("yyyy-MM-dd")
        self.inp_bday.setMinimumDate(QDate(1900, 1, 1))
        self.inp_bday.setSpecialValueText("Not set")
        self.inp_bday.setDate(self.inp_bday.minimumDate())
        self.inp_sponsor = QLineEdit(); self.inp_sponsor.setPlaceholderText("e.g. Schnurbein, Word of Life...")
        self.inp_contact = QLineEdit()
        self.inp_school  = QLineEdit()
        self.inp_parents = QLineEdit()
        self.inp_course  = QLineEdit()
        self.inp_remarks = QTextEdit(); self.inp_remarks.setFixedHeight(54)
        self.inp_status  = QComboBox(); self.inp_status.addItems(["Active", "Inactive/Removed", "Graduated"])

        self.inp_last.setPlaceholderText("Family name")
        self.inp_first.setPlaceholderText("Given name")
        self.inp_grade.setPlaceholderText("Grade or year level")
        self.inp_address.setPlaceholderText("Home address")
        self.inp_city.setPlaceholderText("City")
        self.inp_area.setPlaceholderText("Area or coordinator")
        self.inp_contact.setPlaceholderText("Phone number")
        self.inp_school.setPlaceholderText("School")
        self.inp_parents.setPlaceholderText("Names and occupations")
        self.inp_course.setPlaceholderText("Course, if applicable")
        self.inp_remarks.setPlaceholderText(
            "Important context for the office team"
        )

        def field_block(label_text, field, *, required=False):
            block = QWidget()
            block.setObjectName("FormField")
            block_layout = QVBoxLayout(block)
            block_layout.setContentsMargins(0, 0, 0, 0)
            block_layout.setSpacing(4)
            label = QLabel(label_text + ("  *" if required else ""))
            label.setObjectName("FieldLabel")
            block_layout.addWidget(label)
            block_layout.addWidget(field)
            if not field.accessibleName():
                field.setAccessibleName(label_text)
            return block

        def section(title_text, fields, columns=4, height=86):
            section_widget = QWidget()
            section_widget.setObjectName("FormSection")
            section_widget.setFixedHeight(height)
            section_layout = QVBoxLayout(section_widget)
            section_layout.setContentsMargins(0, 0, 0, 0)
            section_layout.setSpacing(7)
            title = StrongBodyLabel(title_text)
            title.setObjectName("FormSectionTitle")
            section_layout.addWidget(title)
            grid = QGridLayout()
            grid.setHorizontalSpacing(12)
            grid.setVerticalSpacing(7)
            for column in range(columns):
                grid.setColumnStretch(column, 1)
            for row, column, span, label, field, required in fields:
                grid.addWidget(
                    field_block(label, field, required=required),
                    row, column, 1, span,
                )
            section_layout.addLayout(grid)
            return section_widget

        layout.addWidget(section("Student identity", (
            (0, 0, 1, "Last name", self.inp_last, True),
            (0, 1, 1, "First name", self.inp_first, True),
            (0, 2, 1, "Gender", self.inp_gender, False),
            (0, 3, 1, "Birthday", self.inp_bday, False),
        )))
        layout.addWidget(section("Education", (
            (0, 0, 1, "Grade / year", self.inp_grade, False),
            (0, 1, 1, "School", self.inp_school, False),
            (0, 2, 1, "Course / track", self.inp_course, False),
            (0, 3, 1, "Record status", self.inp_status, False),
        )))
        layout.addWidget(section("Support and location", (
            (0, 0, 1, "Area / coordinator", self.inp_area, False),
            (0, 1, 1, "Sponsor", self.inp_sponsor, False),
            (0, 2, 1, "City", self.inp_city, False),
            (0, 3, 1, "Address", self.inp_address, False),
        )))
        layout.addWidget(section("Contact and notes", (
            (0, 0, 1, "Contact number", self.inp_contact, False),
            (0, 1, 1, "Parent / guardian", self.inp_parents, False),
            (0, 2, 2, "Office remarks", self.inp_remarks, False),
        ), height=116))

        footer = QHBoxLayout()
        photo_note = QLabel("Photo can be added after saving")
        photo_note.setObjectName("FormHint")
        layout.addStretch()
        footer.addStretch()
        footer.addWidget(photo_note)

        self.save_form_btn = PrimaryPushButton("Save student")
        self.save_form_btn.clicked.connect(self.save_student_form)
        self.save_form_btn.hide()
        layout.addLayout(footer)

        outer.addWidget(content)
        outer.addStretch()
        self.stacked_widget.addWidget(widget)

    # ── SCREEN 4: EXPENSES ────────────────────────────────────────────────────
    def create_expenses_screen(self):
        widget = QWidget()
        widget.setObjectName("MainContent")
        outer_scroll = QScrollArea()
        outer_scroll.setWidgetResizable(True)
        inner = QWidget()
        inner.setObjectName("MainContent")
        layout = QVBoxLayout(inner)
        layout.setSpacing(20)
        layout.setContentsMargins(0, 0, 0, 0)

        self.expenses_title = StrongBodyLabel("Loading student…")
        self.expenses_title.setObjectName("ExpenseStudentName")

        # School year selector
        sy_card = QFrame()
        sy_card.setObjectName("ExpenseContextBar")
        sy_card.setFixedHeight(56)
        sy_layout = QHBoxLayout(sy_card)
        sy_layout.setContentsMargins(18, 8, 18, 8)
        sy_layout.setSpacing(8)
        sy_lbl = QLabel("School year")
        sy_lbl.setObjectName("FieldLabel")
        self.exp_school_year = QComboBox()
        import datetime
        sy_list = [f"{y}-{y+1}" for y in range(2020, 2032)]
        self.exp_school_year.addItems([ExpenseService.ALL_YEARS] + sy_list)
        cy = datetime.date.today().year
        cm = datetime.date.today().month
        cur_sy = f"{cy}-{cy+1}" if cm >= 6 else f"{cy-1}-{cy}"
        idx = self.exp_school_year.findText(cur_sy)
        # Show the complete synchronized history first. A current-year default
        # made valid imported expenses from prior years look missing.
        self.exp_school_year.setCurrentIndex(0)
        self.exp_school_year.currentTextChanged.connect(self._on_sy_changed)
        self.exp_school_year.setMinimumWidth(150)
        sy_layout.addWidget(self.expenses_title)
        sy_layout.addStretch()
        sy_layout.addWidget(sy_lbl)
        sy_layout.addWidget(self.exp_school_year)
        sync_label = QLabel("Synchronized")
        sync_label.setObjectName("BudgetStatus")
        sy_layout.addWidget(sync_label)
        layout.addWidget(sy_card)

        # ── Budget card ────────────────────────────────────────────────────
        budget_card = Card()
        budget_card.setProperty("tone", "success")
        budget_card.setFixedHeight(150)
        budget_main = QVBoxLayout(budget_card)
        budget_main.setContentsMargins(16, 16, 16, 16)
        budget_main.setSpacing(8)

        budget_hdr = QHBoxLayout()
        self.expense_budget_heading = StrongBodyLabel(
            f"Budget — {self._current_school_year()}"
        )
        self.expense_budget_heading.setObjectName("CardHeading")
        budget_hdr.addWidget(self.expense_budget_heading)
        budget_hdr.addStretch()

        # Edit budget row
        self.budget_input = QLineEdit()
        self.budget_input.setPlaceholderText("Enter budget e.g. 5000")
        self.budget_input.setMaximumWidth(200)
        self.budget_input.setMinimumHeight(40)
        self.budget_input.setAccessibleName("Budget amount in Philippine pesos")
        self.save_budget_btn = PrimaryPushButton("Save budget")
        set_content_hugging_button(self.save_budget_btn)
        self.save_budget_btn.clicked.connect(self.save_budget)
        budget_hdr.addWidget(self.budget_input)
        budget_hdr.addWidget(self.save_budget_btn)
        budget_main.addLayout(budget_hdr)

        budget_figures = QHBoxLayout()
        self.expense_budget_amount_display = QLabel("PHP 0.00")
        self.expense_budget_amount_display.setObjectName("DashboardHeroValue")
        self.expense_spent_display = StrongBodyLabel("PHP 0.00 spent")
        self.expense_remaining_display = QLabel("PHP 0.00 remaining")
        self.expense_remaining_display.setObjectName("Caption")
        budget_figures.addWidget(self.expense_budget_amount_display)
        budget_figures.addSpacing(28)
        budget_figures.addWidget(self.expense_spent_display)
        budget_figures.addStretch()
        budget_figures.addWidget(self.expense_remaining_display)
        budget_main.addLayout(budget_figures)

        # Progress bar row
        self.budget_bar = QProgressBar()
        self.budget_bar.setRange(0, 100)
        self.budget_bar.setValue(0)
        self.budget_bar.setTextVisible(False)
        self.budget_bar.setFixedHeight(9)
        self.budget_bar.setObjectName("BudgetProgress")
        self.budget_bar.setProperty("state", "success")
        self.budget_bar.setAccessibleName("Budget usage")
        self.budget_bar.setAccessibleDescription("No budget usage loaded")
        budget_main.addWidget(self.budget_bar)

        self.budget_status_lbl = QLabel("Budget unallocated for this school year.")
        self.budget_status_lbl.setObjectName("BudgetStatus")
        budget_main.addWidget(self.budget_status_lbl)
        layout.addWidget(budget_card)

        # Add expense card
        add_card = Card()
        self.expense_add_card = add_card
        add_card.setFixedHeight(118)
        add_layout = QVBoxLayout(add_card)
        add_layout.setContentsMargins(16, 16, 16, 16)
        add_layout.setSpacing(8)
        add_hdr = StrongBodyLabel("Add an expense")
        add_hdr.setObjectName("CardHeading")
        add_layout.addWidget(add_hdr)

        add_grid = QGridLayout()
        self.expense_add_grid = add_grid
        add_grid.setHorizontalSpacing(8)
        add_grid.setVerticalSpacing(6)
        self.exp_desc = QLineEdit(); self.exp_desc.setPlaceholderText("Description")
        self.exp_desc.setMinimumHeight(40)
        self.exp_amount = QLineEdit(); self.exp_amount.setPlaceholderText("0.00")
        self.exp_amount.setMinimumHeight(40)
        self.exp_date = QDateEdit()
        self.exp_date.setCalendarPopup(True)
        self.exp_date.setDisplayFormat("yyyy-MM-dd")
        self.exp_date.setDate(QDate.currentDate())
        self.exp_date.setMinimumHeight(40)
        self.exp_date.setMinimumWidth(190)
        self.exp_date.setCursor(Qt.CursorShape.PointingHandCursor)
        self.exp_date.setToolTip("Open the calendar to choose an expense date.")
        self.exp_sy_entry = QComboBox()
        self.exp_sy_entry.addItems(sy_list)
        if idx >= 1: self.exp_sy_entry.setCurrentIndex(idx - 1)
        self.exp_sy_entry.setMinimumHeight(40)
        self.add_expense_btn = PrimaryPushButton("Add expense")
        set_content_hugging_button(self.add_expense_btn)
        self.add_expense_btn.clicked.connect(self.add_expense)
        self.expense_description_label = QLabel("Description")
        self.expense_description_label.setObjectName("FieldLabel")
        self.expense_amount_label = QLabel("Amount (PHP)")
        self.expense_amount_label.setObjectName("FieldLabel")
        self.expense_date_label = QLabel("Expense date")
        self.expense_date_label.setObjectName("FieldLabel")
        self.expense_year_label = QLabel("School year")
        self.expense_year_label.setObjectName("FieldLabel")
        add_grid.addWidget(self.expense_description_label, 0, 0)
        add_grid.addWidget(self.expense_amount_label, 0, 1)
        add_grid.addWidget(self.expense_date_label, 0, 2)
        add_grid.addWidget(self.expense_year_label, 0, 3)
        add_grid.addWidget(self.exp_desc, 1, 0)
        add_grid.addWidget(self.exp_amount, 1, 1)
        add_grid.addWidget(self.exp_date, 1, 2)
        add_grid.addWidget(self.exp_sy_entry, 1, 3)
        add_grid.addWidget(self.add_expense_btn, 1, 4)
        add_grid.setColumnStretch(0, 3)
        add_grid.setColumnStretch(1, 1)
        add_grid.setColumnStretch(2, 1)
        add_grid.setColumnStretch(3, 1)
        add_layout.addLayout(add_grid)
        layout.addWidget(add_card)
        self._relayout_expense_entry()

        # Expenses table
        table_card = Card()
        table_card.setFixedHeight(252)
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(16, 16, 16, 16)
        table_layout.setSpacing(8)
        history_header = QHBoxLayout()
        expense_history_title = StrongBodyLabel("Expense history")
        expense_history_title.setObjectName("CardHeading")
        self.total_label = QLabel("Total: PHP 0.00")
        self.total_label.setObjectName("TotalLabel")
        self.total_label.setAccessibleName("Expense total")
        history_header.addWidget(expense_history_title)
        history_header.addStretch()
        history_header.addWidget(self.total_label)
        table_layout.addLayout(history_header)
        self.expenses_table = QTableWidget(0, 5)
        self.expenses_table.setObjectName("ExpensesTable")
        self.expenses_table.setHorizontalHeaderLabels(["Description", "Amount (PHP)", "Date", "School year", ""])
        self.expenses_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.expenses_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.expenses_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.expenses_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.expenses_table.horizontalHeader().setStretchLastSection(False)
        self.expenses_table.verticalHeader().setVisible(False)
        self.expenses_table.setAlternatingRowColors(True)
        self.expenses_table.setShowGrid(False)
        self.expenses_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.expenses_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.expenses_table.setMinimumHeight(176)
        self.expenses_table.verticalHeader().setMinimumSectionSize(40)
        self.expenses_table.verticalHeader().setDefaultSectionSize(40)
        self.expenses_empty_state = EmptyState(
            "No expenses recorded",
            "Add the first expense for this school year using the form above.",
        )
        self.expenses_empty_state.setAccessibleName("No expenses recorded")
        self.expense_history_stack = QStackedWidget()
        self.expense_history_stack.setMinimumHeight(176)
        self.expense_history_stack.addWidget(self.expenses_table)
        self.expense_history_stack.addWidget(self.expenses_empty_state)
        self.expense_history_stack.setCurrentWidget(self.expenses_empty_state)
        table_layout.addWidget(self.expense_history_stack)
        layout.addWidget(table_card)

        layout.addStretch()
        outer_scroll.setWidget(inner)
        w_layout = QVBoxLayout(widget)
        w_layout.setContentsMargins(0, 0, 0, 0)
        w_layout.addWidget(outer_scroll)
        self.stacked_widget.addWidget(widget)

    def create_coordinators_screen(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(20)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        controls_card = CardWidget()
        controls_card.setObjectName("CoordinatorToolbar")
        controls_card.setFixedHeight(62)
        controls_card.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Maximum,
        )
        controls_layout = QHBoxLayout(controls_card)
        controls_layout.setContentsMargins(12, 12, 12, 12)
        controls_layout.setSpacing(10)

        self.coord_status = QLabel("0 records")
        self.coord_status.setObjectName("Caption")
        self.coord_search = QLineEdit()
        self.coord_search.setPlaceholderText("Search coordinator name, area, or contact")
        self.coord_search.setAccessibleName("Search coordinators")
        self.coord_search.setClearButtonEnabled(True)
        self.coord_search.setFixedWidth(430)
        self.coord_search.setFixedHeight(38)
        self.coord_search.textChanged.connect(self._filter_coordinators)
        self.add_coordinator_btn = PrimaryPushButton("Add coordinator")
        set_content_hugging_button(self.add_coordinator_btn)
        self.add_coordinator_btn.clicked.connect(self._add_coordinator_dialog)
        self.add_coordinator_btn.hide()
        self.refresh_coordinator_btn = QPushButton("Refresh")
        self.refresh_coordinator_btn.setObjectName("SecondaryBtn")
        set_content_hugging_button(self.refresh_coordinator_btn)
        self.refresh_coordinator_btn.clicked.connect(self.load_coordinators)
        controls_layout.addWidget(self.coord_search)
        controls_layout.addWidget(self.coord_status)
        controls_layout.addStretch()
        controls_layout.addWidget(self.refresh_coordinator_btn)
        layout.addWidget(controls_card)

        self.coord_table_card = CardWidget()
        self.coord_table_card.setObjectName("CoordinatorTableCard")
        self.coord_table_card.setFixedHeight(554)
        self.coord_table_card.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Maximum,
        )
        table_layout = QVBoxLayout(self.coord_table_card)
        table_layout.setContentsMargins(12, 12, 12, 12)
        table_layout.setSpacing(0)
        self.coord_table = QTableWidget()
        self.coord_table.setObjectName("CoordTable")
        self.coord_table.setColumnCount(6)
        self.coord_table.setHorizontalHeaderLabels(["NAME", "AREA", "CONTACT", "ASSIGNED", "UPDATED", "ACTIONS"])
        self.coord_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.coord_table.horizontalHeader().setStretchLastSection(False)
        self.coord_table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )
        self.coord_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.coord_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.coord_table.setAlternatingRowColors(True)
        self.coord_table.setShowGrid(False)
        self.coord_table.verticalHeader().setVisible(False)
        self.coord_table.setWordWrap(True)
        self.coord_table.verticalHeader().setMinimumSectionSize(72)
        self.coord_table.verticalHeader().setDefaultSectionSize(72)
        self.coord_table.doubleClicked.connect(self._edit_coordinator_dialog)
        empty_action = ActionButton("Add coordinator")
        empty_action.clicked.connect(self._run_coord_empty_action)
        self.coord_empty_action = empty_action
        self.coord_empty_state = EmptyState(
            "No coordinators yet",
            "Add the first ministry contact to begin the shared directory.",
            empty_action,
        )
        self.coord_results_stack = QStackedWidget()
        self.coord_results_stack.addWidget(self.coord_table)
        self.coord_results_stack.addWidget(self.coord_empty_state)
        self.coord_results_stack.setCurrentWidget(self.coord_empty_state)
        table_layout.addWidget(self.coord_results_stack, 1)
        layout.addWidget(self.coord_table_card)
        layout.addStretch(1)

        self._coord_all_rows = []  # cache for filtering
        self.stacked_widget.addWidget(widget)

    def create_settings_screen(self):
        self.settings_view = SettingsView(self._settings())
        self.stacked_widget.addWidget(self.settings_view)

    def load_coordinators(self):
        self._coordinator_request += 1
        request_id = self._coordinator_request
        self.coord_status.setText("Loading…")
        self.refresh_coordinator_btn.setEnabled(False)
        self.refresh_coordinator_btn.setText("Refreshing…")
        if not self._coord_all_rows:
            self.coord_empty_state.title_label.setText("Loading directory…")
            self.coord_empty_state.description_label.setText(
                "Retrieving the shared coordinator records."
            )
            self.coord_empty_action.setVisible(False)
            self.coord_results_stack.setCurrentWidget(self.coord_empty_state)

        def loaded(data):
            if request_id != self._coordinator_request:
                return
            self._coord_all_rows = data
            self._populate_coord_table(data)
            self._mark_database_updated()
            self._finish_coordinator_refresh(request_id)

        def failed(error):
            if request_id != self._coordinator_request:
                return
            self.coord_status.setText("Unavailable")
            self.coord_empty_state.title_label.setText("Coordinator directory unavailable")
            self.coord_empty_state.description_label.setText(
                "Check the office database connection, then try again."
            )
            self.coord_empty_action.setText("Try again")
            self.coord_empty_action.setProperty("mode", "retry")
            self.coord_results_stack.setCurrentWidget(self.coord_empty_state)
            self._set_coordinator_results_expanded(False)
            self.coord_empty_action.setVisible(True)
            self._finish_coordinator_refresh(request_id)
            logging.getLogger(__name__).error(
                "Coordinator load failed:\n%s", error
            )

        task = self._run_background(
            self.coordinator_repository.list_coordinators,
            loaded,
            failed,
        )
        return task

    def _finish_coordinator_refresh(self, request_id) -> None:
        if request_id != self._coordinator_request:
            return
        self.refresh_coordinator_btn.setEnabled(True)
        self.refresh_coordinator_btn.setText("Refresh")

    def _set_coordinator_results_expanded(self, has_rows):
        self.coord_table_card.setFixedHeight(554)

    def _populate_coord_table(self, rows):
        self.coord_empty_action.setVisible(True)
        self.coord_table.setRowCount(0)
        for r in rows:
            row_idx = self.coord_table.rowCount()
            self.coord_table.insertRow(row_idx)
            contact = "  /  ".join(
                value for value in (r.get("email"), r.get("contact_no")) if value
            ) or "—"
            updated = str(r.get("updated_at") or r.get("created_at") or "—")[:10]
            values = (
                r.get("contact_person") or "Unnamed coordinator",
                r.get("location") or "Area not set",
                contact,
                r.get("assigned") or r.get("fb_page") or "—",
                updated,
            )
            for col, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setToolTip(val)
                item.setData(Qt.ItemDataRole.UserRole, r.get("id"))
                self.coord_table.setItem(row_idx, col, item)
            edit_button = ActionButton("Edit", variant="tertiary")
            set_content_hugging_button(edit_button, height=32)
            edit_button.clicked.connect(
                lambda _checked=False, row=row_idx: self._edit_coordinator_dialog(
                    self.coord_table.model().index(row, 0)
                )
            )
            self.coord_table.setCellWidget(row_idx, 5, edit_button)
            self.coord_table.setRowHeight(row_idx, 72)
        count = len(rows)
        self.coord_status.setText(f"{count} {'record' if count == 1 else 'records'}")
        if rows:
            self.coord_results_stack.setCurrentWidget(self.coord_table)
            self._set_coordinator_results_expanded(True)
            return
        if self.coord_search.text().strip():
            self.coord_empty_state.title_label.setText("No matching coordinators")
            self.coord_empty_state.description_label.setText(
                "Try a different name, location, email address, or phone number."
            )
            self.coord_empty_action.setText("Clear search")
            self.coord_empty_action.setProperty("mode", "clear")
        else:
            self.coord_empty_state.title_label.setText("No coordinators yet")
            self.coord_empty_state.description_label.setText(
                "Add the first ministry contact to begin the shared directory."
            )
            self.coord_empty_action.setText("Add coordinator")
            self.coord_empty_action.setProperty("mode", "add")
        self.coord_results_stack.setCurrentWidget(self.coord_empty_state)
        self._set_coordinator_results_expanded(False)

    def _run_coord_empty_action(self):
        mode = self.coord_empty_action.property("mode")
        if mode == "clear":
            self.coord_search.clear()
        elif mode == "retry":
            self.load_coordinators()
        else:
            self._add_coordinator_dialog()

    def _filter_coordinators(self, text):
        q = text.strip().lower()
        if not q:
            self._populate_coord_table(self._coord_all_rows)
            return
        filtered = [
            r for r in self._coord_all_rows
            if any(q in str(r.get(k) or "").lower()
                   for k in ["location", "contact_person", "email", "contact_no", "fb_page", "remarks"])
        ]
        self._populate_coord_table(filtered)

    def _coord_row_data(self, row_idx):
        item = self.coord_table.item(row_idx, 0)
        record_id = item.data(Qt.ItemDataRole.UserRole) if item else None
        record = next(
            (dict(row) for row in self._coord_all_rows if row.get("id") == record_id),
            {},
        )
        record["_id"] = record_id
        return record

    def _coord_dialog(self, title, prefill=None):
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.setObjectName("CoordinatorEditorDialog")
        dlg.setMinimumWidth(480)
        form = QFormLayout(dlg)
        form.setSpacing(10)
        form.setContentsMargins(22, 20, 22, 22)
        fields = {}
        for label, key in [("Location", "location"), ("Contact person", "contact_person"),
                            ("Email", "email"), ("Contact no.", "contact_no"),
                            ("Facebook page", "fb_page"), ("Remarks", "remarks")]:
            w = QLineEdit()
            w.setMinimumHeight(40)
            w.setAccessibleName(label)
            if prefill:
                w.setText(prefill.get(key, ""))
            form.addRow(label, w)
            fields[key] = w
        btn_row = QHBoxLayout()
        save_btn = PrimaryPushButton("Save")
        set_content_hugging_button(save_btn)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("SecondaryBtn")
        set_content_hugging_button(cancel_btn)
        save_btn.clicked.connect(dlg.accept)
        cancel_btn.clicked.connect(dlg.reject)
        btn_row.addStretch(); btn_row.addWidget(cancel_btn); btn_row.addWidget(save_btn)
        form.addRow(btn_row)
        return dlg, fields

    def _add_coordinator_dialog(self):
        dlg, fields = self._coord_dialog("Add coordinator")
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        record = {k: v.text().strip() for k, v in fields.items()}
        if not record["location"] and not record["contact_person"]:
            self.status_bar.showMessage(
                "Enter a location or contact person before saving.", 5000
            )
            return
        self._run_coordinator_mutation(
            lambda: self.coordinator_repository.insert_coordinator(record),
            success_message="Coordinator added",
            audit_action="create",
            audit_details={"location": record["location"]},
        )

    def _edit_coordinator_dialog(self, index):
        row = index.row()
        prefill = self._coord_row_data(row)
        rec_id = prefill.pop("_id", None)
        dlg, fields = self._coord_dialog("Edit coordinator", prefill)
        # Add delete button
        del_btn = QPushButton("Delete")
        del_btn.setObjectName("DangerBtn")
        set_content_hugging_button(del_btn)
        del_btn.clicked.connect(lambda: dlg.done(2))
        form = dlg.layout()
        last_row = form.rowCount() - 1
        btn_row_item = form.itemAt(last_row, QFormLayout.ItemRole.FieldRole)
        if btn_row_item and btn_row_item.layout():
            btn_row_item.layout().insertWidget(0, del_btn)
        result = dlg.exec()
        if result == QDialog.DialogCode.Accepted and rec_id:
            record = {k: v.text().strip() for k, v in fields.items()}
            self._run_coordinator_mutation(
                lambda: self.coordinator_repository.update_coordinator(rec_id, record),
                success_message="Coordinator updated",
                audit_action="update",
                entity_id=rec_id,
                audit_details={"location": record["location"]},
            )
        elif result == 2 and rec_id:
            confirm = QMessageBox.question(
                self,
                "Delete coordinator",
                "Delete this coordinator from the shared directory?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm == QMessageBox.StandardButton.Yes:
                self._run_coordinator_mutation(
                    lambda: self.coordinator_repository.delete_coordinator(rec_id),
                    success_message="Coordinator deleted",
                    audit_action="delete",
                    entity_id=rec_id,
                    audit_details={"location": prefill.get("location", "")},
                )

    def _run_coordinator_mutation(
        self,
        function,
        *,
        success_message,
        audit_action,
        entity_id=None,
        audit_details=None,
    ):
        """Run coordinator writes without freezing navigation or input."""
        self.add_coordinator_btn.setEnabled(False)
        self.refresh_coordinator_btn.setEnabled(False)
        self.coord_table.setEnabled(False)
        self.coord_status.setText("Saving…")

        def succeeded(result):
            saved_id = entity_id
            if saved_id is None and result:
                saved_id = result[0].get("id")
            self.add_coordinator_btn.setEnabled(True)
            self.coord_table.setEnabled(True)
            self._audit(
                audit_action,
                "coordinator",
                saved_id,
                audit_details,
            )
            self.load_coordinators()
            self.status_bar.showMessage(success_message, 4000)

        def failed(error):
            self.add_coordinator_btn.setEnabled(True)
            self.refresh_coordinator_btn.setEnabled(True)
            self.coord_table.setEnabled(True)
            self.coord_status.setText("Save failed")
            self.status_bar.showMessage(
                "Could not save the coordinator. Check the office connection and try again.",
                8000,
            )
            logging.getLogger(__name__).error(
                "Coordinator mutation failed:\n%s", error
            )

        return self._run_background(function, succeeded, failed)

    def create_workbook_screen(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        layout.setSpacing(20)

        self.workbook_state_badge = StatusBadge(
            "No workbook", role="WorkbookStateBadge"
        )

        def workbook_btn(text, slot, object_name=None, icon=None):
            variant = {
                "SecondaryBtn": "secondary",
                "DangerBtn": "danger",
                "SuccessBtn": "success",
            }.get(object_name, "primary")
            button = ActionButton(text, variant=variant)
            if object_name:
                button.setObjectName(object_name)
            button.setProperty("density", "compact")
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            button.setFixedHeight(40)
            button.setAccessibleName(text)
            button.clicked.connect(slot)
            return button

        controls = Card()
        controls.setObjectName("WorkbookToolbar")
        controls.setFixedHeight(86)
        controls.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Maximum,
        )
        controls_layout = QHBoxLayout(controls)
        controls_layout.setContentsMargins(16, 12, 16, 12)
        controls_layout.setSpacing(12)

        file_row = QHBoxLayout()
        file_row.setSpacing(12)
        file_icon = QLabel("XLSX")
        file_icon.setObjectName("WorkbookFileIcon")
        file_icon.setFixedSize(46, 46)
        file_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        file_row.addWidget(file_icon)
        file_meta = QVBoxLayout()
        file_meta.setSpacing(2)

        self.workbook_path_label = QLabel("No workbook loaded")
        self.workbook_path_label.setObjectName("WorkbookFileLabel")
        self.workbook_path_label.setWordWrap(True)
        self.workbook_status_label = QLabel("No workbook selected.")
        self.workbook_status_label.setObjectName("WorkbookPathLabel")
        self.workbook_status_label.setWordWrap(True)
        file_meta.addWidget(self.workbook_path_label)
        file_meta.addWidget(self.workbook_status_label)
        file_row.addLayout(file_meta, 1)
        file_row.addWidget(self.workbook_state_badge, alignment=Qt.AlignmentFlag.AlignVCenter)

        self.workbook_open_saved_btn = workbook_btn(
            "Reload file",
            self.load_saved_workbook,
            icon=QStyle.StandardPixmap.SP_DialogOpenButton,
        )
        self.workbook_choose_btn = workbook_btn(
            "Choose file",
            self.choose_workbook_file,
            "SecondaryBtn",
            QStyle.StandardPixmap.SP_DirIcon,
        )
        self.workbook_save_btn = workbook_btn(
            "Save workbook",
            self.save_workbook_tabs,
            "SuccessBtn",
            QStyle.StandardPixmap.SP_DialogSaveButton,
        )
        self.workbook_reload_btn = workbook_btn(
            "Reload",
            lambda: self.load_workbook_tabs(self._workbook_path) if self._workbook_path else self.choose_workbook_file(),
            "SecondaryBtn",
            QStyle.StandardPixmap.SP_BrowserReload,
        )
        self.workbook_excel_btn = workbook_btn(
            "Excel",
            self.open_workbook_in_excel,
            "SecondaryBtn",
            QStyle.StandardPixmap.SP_FileIcon,
        )

        self.workbook_save_btn.hide()
        self.workbook_reload_btn.hide()
        self.workbook_excel_btn.hide()
        controls_layout.addLayout(file_row, 1)
        controls_layout.addWidget(self.workbook_open_saved_btn)
        controls_layout.addWidget(self.workbook_choose_btn)

        self.workbook_sheet_controls = Card()
        self.workbook_sheet_controls.setObjectName("WorkbookSheetBar")
        self.workbook_sheet_controls.setFixedHeight(52)
        sheet_row = QHBoxLayout(self.workbook_sheet_controls)
        sheet_row.setContentsMargins(14, 7, 14, 7)
        sheet_row.setSpacing(Spacing.XS)
        sheet_lbl = QLabel("Sheet")
        sheet_lbl.setObjectName("ToolbarLabel")
        self.workbook_sheet_combo = QComboBox()
        self.workbook_sheet_combo.setObjectName("WorkbookSheetCombo")
        self.workbook_sheet_combo.setMinimumWidth(230)
        self.workbook_sheet_combo.currentIndexChanged.connect(self._on_workbook_sheet_combo_changed)
        self.workbook_sheet_summary_label = QLabel("No sheet selected")
        self.workbook_sheet_summary_label.setObjectName("WorkbookSheetSummary")
        sheet_row.addWidget(sheet_lbl)
        sheet_row.addWidget(self.workbook_sheet_combo)
        sheet_row.addWidget(self.workbook_sheet_summary_label, 1)
        layout.addWidget(controls)
        layout.addWidget(self.workbook_sheet_controls)

        self.workbook_add_column_btn = workbook_btn("Insert column", self.insert_workbook_column, "SecondaryBtn")
        self.workbook_delete_column_btn = workbook_btn("Delete column", self.delete_workbook_column, "DangerBtn")
        self.workbook_sync_current_btn = workbook_btn("Sync current", self.sync_current_sheet_to_supabase)
        self.workbook_sync_all_btn = workbook_btn("Sync all", self.sync_all_workbook_sheets_to_supabase, "SecondaryBtn")
        self.workbook_export_btn = workbook_btn(
            "Export students",
            self.student_list_view.export_all_students_to_excel,
            "SecondaryBtn",
        )

        self.workbook_data_actions = QWidget()
        action_row = QHBoxLayout(self.workbook_data_actions)
        action_row.setContentsMargins(0, 0, 0, 0)
        action_row.setSpacing(Spacing.XS)
        action_row.addWidget(self.workbook_add_column_btn)
        action_row.addWidget(self.workbook_delete_column_btn)
        action_row.addStretch()
        action_row.addWidget(self.workbook_sync_current_btn)
        action_row.addWidget(self.workbook_sync_all_btn)
        action_row.addWidget(self.workbook_export_btn)
        self.workbook_data_actions.hide()

        self.workbook_empty_card = EmptyState(
            "No workbook open",
            "Use Reload file or Choose file above. After loading, select a sheet and review its rows.",
        )
        self.workbook_empty_card.setFixedHeight(458)
        self.workbook_empty_card.setAccessibleName("No workbook open")
        layout.addWidget(self.workbook_empty_card)

        self.workbook_tabs = QTabWidget()
        self.workbook_tabs.setDocumentMode(True)
        self.workbook_tabs.setUsesScrollButtons(True)
        self.workbook_tabs.setElideMode(Qt.TextElideMode.ElideRight)
        self.workbook_tabs.setMovable(False)
        self.workbook_tabs.setFixedHeight(458)
        self.workbook_tabs.currentChanged.connect(self._on_workbook_tab_changed)
        self.workbook_tabs.setVisible(False)
        layout.addWidget(self.workbook_tabs)
        layout.addStretch()

        self._refresh_workbook_controls()
        self.stacked_widget.addWidget(widget)
    def _settings(self):
        return QSettings("YWAMBalut", "SSMStudentProfilingSystem")

    def _refresh_workbook_controls(self):
        has_workbook = bool(self._workbook)
        for button_name in (
            "workbook_save_btn",
            "workbook_reload_btn",
            "workbook_excel_btn",
            "workbook_add_column_btn",
            "workbook_delete_column_btn",
            "workbook_sync_current_btn",
            "workbook_sync_all_btn",
        ):
            button = getattr(self, button_name, None)
            if button is not None:
                button.setEnabled(has_workbook and not self._workbook_busy)
        if hasattr(self, "workbook_save_btn"):
            self.workbook_save_btn.setEnabled(
                has_workbook and self._workbook_dirty and not self._workbook_busy
            )
            self.workbook_save_btn.hide()
            self.workbook_reload_btn.hide()
            self.workbook_excel_btn.hide()
        if hasattr(self, "workbook_sheet_controls"):
            self.workbook_sheet_controls.setVisible(has_workbook)
        if hasattr(self, "workbook_data_actions"):
            self.workbook_data_actions.setVisible(has_workbook)
        if hasattr(self, "workbook_sheet_combo"):
            self.workbook_sheet_combo.setEnabled(
                has_workbook and self.workbook_tabs.count() > 0 and not self._workbook_busy
            )
        if hasattr(self, "workbook_tabs"):
            self.workbook_tabs.setEnabled(not self._workbook_busy)
        self._refresh_workbook_state_badge()

    def _set_workbook_busy(self, busy, message=""):
        self._workbook_busy = busy
        self._refresh_workbook_controls()
        if message:
            self.status_bar.showMessage(message, 30000 if busy else 4000)

    def _refresh_workbook_state_badge(self):
        if not hasattr(self, "workbook_state_badge"):
            return
        if not self._workbook:
            text, state = "No workbook", "empty"
        elif self._workbook_dirty:
            text, state = "Unsaved", "dirty"
        else:
            text, state = "Saved", "saved"
        self.workbook_state_badge.set_state(state, text)

    def _on_workbook_sheet_combo_changed(self, index):
        if not hasattr(self, "workbook_tabs") or index < 0:
            return
        if self.workbook_tabs.currentIndex() != index:
            self.workbook_tabs.setCurrentIndex(index)

    def _update_workbook_sheet_summary(self, index=None):
        if not hasattr(self, "workbook_sheet_summary_label"):
            return
        if not self._workbook or not hasattr(self, "workbook_tabs") or self.workbook_tabs.count() == 0:
            self.workbook_sheet_summary_label.setText("No sheet selected")
            return
        if index is None:
            index = self.workbook_tabs.currentIndex()
        if index < 0:
            self.workbook_sheet_summary_label.setText("No sheet selected")
            return
        table = self.workbook_tabs.widget(index)
        sheet_name = self.workbook_tabs.tabText(index)
        if isinstance(table, QTableWidget) and sheet_name in self._loaded_workbook_sheets:
            self.workbook_sheet_summary_label.setText(f"{sheet_name}: {table.rowCount()} rows x {table.columnCount()} columns")
        else:
            self.workbook_sheet_summary_label.setText(f"{sheet_name}: not loaded")

    def load_saved_workbook(self):
        saved_path = self._settings().value("workbook_path", "", type=str)
        if saved_path and os.path.exists(saved_path):
            self.load_workbook_tabs(saved_path)
            return

        if saved_path:
            QMessageBox.information(
                self,
                "Saved workbook missing",
                f"The saved workbook path was not found:\n{saved_path}\n\nChoose the workbook location for this computer."
            )
        self.choose_workbook_file()

    def choose_workbook_file(self):
        saved_path = self._settings().value("workbook_path", "", type=str)
        start_dir = os.path.dirname(self._workbook_path or saved_path) if (self._workbook_path or saved_path) else os.path.expanduser("~")
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open SSM Workbook",
            start_dir,
            "Excel files (*.xlsx *.xlsm *.xltx *.xltm)"
        )
        if path:
            self.load_workbook_tabs(path)

    def open_workbook_in_excel(self):
        path = self._workbook_path
        if not path:
            saved_path = self._settings().value("workbook_path", "", type=str)
            start_dir = os.path.dirname(saved_path) if saved_path else os.path.expanduser("~")
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Open workbook in Excel",
                start_dir,
                "Excel files (*.xlsx *.xlsm *.xltx *.xltm)"
            )
            if not path:
                return

        if self._workbook_dirty:
            confirm = QMessageBox.question(
                self,
                "Unsaved workbook changes",
                "Save workbook changes before opening it in Excel?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )
            if confirm == QMessageBox.StandardButton.Cancel:
                return
            if confirm == QMessageBox.StandardButton.Yes:
                self.save_workbook_tabs()

        try:
            os.startfile(path)
            self.status_bar.showMessage("Opened workbook in Excel", 4000)
        except Exception as e:
            QMessageBox.critical(self, "Open in Excel failed", f"Could not open workbook:\n({type(e).__name__}) {e}")

    def load_workbook_tabs(self, path):
        if not path:
            return False
        if self._workbook_dirty and os.path.abspath(path) != os.path.abspath(self._workbook_path or ""):
            decision = QMessageBox.question(
                self,
                "Unsaved workbook changes",
                "Save the current workbook before opening another file?",
                QMessageBox.StandardButton.Yes |
                QMessageBox.StandardButton.No |
                QMessageBox.StandardButton.Cancel,
            )
            if decision == QMessageBox.StandardButton.Cancel:
                return False
            if decision == QMessageBox.StandardButton.Yes and not self.save_workbook_tabs():
                return False
        try:
            from openpyxl import load_workbook

            keep_vba = os.path.splitext(path)[1].lower() in (".xlsm", ".xltm")
            new_workbook = load_workbook(
                path,
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
            )
            with self._workbook_lock:
                if self._workbook and hasattr(self._workbook, "close"):
                    self._workbook.close()
                self._workbook = new_workbook
            self._workbook_path = path
            self._workbook_mtime_ns = os.stat(path).st_mtime_ns
            self._settings().setValue("workbook_path", path)
            self._loaded_workbook_sheets.clear()
            self._workbook_dirty = False
            self._workbook_revision += 1
            self._invalidate_master_reference_cache()
            self.workbook_tabs.blockSignals(True)
            self.workbook_tabs.clear()
            self.workbook_sheet_combo.blockSignals(True)
            self.workbook_sheet_combo.clear()

            for sheet_name in self._workbook.sheetnames:
                table = QTableWidget()
                table.setObjectName("WorkbookTable")
                table.setAlternatingRowColors(True)
                table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
                table.setEditTriggers(
                    QTableWidget.EditTrigger.DoubleClicked |
                    QTableWidget.EditTrigger.EditKeyPressed |
                    QTableWidget.EditTrigger.AnyKeyPressed
                )
                table.horizontalHeader().setDefaultSectionSize(150)
                table.horizontalHeader().setMinimumSectionSize(96)
                table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                table.horizontalHeader().setStretchLastSection(False)
                table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
                table.setHorizontalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
                table.verticalHeader().setDefaultSectionSize(42)
                table.verticalHeader().setMinimumSectionSize(42)
                table.setWordWrap(False)
                table.cellChanged.connect(self._on_workbook_cell_changed)
                self.workbook_tabs.addTab(table, sheet_name)

            self.workbook_sheet_combo.addItems(self._workbook.sheetnames)
            if self._workbook.sheetnames:
                self.workbook_sheet_combo.setCurrentIndex(0)
            self.workbook_sheet_combo.blockSignals(False)
            self.workbook_tabs.blockSignals(False)
            self.workbook_path_label.setText(os.path.basename(path))
            self.workbook_path_label.setToolTip(path)
            self.workbook_status_label.setText(path)
            self.workbook_empty_card.setVisible(False)
            self.workbook_tabs.setVisible(True)
            self._refresh_workbook_controls()
            if self.workbook_tabs.count():
                self.workbook_tabs.setCurrentIndex(0)
                self._load_workbook_sheet(0)
            return True
        except Exception as e:
            QMessageBox.critical(self, "Workbook error", f"Could not open workbook:\n({type(e).__name__}) {e}")
            return False

    def _on_workbook_tab_changed(self, index):
        if index < 0:
            self._update_workbook_sheet_summary(index)
            return
        if hasattr(self, "workbook_sheet_combo") and self.workbook_sheet_combo.currentIndex() != index:
            self.workbook_sheet_combo.blockSignals(True)
            self.workbook_sheet_combo.setCurrentIndex(index)
            self.workbook_sheet_combo.blockSignals(False)
        self._load_workbook_sheet(index)

    def _current_workbook_table(self):
        if not hasattr(self, "workbook_tabs"):
            return None
        index = self.workbook_tabs.currentIndex()
        if index < 0:
            return None
        table = self.workbook_tabs.widget(index)
        return table if isinstance(table, QTableWidget) else None

    def _sync_workbook_side_scroll(self, table=None):
        pass  # external scroll bar removed; native table scrollbar is used

    def _on_workbook_h_scroll_changed(self, value):
        pass  # external scroll bar removed

    def _set_workbook_column_widths(self, table):
        for col in range(table.columnCount()):
            table.setColumnWidth(col, 150)

    def _load_workbook_sheet(self, index):
        if not self._workbook or index < 0 or self._workbook_busy:
            return
        sheet_name = self.workbook_tabs.tabText(index)
        if sheet_name in self._loaded_workbook_sheets:
            self._update_workbook_sheet_summary(index)
            return

        table = self.workbook_tabs.widget(index)
        self._workbook_loading_request += 1
        request_id = self._workbook_loading_request
        self._loading_workbook_sheet = True
        table.blockSignals(True)
        table.setEnabled(False)
        table.clear()
        table.setRowCount(0)
        table.setColumnCount(0)
        self.workbook_sheet_summary_label.setText(f"Loading {sheet_name}...")
        self._set_workbook_busy(True, f"Loading {sheet_name}...")

        def extract_rows():
            with self._workbook_lock:
                worksheet = self._workbook[sheet_name]
                max_cols = worksheet.max_column or 0
                rows = [
                    [self._format_excel_value(value) for value in row]
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=worksheet.max_row,
                        max_col=max_cols,
                        values_only=False,
                    )
                ]
            return rows, max_cols

        def prepare(result):
            if request_id != self._workbook_loading_request:
                return
            rows, max_cols = result
            table.setRowCount(len(rows))
            table.setColumnCount(max_cols)
            if rows:
                headers = [
                    value if value else self._excel_column_label(column + 1)
                    for column, value in enumerate(rows[0])
                ]
            else:
                headers = [self._excel_column_label(column + 1) for column in range(max_cols)]
            table.setHorizontalHeaderLabels(headers)
            self._set_workbook_column_widths(table)

            def populate(start=0):
                if request_id != self._workbook_loading_request:
                    return
                end = min(start + 100, len(rows))
                for row_index in range(start, end):
                    for column_index, value in enumerate(rows[row_index]):
                        table.setItem(row_index, column_index, QTableWidgetItem(value))
                if end < len(rows):
                    self.workbook_sheet_summary_label.setText(
                        f"Loading {sheet_name}... {end}/{len(rows)} rows"
                    )
                    QTimer.singleShot(0, lambda: populate(end))
                    return
                table.blockSignals(False)
                table.setEnabled(True)
                self._loading_workbook_sheet = False
                self._loaded_workbook_sheets.add(sheet_name)
                self._set_workbook_busy(False)
                self._update_workbook_sheet_summary(index)
                if not self._workbook_dirty:
                    self.workbook_status_label.setText(self._workbook_path or "")

            populate()

        def failed(error):
            if request_id != self._workbook_loading_request:
                return
            table.blockSignals(False)
            table.setEnabled(True)
            self._loading_workbook_sheet = False
            self._set_workbook_busy(False)
            self.workbook_sheet_summary_label.setText(f"Could not load {sheet_name}")
            QMessageBox.critical(
                self,
                "Workbook sheet error",
                error.strip().splitlines()[-1],
            )

        return self._run_background(extract_rows, prepare, failed)

    def _on_workbook_cell_changed(self, row, column):
        if self._loading_workbook_sheet or not self._workbook:
            return
        table = self.sender()
        index = self.workbook_tabs.indexOf(table)
        if index < 0:
            return

        sheet_name = self.workbook_tabs.tabText(index)
        item = table.item(row, column)
        text = item.text() if item else ""
        with self._workbook_lock:
            self._workbook[sheet_name].cell(row=row + 1, column=column + 1).value = text if text != "" else None
        self._workbook_dirty = True
        self._workbook_revision += 1
        self._invalidate_master_reference_cache()
        self.workbook_status_label.setText(f"Unsaved changes in {sheet_name}.")
        self._refresh_workbook_controls()
        self._update_workbook_sheet_summary(index)

    def insert_workbook_column(self):
        if not self._workbook:
            QMessageBox.information(self, "Workbook", "Open a workbook first.")
            return

        index = self.workbook_tabs.currentIndex()
        if index < 0:
            return

        sheet_name = self.workbook_tabs.tabText(index)
        table = self.workbook_tabs.widget(index)
        ws = self._workbook[sheet_name]

        if sheet_name not in self._loaded_workbook_sheets:
            self._load_workbook_sheet(index)

        selected_col = table.currentColumn()
        insert_col = selected_col if selected_col >= 0 else table.columnCount()
        excel_col = insert_col + 1
        with self._workbook_lock:
            ws.insert_cols(excel_col)

        table.blockSignals(True)
        table.insertColumn(insert_col)
        table.setHorizontalHeaderItem(insert_col, QTableWidgetItem(self._excel_column_label(excel_col)))
        table.setColumnWidth(insert_col, 150)

        if table.rowCount() == 0:
            table.setRowCount(1)

        for row in range(table.rowCount()):
            table.setItem(row, insert_col, QTableWidgetItem(""))

        table.item(0, insert_col).setText("New column")
        with self._workbook_lock:
            ws.cell(row=1, column=excel_col).value = "New column"
        self._refresh_workbook_table_headers(table)
        table.blockSignals(False)

        self._workbook_dirty = True
        self._workbook_revision += 1
        self._invalidate_master_reference_cache()
        self._refresh_workbook_controls()
        self._update_workbook_sheet_summary(index)
        table.setCurrentCell(0, insert_col)
        table.editItem(table.item(0, insert_col))
        self.workbook_status_label.setText(f"Inserted column in {sheet_name}.")

    def delete_workbook_column(self):
        if not self._workbook:
            QMessageBox.information(self, "Workbook", "Open a workbook first.")
            return

        index = self.workbook_tabs.currentIndex()
        if index < 0:
            return

        sheet_name = self.workbook_tabs.tabText(index)
        table = self.workbook_tabs.widget(index)
        ws = self._workbook[sheet_name]

        if sheet_name not in self._loaded_workbook_sheets:
            self._load_workbook_sheet(index)

        selected_col = table.currentColumn()
        if selected_col < 0:
            QMessageBox.information(self, "Delete column", "Select a column or cell first.")
            return

        col_label = self._excel_column_label(selected_col + 1)
        confirm = QMessageBox.question(
            self,
            "Delete column",
            f"Delete column {col_label} from '{sheet_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        with self._workbook_lock:
            ws.delete_cols(selected_col + 1)
        table.blockSignals(True)
        table.removeColumn(selected_col)
        self._refresh_workbook_table_headers(table)
        table.blockSignals(False)

        self._workbook_dirty = True
        self._workbook_revision += 1
        self._invalidate_master_reference_cache()
        self._refresh_workbook_controls()
        self._update_workbook_sheet_summary(index)
        self.workbook_status_label.setText(f"Deleted column {col_label} from {sheet_name}.")

    def _refresh_workbook_table_headers(self, table):
        for col in range(table.columnCount()):
            table.setHorizontalHeaderItem(col, QTableWidgetItem(self._excel_column_label(col + 1)))

    def save_workbook_tabs(self):
        if not self._workbook or not self._workbook_path:
            QMessageBox.information(self, "Workbook", "No workbook is open.")
            return False
        try:
            if os.path.exists(self._workbook_path):
                current_mtime = os.stat(self._workbook_path).st_mtime_ns
                if self._workbook_mtime_ns is not None and current_mtime != self._workbook_mtime_ns:
                    QMessageBox.warning(
                        self,
                        "Workbook changed on another computer",
                        "The workbook changed after you opened it. Your copy was not saved.\n\n"
                        "Reload the workbook and re-apply your changes, or save a separate copy in Excel.",
                    )
                    return False

            self._backup_workbook_file()
            directory = os.path.dirname(os.path.abspath(self._workbook_path))
            suffix = os.path.splitext(self._workbook_path)[1] or ".xlsx"
            handle, temporary_path = tempfile.mkstemp(
                prefix=".ssm-save-", suffix=suffix, dir=directory
            )
            os.close(handle)
            try:
                with self._workbook_lock:
                    self._workbook.save(temporary_path)
                os.replace(temporary_path, self._workbook_path)
            finally:
                if os.path.exists(temporary_path):
                    os.remove(temporary_path)

            self._workbook_mtime_ns = os.stat(self._workbook_path).st_mtime_ns
            self._workbook_dirty = False
            self._workbook_revision += 1
            self._invalidate_master_reference_cache()
            self.workbook_status_label.setText(self._workbook_path)
            self._refresh_workbook_controls()
            self.status_bar.showMessage("Workbook saved", 4000)
            return True
        except PermissionError:
            QMessageBox.warning(
                self,
                "Workbook is open",
                "Could not save the workbook. Close it in Excel, then click Save workbook again."
            )
            return False
        except Exception as e:
            QMessageBox.critical(self, "Save error", f"Could not save workbook:\n({type(e).__name__}) {e}")
            return False

    def _backup_workbook_file(self):
        if not self._workbook_path or not os.path.exists(self._workbook_path):
            return
        backup_dir = os.path.join(os.path.dirname(self._workbook_path), "SSM Backups")
        os.makedirs(backup_dir, exist_ok=True)
        stem, extension = os.path.splitext(os.path.basename(self._workbook_path))
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copy2(
            self._workbook_path,
            os.path.join(backup_dir, f"{stem}-{timestamp}{extension}"),
        )
        backups = sorted(
            (
                os.path.join(backup_dir, name)
                for name in os.listdir(backup_dir)
                if name.startswith(f"{stem}-") and name.endswith(extension)
            ),
            key=os.path.getmtime,
            reverse=True,
        )
        for old_backup in backups[10:]:
            os.remove(old_backup)

    def sync_current_sheet_to_supabase(self):
        if not self._workbook:
            QMessageBox.information(self, "Sync", "Open a workbook first.")
            return
        index = self.workbook_tabs.currentIndex()
        if index < 0:
            return
        sheet_name = self.workbook_tabs.tabText(index)
        if self._is_old_master_sheet(sheet_name):
            latest_master = self._latest_master_sheet_name()
            QMessageBox.warning(
                self,
                "Old master sheet",
                f"'{sheet_name}' looks like an older master list.\n\nUse '{latest_master}' for student sync, or export from Supabase instead."
            )
            return
        self._set_workbook_busy(True, f"Syncing {sheet_name}...")

        def sync():
            with self._workbook_lock:
                return self._sync_workbook_sheet_to_supabase(sheet_name)

        def synced(count):
            self._set_workbook_busy(False)
            if count is None:
                QMessageBox.information(self, "Sync", f"'{sheet_name}' is not a supported Supabase sync sheet.")
                return
            QMessageBox.information(self, "Sync complete", f"Synced {count} records from '{sheet_name}'.")
            self.workbook_status_label.setText(f"Synced {count} records from {sheet_name}.")
            self._audit("sync", "workbook_sheet", sheet_name, {"record_count": count})
            self._on_students_changed()

        def failed(error):
            self._set_workbook_busy(False)
            QMessageBox.critical(
                self,
                "Sync failed",
                f"Could not sync '{sheet_name}':\n{error.strip().splitlines()[-1]}",
            )

        return self._run_background(sync, synced, failed)

    def sync_all_workbook_sheets_to_supabase(self):
        if not self._workbook:
            QMessageBox.information(self, "Sync", "Open a workbook first.")
            return
        self._set_workbook_busy(True, "Syncing workbook sheets...")

        def sync_all():
            with self._workbook_lock:
                results = []
                latest_master = self._latest_master_sheet_name()
                for sheet_name in self._workbook.sheetnames:
                    if "master" in sheet_name.lower() and sheet_name != latest_master:
                        continue
                    count = self._sync_workbook_sheet_to_supabase(sheet_name)
                    if count is not None:
                        results.append(f"{sheet_name}: {count}")
            return results

        def synced(results):
            self._set_workbook_busy(False)
            if not results:
                QMessageBox.information(self, "Sync", "No supported workbook sheets were found.")
                return
            QMessageBox.information(self, "Sync complete", "Synced sheets:\n" + "\n".join(results))
            self.workbook_status_label.setText(f"Synced {len(results)} sheets to Supabase.")
            self._audit("sync_all", "workbook", details={"results": results})
            self._on_students_changed()

        def failed(error):
            self._set_workbook_busy(False)
            QMessageBox.critical(
                self,
                "Sync failed",
                f"Could not sync workbook:\n{error.strip().splitlines()[-1]}",
            )

        return self._run_background(sync_all, synced, failed)

    def _sync_workbook_sheet_to_supabase(self, sheet_name):
        return self.workbook_import_service.sync_sheet(self._workbook, sheet_name)

    def _is_old_master_sheet(self, sheet_name):
        latest_master = self._latest_master_sheet_name()
        return "master" in sheet_name.lower() and latest_master and sheet_name != latest_master

    def _sync_masterlist_sheet(self, sheet_name):
        return self.workbook_import_service.sync_masterlist_sheet(self._workbook, sheet_name)

    def _sync_coordinator_sheet(self, sheet_name):
        return self.workbook_import_service.sync_coordinator_sheet(self._workbook, sheet_name)

    def _sync_donor_sheet(self, sheet_name):
        return self.workbook_import_service.sync_donor_sheet(self._workbook, sheet_name)

    def _sync_movements_sheet(self, sheet_name):
        return self.workbook_import_service.sync_movements_sheet(self._workbook, sheet_name)

    def _insert_chunks(self, table_name, records, size=100):
        self.workbook_repository.insert_records(table_name, records, chunk_size=size)
    def _sheet_values(self, sheet_name):
        return self.workbook_import_service.sheet_values(self._workbook, sheet_name)

    def _find_header_row(self, rows, required):
        return self.workbook_import_service.find_header_row(rows, required)

    def _header_map(self, row):
        return self.workbook_import_service.header_map(row)

    def _normalize_header(self, value):
        return self.workbook_import_service.normalize_header(value)

    def _safe_cell(self, row, index):
        return self.workbook_import_service.safe_cell(row, index)

    def _school_year_from_sheet(self, sheet_name):
        return self.workbook_import_service.school_year_from_sheet(sheet_name)

    def _format_excel_value(self, value):
        return self.workbook_import_service.format_excel_value(value)

    @staticmethod
    def _excel_column_label(number):
        label = ""
        while number:
            number, remainder = divmod(number - 1, 26)
            label = chr(65 + remainder) + label
        return label

    # ── LIST ──────────────────────────────────────────────────────────────────
    def _latest_master_sheet_name(self):
        return self._latest_master_sheet_name_from_names(self._workbook.sheetnames)

    def _sheet_year_score(self, sheet_name):
        return self.masterlist_service.sheet_year_score(sheet_name)
    def _load_profile(self, sid):
        """Load a profile off the UI thread and ignore stale responses."""
        self._profile_request += 1
        request_id = self._profile_request
        self.lbl_profile_name.setText("Loading profile…")
        self.lbl_profile_status.setText("Fetching student record")
        self.profile_progress.setValue(0)
        self.profile_completion_label.setText("0%")
        self.profile_budget_label.setText("Loading budget…")
        self.profile_budget_bar.setValue(0)
        self.remarks_edit.setEnabled(False)
        self.save_remarks_btn.setEnabled(False)
        for button in (
            self.edit_btn,
            self.deactivate_btn,
            self.remove_student_btn,
            self.change_photo_btn,
            self.remove_photo_btn,
            self.header_profile_photo_button,
            self.header_profile_expenses_button,
        ):
            button.setEnabled(False)
        for action in (
            self.profile_change_photo_action,
            self.profile_remove_photo_action,
            self.profile_status_action,
            self.profile_remove_action,
        ):
            action.setEnabled(False)

        def fetch():
            student = self.student_repository.get_student_single(sid)
            student = self._apply_current_master_status(student)
            summary = self.expense_service.get_financial_summary(
                sid, self._current_school_year()
            )
            return student, summary

        def loaded(result):
            if request_id != self._profile_request or sid != self.current_student_id:
                return
            student, summary = result
            self._apply_profile(student, summary)
            self._loaded_remarks = self.remarks_edit.toPlainText()
            self.remarks_edit.setEnabled(True)
            self._update_remarks_dirty_state()
            self.edit_btn.setEnabled(True)
            self.remove_student_btn.setEnabled(True)
            self.change_photo_btn.setEnabled(True)
            self.header_profile_photo_button.setEnabled(True)
            self.header_profile_expenses_button.setEnabled(True)
            for action in (
                self.profile_change_photo_action,
                self.profile_remove_photo_action,
                self.profile_status_action,
                self.profile_remove_action,
            ):
                action.setEnabled(True)
            self.lbl_profile_name.setAccessibleDescription(
                f"Student profile for {self.lbl_profile_name.text()}"
            )
            self.status_bar.showMessage("Student profile updated", 2500)

        def failed(error):
            if request_id != self._profile_request or sid != self.current_student_id:
                return
            self.lbl_profile_name.setText("Profile unavailable")
            self.lbl_profile_status.setText("Could not load student record")
            self.remarks_edit.setEnabled(False)
            self.status_bar.showMessage(
                "Could not load this profile. Check the office connection and refresh.",
                8000,
            )
            logging.getLogger(__name__).error("Profile load failed:\n%s", error)

        return self._run_background(fetch, loaded, failed)

    def _apply_profile(self, s, financial_summary=None):
        try:
            full_status, short_status, _status_token = self._status_style(s.get("status"))
            inactive = (full_status != "Active")
            grade_label = self.student_service.format_grade_label(s.get("grade"))
            display_status = (
                "Graduating"
                if full_status == "Active" and grade_label == "Graduating"
                else full_status
            )
            
            # 1. Update Buttons and Status
            master_status = s.get("_status_source") == "masterlist"
            self.deactivate_btn.hide()
            self.profile_status_action.setEnabled(not master_status)
            self.deactivate_btn.setEnabled(True)
            self.deactivate_btn.setToolTip("")
            if inactive:
                self.deactivate_btn.setText("Mark active")
                self.profile_status_action.setText("Mark active")
                self.deactivate_btn.setObjectName("SuccessBtn")
                self.deactivate_btn.style().unpolish(self.deactivate_btn)
                self.deactivate_btn.style().polish(self.deactivate_btn)
            else:
                self.deactivate_btn.setText("Mark inactive")
                self.profile_status_action.setText("Mark inactive")
                self.deactivate_btn.setObjectName("WarningBtn")
                self.deactivate_btn.style().unpolish(self.deactivate_btn)
                self.deactivate_btn.style().polish(self.deactivate_btn)
            status_key = {
                "Active": "active",
                "Inactive/Removed": "inactive",
                "Graduated": "graduated",
                "Graduating": "graduated",
            }.get(display_status, "inactive")
            self.lbl_profile_status.setText(display_status)
            self.lbl_profile_status.setProperty("status", status_key)
            self.lbl_profile_status.setStyleSheet("")
            self.profile_summary_accent.setProperty("status", status_key)
            self.profile_summary_accent.style().unpolish(self.profile_summary_accent)
            self.profile_summary_accent.style().polish(self.profile_summary_accent)
            self.lbl_profile_status.style().unpolish(self.lbl_profile_status)
            self.lbl_profile_status.style().polish(self.lbl_profile_status)

            # 2. Update Header Name
            gender = " ".join(str(s.get("gender") or "").strip().split()).casefold()
            gender_suffix = (
                " (M)" if gender in {"m", "male"}
                else " (F)" if gender in {"f", "female"}
                else ""
            )
            display_name = ", ".join(
                part for part in (
                    str(s.get("last_name") or "").strip(),
                    str(s.get("first_name") or "").strip(),
                )
                if part
            )
            self.lbl_profile_name.setText(
                f"{display_name or 'Student name'}{gender_suffix}"
            )
            self.profile_meta_label.setText(
                f"{s.get('area') or 'Area not set'}  /  "
                f"{s.get('sponsor') or 'Sponsor not set'}"
            )

            # 3. Safely update all grid labels
            fields = [
                "gender", "grade", "area", "city", "address", 
                "birthday", "contact", "sponsor", "school", 
                "parents", "course"
            ]
            
            for field in fields:
                if field not in self.profile_data_labels:
                    continue
                val = s.get(field)
                # Ensure we handle None or empty strings gracefully
                display_text = (
                    str(val).strip()
                    if val and str(val).strip()
                    else "Not set"
                )
                self.profile_data_labels[field].setText(display_text)
                self.profile_data_labels[field].setAccessibleDescription(
                    f"{field.replace('_', ' ').title()}: {display_text}"
                )
            self.profile_data_labels["support_status"].setText(display_status)
            self.profile_data_labels["school_year"].setText(
                self._current_school_year()
            )

            self.remarks_edit.setPlainText(s.get("remarks") or "")
            self.remarks_display.setText(
                (s.get("remarks") or "").strip() or "No office notes recorded."
            )
            self.profile_next_action_copy.setText(
                "Assign a sponsor or record why one is not required."
                if not str(s.get("sponsor") or "").strip()
                else "Review during the next office follow-up."
            )

            photo_url = s.get("photo_url")
            self.header_profile_photo_button.setText(
                "Change photo" if photo_url else "Add photo"
            )
            self.header_profile_photo_button.setToolTip(
                "Replace the current student photo"
                if photo_url
                else "Upload a photo for this student"
            )
            self._load_photo_from_url(self.photo_label, photo_url)
            self.remove_photo_btn.hide()
            self.profile_remove_photo_action.setVisible(bool(photo_url))
            if not photo_url:
                initials = "".join(
                    part[:1].upper()
                    for part in (s.get("first_name"), s.get("last_name"))
                    if str(part or "").strip()
                )[:2]
                self.photo_label.setText(initials or "SSM")

            completion = self._profile_completion_percent(s)
            self.profile_completion_label.setText(f"{completion}%")
            animate_progress(
                self.profile_progress,
                completion,
                motion_enabled=not self._reduce_motion,
            )

            budget = self.expense_service.budget_card_status(financial_summary)
            self.profile_budget_label.setText(
                budget.get("detail") or budget.get("title") or "Unallocated this year"
            )
            self.profile_budget_bar.setProperty(
                "state", budget.get("state", "neutral")
            )
            self.profile_budget_bar.style().unpolish(self.profile_budget_bar)
            self.profile_budget_bar.style().polish(self.profile_budget_bar)
            animate_progress(
                self.profile_budget_bar,
                budget.get("percent", 0),
                motion_enabled=not self._reduce_motion,
            )

        except Exception as e:
            self.status_bar.showMessage(f"Load error ({type(e).__name__}): {e}", 8000)
    

    # ── PHOTO ─────────────────────────────────────────────────────────────────
    # Local photo cache: {url_without_query -> local_path}
    _photo_cache: dict = {}

    def _photo_cache_path(self, url: str) -> str:
        return self.photo_service.photo_cache_path(url)

    def _load_photo_from_url(self, label, url, log=None):
        """Show photo from local cache instantly; download in background if missing.

        A generation counter (_photo_gen) is stamped at call-time so that if
        the user navigates to another student before the download finishes, the
        stale callback detects the mismatch and discards its result instead of
        overwriting the newly-loaded profile photo.
        """
        def L(msg):
            if log is not None: log.append(msg)

        # Invalidate every older download before handling this request. This
        # must happen even when the new profile has no photo or uses the cache.
        self._photo_gen = getattr(self, "_photo_gen", 0) + 1
        my_gen = self._photo_gen
        if not url:
            label.clear(); label.setText("No photo"); return

        cache_path = self._photo_cache_path(url)

        # Serve from disk cache immediately if available — no network needed.
        if os.path.exists(cache_path):
            self._set_photo_local(label, cache_path)
            return

        def fetch():
            _path, data = self.photo_service.download_photo_to_cache(url)
            return data

        def apply(data):
            if getattr(self, "_photo_gen", 0) != my_gen:
                return
            pix = QPixmap()
            if pix.loadFromData(data) and not pix.isNull():
                w = label.width() or 140
                h = label.height() or 170
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                scaled = self._scale_cover(pix, w, h)
                if label.objectName() == "ProfileAvatar":
                    scaled = self._circle_crop(scaled, w, h)
                label.setPixmap(scaled)
            else:
                label.clear()
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                label.setText("No photo")

        def fail(error):
            L(f"Photo fetch failed: {error.strip().splitlines()[-1]}")
            if getattr(self, "_photo_gen", 0) == my_gen:
                label.clear()
                label.setText("No photo")

        self._run_background(fetch, apply, fail)

    @staticmethod
    def _scale_cover(pixmap: QPixmap, w: int, h: int) -> QPixmap:
        """Scale pixmap to fill w x h (cover), then centre-crop — no letterbox."""
        scaled = pixmap.scaled(w, h,
            Qt.AspectRatioMode.KeepAspectRatioByExpanding,
            Qt.TransformationMode.SmoothTransformation)
        if scaled.width() > w or scaled.height() > h:
            x = (scaled.width()  - w) // 2
            y = (scaled.height() - h) // 2
            scaled = scaled.copy(x, y, w, h)
        return scaled

    @staticmethod
    def _circle_crop(pixmap: QPixmap, w: int, h: int) -> QPixmap:
        """Clip profile photography to the circular Figma avatar shape."""
        result = QPixmap(w, h)
        result.fill(Qt.GlobalColor.transparent)
        painter = QPainter(result)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        clip = QPainterPath()
        clip.addEllipse(QRectF(0, 0, w, h))
        painter.setClipPath(clip)
        painter.drawPixmap(0, 0, pixmap)
        painter.end()
        return result

    def _set_photo_local(self, label, path):
        if path and os.path.exists(path):
            w = label.width() or 140
            h = label.height() or 170
            pix = self._scale_cover(QPixmap(path), w, h)
            if label.objectName() == "ProfileAvatar":
                pix = self._circle_crop(pix, w, h)
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label.setPixmap(pix)
        else:
            label.clear()
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label.setText("No photo")

    def _upload_photo(self, local_path, student_id, log=None):
        return self.photo_service.upload_photo(local_path, student_id, log)

    def _cache_uploaded_photo(self, source_path, url):
        try:
            return self.photo_service.cache_uploaded_photo(source_path, url)
        except Exception:
            return None

    def change_photo(self):
        if not self.current_student_id:
            QMessageBox.warning(self, "No student", "No student selected.")
            return
        path, _ = QFileDialog.getOpenFileName(self, "Select photo", "", "Images (*.png *.jpg *.jpeg *.bmp)")
        if not path:
            return
        sid = self.current_student_id
        self.status_bar.showMessage("Uploading photo... please wait")
        self.change_photo_btn.setEnabled(False)
        self.header_profile_photo_button.setEnabled(False)

        def upload():
            log = []
            logging.getLogger(__name__).info("Uploading photo for student %s", sid)
            url = self._upload_photo(path, sid, log)
            self.student_repository.update_photo_url(sid, url)
            self._cache_uploaded_photo(path, url)
            return url

        def done(_url):
            self.change_photo_btn.setEnabled(True)
            self.header_profile_photo_button.setEnabled(True)
            if self.current_student_id == sid:
                self._load_profile(sid)
            self.status_bar.showMessage("Photo updated", 5000)

        def failed(error):
            self.change_photo_btn.setEnabled(True)
            self.header_profile_photo_button.setEnabled(True)
            logging.getLogger(__name__).error("Photo upload failed:\n%s", error)
            self.status_bar.showMessage("Photo upload failed", 5000)
            QMessageBox.critical(self, "Photo upload failed", error.strip().splitlines()[-1])

        self._run_background(upload, done, failed)

    def remove_photo(self):
        """Delete the student's photo from Storage and clear photo_url in the DB."""
        if not self.current_student_id:
            return
        confirm = QMessageBox.question(
            self, "Remove photo",
            "Remove this student's photo? This cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        sid = self.current_student_id
        self.status_bar.showMessage("Removing photo…", 30000)
        self.remove_photo_btn.setEnabled(False)

        def remove():
            self.student_repository.update_photo_url(sid, None)
            self.photo_service.clear_student_cache(sid)
            self.photo_service.delete_storage_photo_variants(sid)

        def done(_result):
            self.remove_photo_btn.setEnabled(True)
            if self.current_student_id == sid:
                self._load_profile(sid)
            self.status_bar.showMessage("Photo removed", 4000)

        def failed(error):
            self.remove_photo_btn.setEnabled(True)
            self.status_bar.showMessage("Remove failed", 5000)
            QMessageBox.critical(self, "Remove photo failed", error.strip().splitlines()[-1])

        self._run_background(remove, done, failed)

    def pick_add_photo(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select photo", "", "Images (*.png *.jpg *.jpeg *.bmp)")
        if path:
            self._pending_photo = path
            self._set_photo_local(self.add_photo_label, path)

    # ── REMARKS ───────────────────────────────────────────────────────────────
    def _update_remarks_dirty_state(self) -> None:
        button = getattr(self, "save_remarks_btn", None)
        if button is None:
            return
        dirty = bool(self.current_student_id) and (
            self.remarks_edit.toPlainText() != getattr(self, "_loaded_remarks", "")
        )
        button.setEnabled(dirty and self.remarks_edit.isEnabled())
        button.setText("Save remarks")
        button.setAccessibleDescription(
            "Save changed student remarks" if dirty else "No unsaved remark changes"
        )

    def save_remarks(self):
        if not self.current_student_id:
            return
        student_id = self.current_student_id
        value = self.remarks_edit.toPlainText()
        self.save_remarks_btn.setEnabled(False)
        self.save_remarks_btn.setText("Saving…")

        def saved(_result):
            if student_id != self.current_student_id:
                return
            self._loaded_remarks = value
            self._update_remarks_dirty_state()
            self._audit("update_remarks", "student", student_id)
            self.status_bar.showMessage("Remarks saved", 3000)

        def failed(error):
            if student_id == self.current_student_id:
                self.save_remarks_btn.setText("Save remarks")
                self._update_remarks_dirty_state()
            self.status_bar.showMessage(
                "Could not save remarks. Check the office connection and try again.",
                8000,
            )
            logging.getLogger(__name__).error("Remark save failed:\n%s", error)

        return self._run_background(
            lambda: self.student_repository.update_student(
                student_id,
                {"remarks": value},
            ),
            saved,
            failed,
        )

    # ── TOGGLE STATUS ─────────────────────────────────────────────────────────
    def toggle_active_status(self):
        if not self.current_student_id:
            return
        student_id = self.current_student_id
        original_text = self.deactivate_btn.text()
        self.deactivate_btn.setEnabled(False)
        self.deactivate_btn.setText("Updating…")

        def update():
            student = self.student_repository.get_student_single(student_id)
            current = self._apply_current_master_status(student).get("status")
            new_status = "Active" if current == "Inactive/Removed" else "Inactive/Removed"
            self.student_repository.update_status(student_id, new_status)
            return new_status

        def updated(new_status):
            if student_id != self.current_student_id:
                return
            self._audit(
                "status_change", "student", student_id,
                {"status": new_status},
            )
            self._load_profile(student_id)
            self.status_bar.showMessage(f"Status set to {new_status}", 3000)
            self._on_students_changed()

        def failed(error):
            if student_id == self.current_student_id:
                self.deactivate_btn.setEnabled(True)
                self.deactivate_btn.setText(original_text)
            self.status_bar.showMessage(
                "Could not update the student status. Check the office connection and try again.",
                8000,
            )
            logging.getLogger(__name__).error(
                "Student status update failed:\n%s", error
            )

        return self._run_background(update, updated, failed)

    # ── ADD / EDIT FORM ───────────────────────────────────────────────────────
    def remove_current_student(self):
        if not self.current_student_id:
            return
        displayed_name = self.lbl_profile_name.text().strip()
        name = (
            displayed_name
            if displayed_name not in {"", "Loading profile…", "Profile unavailable"}
            else "this student"
        )
        confirm = QMessageBox.question(
            self,
            "Remove student",
            (
                f"Remove {name}?\n\n"
                "This will permanently delete the student record and related "
                "expenses, budgets, donor links, and movement entries."
            ),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        sure = QMessageBox.question(
            self,
            "Remove this student?",
            (
                f"Are you sure you want to permanently remove {name}?\n\n"
                "This deletes the live Supabase data for everyone using the system."
            ),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if sure != QMessageBox.StandardButton.Yes:
            return

        student_id = self.current_student_id

        def remove():
            return self.student_repository.delete_student_with_related(student_id)

        def removed(result):
            deleted = result.get("students", [])
            if not deleted:
                self.status_bar.showMessage("Student was not found or was already removed.", 5000)
            else:
                self.status_bar.showMessage(f"Removed {name}.", 5000)
            self._audit("delete", "student", student_id, details={"name": name})
            self.current_student_id = None
            self._on_students_changed()
            self.nav_students()

        def failed(error):
            self.status_bar.showMessage(
                "Could not remove the student. Check the office connection and try again.",
                8000,
            )
            logging.getLogger(__name__).error(
                "Student removal failed:\n%s", error
            )

        self._run_background(remove, removed, failed)

    def _open_add_screen(self):
        self._clear_form()
        self._switch_page(3)

    def open_edit_screen(self):
        if not self.current_student_id:
            return
        student_id = self.current_student_id
        self._student_form_request += 1
        request_id = self._student_form_request
        self._editing_id = student_id
        self.form_title_label.setText("Loading student record…")
        self.save_form_btn.setEnabled(False)
        self._switch_page(3)
        self.header_add_button.setEnabled(False)

        def fetch():
            student = self.student_repository.get_student_single(student_id)
            return student, self._apply_current_master_status(student)

        def loaded(result):
            if (
                request_id != self._student_form_request
                or student_id != self.current_student_id
                or self.stacked_widget.currentIndex() != 3
            ):
                return
            student, resolved = result
            self._apply_student_form(student, resolved)
            self.save_form_btn.setEnabled(True)
            self.header_add_button.setEnabled(True)

        def failed(error):
            if request_id != self._student_form_request:
                return
            self.save_form_btn.setEnabled(True)
            self.header_add_button.setEnabled(True)
            self.status_bar.showMessage(
                "Could not load this student for editing. Refresh the profile and try again.",
                8000,
            )
            logging.getLogger(__name__).error("Edit form load failed:\n%s", error)

        return self._run_background(fetch, loaded, failed)

    def _apply_student_form(self, student, s):
        try:
            self.form_title_label.setText("Edit student record")
            self._editing_id = self.current_student_id
            self._editing_snapshot = {
                field: student.get(field)
                for field in (
                    "last_name", "first_name", "gender", "grade", "address",
                    "city", "area", "birthday", "sponsor", "contact", "school",
                    "parents", "course", "remarks", "status",
                )
            }
            self.inp_last.setText(s.get("last_name") or "")
            self.inp_first.setText(s.get("first_name") or "")
            self.inp_gender.setCurrentIndex(max(self.inp_gender.findText(s.get("gender") or ""), 0))
            self.inp_grade.setText(s.get("grade") or "")
            self.inp_address.setText(s.get("address") or "")
            self.inp_city.setText(s.get("city") or "")
            self.inp_area.setText(s.get("area") or "")
            birthday_str = s.get("birthday") or ""
            if birthday_str:
                self.inp_bday.setDate(QDate.fromString(birthday_str, "yyyy-MM-dd"))
            else:
                self.inp_bday.setDate(self.inp_bday.minimumDate())
            self.inp_sponsor.setText(s.get("sponsor") or "")
            self.inp_contact.setText(s.get("contact") or "")
            self.inp_school.setText(s.get("school") or "")
            self.inp_parents.setText(s.get("parents") or "")
            self.inp_course.setText(s.get("course") or "")
            self.inp_remarks.setPlainText(s.get("remarks") or "")
            self.inp_status.setCurrentIndex(max(self.inp_status.findText(s.get("status") or "Active"), 0))
            master_status = s.get("_status_source") == "masterlist"
            self.inp_status.setEnabled(not master_status)
            self.inp_status.setToolTip("Edit the latest masterlist workbook to change this status." if master_status else "")
            self._pending_photo = None
            self.add_photo_label.clear(); self.add_photo_label.setText("No photo")
            self._switch_page(3)
            self.page_title_label.setText("Edit student")
            self.page_subtitle_label.setText(
                "Update identity, support, and office record details"
            )
        except Exception as e:
            self.status_bar.showMessage(f"Load error ({type(e).__name__}): {e}", 8000)

    def _clear_form(self):
        for w in (self.inp_last, self.inp_first, self.inp_grade, self.inp_address,
                  self.inp_city, self.inp_area, self.inp_sponsor,
                  self.inp_contact, self.inp_school, self.inp_parents, self.inp_course):
            w.clear()
            if w in (self.inp_last, self.inp_first):
                w.setProperty("invalid", False)
                w.style().unpolish(w)
                w.style().polish(w)
        self.inp_bday.setDate(self.inp_bday.minimumDate())
        self.inp_remarks.clear()
        self.inp_gender.setCurrentIndex(0)
        self.inp_status.setCurrentIndex(0)
        self.inp_status.setEnabled(True)
        self.inp_status.setToolTip("")
        self.add_photo_label.clear(); self.add_photo_label.setText("No photo")
        self._pending_photo = None
        self._editing_id = None
        self._editing_snapshot = None
        self.form_title_label.setText("New student record")

    def save_student_form(self):
        required_fields = (
            (self.inp_last, "Enter the student's family name."),
            (self.inp_first, "Enter the student's given name."),
        )
        missing_field = None
        missing_message = ""
        for field, message in required_fields:
            invalid = not field.text().strip()
            field.setProperty("invalid", invalid)
            field.style().unpolish(field)
            field.style().polish(field)
            if invalid and missing_field is None:
                missing_field = field
                missing_message = message
        if missing_field is not None:
            missing_field.setFocus()
            self.status_bar.showMessage(missing_message, 6000)
            return

        bday_text = (
            self.inp_bday.date().toString("yyyy-MM-dd")
            if self.inp_bday.date() != self.inp_bday.minimumDate()
            else None
        )

        payload = {
            "last_name":  self.inp_last.text().strip(),
            "first_name": self.inp_first.text().strip(),
            "gender":     self.inp_gender.currentText(),
            "grade":      self.inp_grade.text(),
            "address":    self.inp_address.text(),
            "city":       self.inp_city.text(),
            "area":       self.inp_area.text(),
            "birthday":   bday_text,
            "sponsor":    self.inp_sponsor.text(),
            "contact":    self.inp_contact.text(),
            "school":     self.inp_school.text(),
            "parents":    self.inp_parents.text(),
            "course":     self.inp_course.text(),
            "remarks":    self.inp_remarks.toPlainText(),
            "status":     self.inp_status.currentText(),
        }
        editing = self._editing_id
        editing_snapshot = dict(self._editing_snapshot or {})
        pending_photo = self._pending_photo
        self.save_form_btn.setEnabled(False)
        self.header_add_button.setEnabled(False)
        self.status_bar.showMessage("Saving student...", 30000)

        def save():
            if editing:
                current = self.student_repository.get_student_single(
                    editing,
                    columns=",".join(editing_snapshot),
                )
                changed_elsewhere = any(
                    str(current.get(field) or "").strip()
                    != str(original or "").strip()
                    for field, original in editing_snapshot.items()
                )
                if changed_elsewhere:
                    raise RuntimeError(
                        "This student was changed on another computer. "
                        "Reopen the profile and apply your changes again."
                    )
                self.student_repository.update_student(editing, payload)
                student_id = editing
            else:
                rows = self.student_repository.insert_student(payload)
                if not rows or "id" not in rows[0]:
                    raise RuntimeError("The database did not return the new student ID.")
                student_id = rows[0]["id"]

            if pending_photo:
                url = self._upload_photo(pending_photo, student_id)
                self.student_repository.update_photo_url(student_id, url)
                self._cache_uploaded_photo(pending_photo, url)
            return student_id, bool(editing)

        def saved(result):
            student_id, was_editing = result
            self.save_form_btn.setEnabled(True)
            self.header_add_button.setEnabled(True)
            self._clear_form()
            self._audit("update" if was_editing else "create", "student", student_id)
            self._on_students_changed()
            if was_editing:
                self.current_student_id = student_id
                self._load_profile(student_id)
                self._switch_page(2)
            else:
                self.nav_students()
            self.status_bar.showMessage("Student saved", 4000)

        def failed(error):
            self.save_form_btn.setEnabled(True)
            self.header_add_button.setEnabled(True)
            logging.getLogger(__name__).error("Student save failed:\n%s", error)
            self.status_bar.showMessage("Student save failed", 5000)
            QMessageBox.critical(self, "Save error", error.strip().splitlines()[-1])

        self._run_background(save, saved, failed)

    # ── EXPENSES ──────────────────────────────────────────────────────────────
    def open_expenses_screen(self):
        if not self.current_student_id:
            return
        student_id = self.current_student_id
        self.expenses_title.setText("Loading student…")
        self.expenses_title.setAccessibleDescription(
            "Loading the selected student's expense records"
        )
        self.expenses_table.setRowCount(0)
        self._switch_page(4)
        self._refresh_expenses_view()

        def loaded(student):
            if student_id != self.current_student_id:
                return
            display_name = (
                f"{student.get('last_name', '')}, {student.get('first_name', '')}"
            ).strip(", ")
            self.expenses_title.setText(display_name or "Student")
            self.expenses_title.setAccessibleDescription(
                f"Expense records for {display_name or 'the selected student'}"
            )

        def failed(error):
            if student_id != self.current_student_id:
                return
            self.expenses_title.setText("Expenses")
            logging.getLogger(__name__).error(
                "Expense student heading failed:\n%s", error
            )

        return self._run_background(
            lambda: self.student_repository.get_student_single(
                student_id,
                columns="last_name,first_name",
            ),
            loaded,
            failed,
        )

    def _on_sy_changed(self):
        """Called when school year filter changes; reload expenses and budget."""
        sy = self.exp_school_year.currentText()
        if sy != ExpenseService.ALL_YEARS:
            idx = self.exp_sy_entry.findText(sy)
            if idx >= 0:
                self.exp_sy_entry.setCurrentIndex(idx)
        self._refresh_expenses_view()

    def _refresh_expenses_view(self):
        if not self.current_student_id:
            return

        self._expense_request += 1
        request_id = self._expense_request
        student_id = self.current_student_id
        sy_filter = self.exp_school_year.currentText()
        budget_year = (
            self._current_school_year()
            if sy_filter == ExpenseService.ALL_YEARS
            else sy_filter
        )
        self.expense_budget_heading.setText(f"Budget — {budget_year}")
        self.expenses_table.setRowCount(0)
        self.expenses_empty_state.title_label.setText("Updating expense history…")
        self.expenses_empty_state.description_label.setText(
            "Retrieving synchronized budget and expense figures."
        )
        self.expenses_empty_state.setAccessibleName("Updating expense history")
        self.expense_history_stack.setCurrentWidget(self.expenses_empty_state)
        self.total_label.setText("Total: Updating…")
        self.budget_status_lbl.setText("Updating synchronized figures…")
        self.budget_input.setEnabled(False)
        self.save_budget_btn.setEnabled(False)
        self.add_expense_btn.setEnabled(False)

        def do_work():
            expenses = self.expense_service.list_expenses(student_id, sy_filter)
            summary = self.expense_service.get_financial_summary(
                student_id,
                budget_year,
            )
            budget_expenses = (
                self.expense_service.expenses_for_school_year(
                    expenses,
                    budget_year,
                )
                if sy_filter == ExpenseService.ALL_YEARS
                else expenses
            )
            return (
                self.expense_service.reconcile_summary(
                    summary,
                    budget_expenses,
                    budget_year,
                ),
                expenses,
                budget_year,
            )

        def on_done(result):
            if (
                request_id != self._expense_request
                or student_id != self.current_student_id
                or sy_filter != self.exp_school_year.currentText()
            ):
                return
            summary_data, expenses_data, budget_year = result

            # Update budget display
            budget_info = self.expense_service.budget_usage(summary_data)
            self.expense_budget_heading.setText(f"Budget — {budget_year}")
            animate_progress(
                self.budget_bar,
                budget_info["percent"],
                motion_enabled=not self._reduce_motion,
            )
            self.budget_bar.setProperty("state", budget_info["state"])
            self.budget_bar.style().unpolish(self.budget_bar)
            self.budget_bar.style().polish(self.budget_bar)
            self.budget_status_lbl.setText(budget_info["message"])
            self.budget_status_lbl.setProperty("state", budget_info["state"])
            self.budget_status_lbl.style().unpolish(self.budget_status_lbl)
            self.budget_status_lbl.style().polish(self.budget_status_lbl)
            budget_amount = budget_info.get("budget", 0)
            self.expense_budget_amount_display.setText(
                f"PHP {budget_amount:,.2f}"
            )
            self.expense_spent_display.setText(
                f"PHP {budget_info.get('spent', 0):,.2f} spent"
            )
            self.expense_remaining_display.setText(
                f"PHP {budget_info.get('remaining', 0):,.2f} remaining"
            )
            self.budget_input.setText(f"{budget_amount:,.2f}" if budget_amount > 0 else "")
            can_edit_budget = sy_filter != ExpenseService.ALL_YEARS
            self.budget_input.setEnabled(can_edit_budget)
            self.save_budget_btn.setEnabled(can_edit_budget)
            self.save_budget_btn.setText("Save budget")
            self.add_expense_btn.setEnabled(True)
            self.add_expense_btn.setText("Add expense")
            self.budget_bar.setAccessibleDescription(
                f"{budget_info['percent']} percent used. {budget_info['message']}"
            )

            # Populate expenses table
            for exp in expenses_data:
                self._add_expense_to_table(exp)
            self.expenses_empty_state.title_label.setText("No expenses recorded")
            self.expenses_empty_state.description_label.setText(
                (
                    "No expenses have been recorded for this student."
                    if sy_filter == ExpenseService.ALL_YEARS
                    else "Add the first expense for this school year using the form above."
                )
            )
            self.expenses_empty_state.setAccessibleName("No expenses recorded")
            self.expense_history_stack.setCurrentWidget(
                self.expenses_table if expenses_data else self.expenses_empty_state
            )

            # Update total label
            total = self.expense_service.calculate_total(expenses_data)
            self.total_label.setText(self.expense_service.total_label(total, sy_filter))
            self.total_label.setAccessibleDescription(
                f"Synchronized total is PHP {total:,.2f} for {sy_filter}"
            )
            self._mark_database_updated()

        def on_error(error):
            if (
                request_id != self._expense_request
                or student_id != self.current_student_id
                or sy_filter != self.exp_school_year.currentText()
            ):
                return
            can_edit_budget = sy_filter != ExpenseService.ALL_YEARS
            self.budget_input.setEnabled(can_edit_budget)
            self.save_budget_btn.setEnabled(can_edit_budget)
            self.add_expense_btn.setEnabled(True)
            self.save_budget_btn.setText("Save budget")
            self.add_expense_btn.setText("Add expense")
            self.total_label.setText("Total unavailable")
            self.budget_status_lbl.setText("Budget figures unavailable")
            self.expenses_empty_state.title_label.setText("Expense history unavailable")
            self.expenses_empty_state.description_label.setText(
                "Check the office database connection, then refresh this view."
            )
            self.expenses_empty_state.setAccessibleName("Expense history unavailable")
            self.expense_history_stack.setCurrentWidget(self.expenses_empty_state)
            self.status_bar.showMessage(
                "Could not load expenses. Check the office connection and refresh.",
                8000,
            )
            logging.getLogger(__name__).error(
                "Expense refresh failed:\n%s", error
            )

        return self._run_background(do_work, on_done, on_error)

    def _add_expense_to_table(self, exp):
        row_idx = self.expenses_table.rowCount()
        self.expenses_table.insertRow(row_idx)
        self.expenses_table.setRowHeight(row_idx, 40)
        description_item = QTableWidgetItem(exp.get("description", ""))
        description_item.setToolTip(description_item.text())
        self.expenses_table.setItem(row_idx, 0, description_item)
        amount = exp.get("amount", 0) or 0
        amount_item = QTableWidgetItem(f"{amount:,.2f}")
        amount_item.setTextAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )
        amount_item.setToolTip(f"PHP {amount:,.2f}")
        self.expenses_table.setItem(row_idx, 1, amount_item)
        self.expenses_table.setItem(row_idx, 2, QTableWidgetItem(exp.get("date") or ""))
        self.expenses_table.setItem(row_idx, 3, QTableWidgetItem(exp.get("school_year") or ""))
        del_btn = ActionButton("Delete", variant="danger")
        del_btn.setProperty("density", "compact")
        expense_id = exp["id"]
        del_btn.clicked.connect(lambda _, e=expense_id: self.delete_expense(e))
        action_cell = QWidget()
        action_cell.setObjectName("TableActionCell")
        action_layout = QHBoxLayout(action_cell)
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.setSpacing(0)
        action_layout.addWidget(del_btn)
        self.expenses_table.setCellWidget(row_idx, 4, action_cell)

    def save_budget(self):
        """Upsert budget for current student + school year."""
        sy = self.exp_school_year.currentText()
        if sy == ExpenseService.ALL_YEARS:
            QMessageBox.warning(self, "Budget", "Please select a specific school year before saving a budget.")
            return
        try:
            amount = self.expense_service.parse_amount(self.budget_input.text())
        except ValueError:
            self.status_bar.showMessage("Invalid budget amount", 4000)
            return
        student_id = self.current_student_id
        self.save_budget_btn.setEnabled(False)
        self.save_budget_btn.setText("Saving…")

        def saved(_result):
            if student_id != self.current_student_id:
                return
            self._audit(
                "save_budget", "student", student_id,
                {"school_year": sy, "amount": amount},
            )
            self.status_bar.showMessage(f"Budget saved: PHP {amount:,.2f} for {sy}", 4000)
            self._refresh_expenses_view()

        def failed(error):
            if student_id == self.current_student_id:
                self.save_budget_btn.setEnabled(True)
                self.save_budget_btn.setText("Save budget")
            self.status_bar.showMessage(
                "Could not save the budget. Check the office connection and try again.",
                8000,
            )
            logging.getLogger(__name__).error("Budget save failed:\n%s", error)

        return self._run_background(
            lambda: self.expense_service.save_budget(student_id, sy, amount),
            saved,
            failed,
        )

    def add_expense(self):
        if not self.current_student_id:
            QMessageBox.information(self, "Select student", "Select a student before adding an expense.")
            return
        desc = self.exp_desc.text().strip()
        if not desc:
            self.status_bar.showMessage("Description is required", 4000)
            return
        try:
            amount = self.expense_service.parse_amount(self.exp_amount.text())
        except ValueError:
            self.status_bar.showMessage("Invalid amount - enter a number like 250.00", 4000)
            return
        school_year = self.exp_sy_entry.currentText()
        expense_date = self.exp_date.date().toString("yyyy-MM-dd")
        student_id = self.current_student_id
        self.add_expense_btn.setEnabled(False)
        self.add_expense_btn.setText("Adding…")

        def added(_result):
            if student_id != self.current_student_id:
                return
            self._audit(
                "add_expense", "student", student_id,
                {"school_year": school_year, "amount": amount, "description": desc},
            )
            self.exp_desc.clear()
            self.exp_amount.clear()
            self.exp_date.setDate(QDate.currentDate())
            if self.exp_school_year.currentText() not in (ExpenseService.ALL_YEARS, school_year):
                self.exp_school_year.setCurrentText(school_year)
            else:
                self._refresh_expenses_view()
            self.status_bar.showMessage("Expense added", 4000)

        def failed(error):
            if student_id == self.current_student_id:
                self.add_expense_btn.setEnabled(True)
                self.add_expense_btn.setText("Add expense")
            self.status_bar.showMessage(
                "Could not add the expense. Check the details and office connection.",
                8000,
            )
            logging.getLogger(__name__).error("Expense add failed:\n%s", error)

        return self._run_background(
            lambda: self.expense_service.add_expense(
                student_id,
                desc,
                amount,
                expense_date,
                school_year,
            ),
            added,
            failed,
        )

    def delete_expense(self, eid):
        confirm = QMessageBox.question(
            self,
            "Delete expense",
            "Delete this expense from the student's history?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return
        student_id = self.current_student_id
        self.expenses_table.setEnabled(False)

        def deleted(_result):
            self._audit("delete", "expense", eid)
            self._refresh_expenses_view()
            self.expenses_table.setEnabled(True)
            self.status_bar.showMessage("Expense deleted", 4000)

        def failed(error):
            if student_id == self.current_student_id:
                self.expenses_table.setEnabled(True)
            self.status_bar.showMessage(
                "Could not delete the expense. Check the office connection and try again.",
                8000,
            )
            logging.getLogger(__name__).error("Expense delete failed:\n%s", error)

        return self._run_background(
            lambda: self.expense_service.delete_expense(eid),
            deleted,
            failed,
        )

    def closeEvent(self, event):
        if self._workbook_dirty:
            decision = QMessageBox.question(
                self,
                "Unsaved workbook changes",
                "Save workbook changes before closing the application?",
                QMessageBox.StandardButton.Yes |
                QMessageBox.StandardButton.No |
                QMessageBox.StandardButton.Cancel,
            )
            if decision == QMessageBox.StandardButton.Cancel:
                event.ignore()
                return
            if decision == QMessageBox.StandardButton.Yes and not self.save_workbook_tabs():
                event.ignore()
                return
        if self._workbook and hasattr(self._workbook, "close"):
            self._workbook.close()
        event.accept()

# ── ENTRY POINT ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    register_app_fonts()
    icon_path = resource_path(APP_ICON_ASSET)
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))

    # 1. Create the client, guiding normal users through one-time setup when needed.
    while True:
        try:
            sb = get_supabase()
            break
        except RuntimeError as error:
            setup = ConfigurationDialog(str(error))
            if setup.exec() != QDialog.DialogCode.Accepted:
                raise SystemExit(0) from error
            reset_supabase_clients()

    # 2. Choose an office profile and recover in-place from connection errors.
    while True:
        splash = StartupDialog()
        splash.queue_ping(sb)
        result = splash.exec()
        if result == QDialog.DialogCode.Accepted and splash.success:
            break
        if splash.recovery_action != "settings":
            raise SystemExit(0)

        setup = ConfigurationDialog(friendly_connection_error(splash.error_msg))
        if setup.exec() != QDialog.DialogCode.Accepted:
            raise SystemExit(0)
        reset_supabase_clients()
        sb = get_supabase()

    # 3. Launch the shared office workspace under the selected audit identity.
    window = StudentApp(sb, initial_user=splash.selected_user)
    window.show()
    sys.exit(app.exec())
